# -*- coding: utf-8 -*-
"""
fason_db.py — BBA Fason İrsaliye Takip Modülü
Firma yönetimi dahil versiyon.
"""

import os
import sys
import sqlite3
from datetime import datetime
from flask import Blueprint, request, jsonify, session

fason_bp = Blueprint("fason", __name__)

ORTAK_KLASOR = r"K:\Warehouse\Yeşilovacık\12_Paylaşım Klasörü\01-BBA\bba-tool"


def _db_klasor_bul():
    if os.path.isdir(ORTAK_KLASOR):
        return ORTAK_KLASOR
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


DB_KLASOR = _db_klasor_bul()
DB_YOL = os.path.join(DB_KLASOR, "fason.db")


def _fason_export_klasor():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_db():
    conn = sqlite3.connect(DB_YOL)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _kolon_var_mi(conn, tablo, kolon):
    r = conn.execute(f"PRAGMA table_info({tablo})").fetchall()
    return any(row[1] == kolon for row in r)


def _migrate_ek_kolonlar(conn):
    yeni_kolonlar = [
        ("irsaliye_tarihi", "TEXT"),
        ("stok_miktari",    "REAL"),
        ("giris_miktari",   "REAL"),
        ("otis_stok",       "REAL"),
        ("otis_giris",      "REAL"),
        ("toplam_fiyat",    "REAL"),
        ("firma_id",        "INTEGER"),
    ]
    for kolon_ad, tip in yeni_kolonlar:
        if not _kolon_var_mi(conn, "fason_irsaliye", kolon_ad):
            try:
                conn.execute(f"ALTER TABLE fason_irsaliye ADD COLUMN {kolon_ad} {tip}")
                print(f"[Fason DB] Kolon eklendi: {kolon_ad}")
            except Exception as e:
                print(f"[Fason DB] Kolon ekleme hatasi ({kolon_ad}): {e}")


def init_fason_db():
    conn = get_db()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS fason_durum (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ad TEXT NOT NULL UNIQUE,
                renk TEXT DEFAULT 'text3',
                ikon TEXT DEFAULT 'fa-circle',
                sira INTEGER DEFAULT 0,
                aktif INTEGER DEFAULT 1,
                olusturulma TEXT DEFAULT (datetime('now', 'localtime'))
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS fason_firma (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ad TEXT NOT NULL UNIQUE,
                aktif INTEGER DEFAULT 1,
                olusturulma TEXT DEFAULT (datetime('now', 'localtime'))
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS fason_irsaliye (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                irsaliye_no TEXT NOT NULL,
                aciklama TEXT DEFAULT '',
                giren_kullanici TEXT NOT NULL,
                girilme_tarihi TEXT DEFAULT (datetime('now', 'localtime')),
                durum_id INTEGER,
                durum_notu TEXT DEFAULT '',
                durum_guncelleyen TEXT DEFAULT '',
                durum_guncelleme_tarihi TEXT DEFAULT '',
                FOREIGN KEY (durum_id) REFERENCES fason_durum(id)
            )
        """)

        _migrate_ek_kolonlar(conn)

        conn.execute("CREATE INDEX IF NOT EXISTS idx_irsaliye_no ON fason_irsaliye(irsaliye_no)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_durum ON fason_irsaliye(durum_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_firma ON fason_irsaliye(firma_id)")

        c = conn.execute("SELECT COUNT(*) FROM fason_durum").fetchone()
        if c[0] == 0:
            varsayilan_durumlar = [
                ("Fatura kaydı bekleniyor", "amber",  "fa-hourglass-half", 10),
                ("Çıkış yok",               "red",    "fa-ban",             20),
                ("Kısmi çıkış",             "purple", "fa-clock-rotate-left", 30),
                ("Çıkış yapıldı",           "green",  "fa-check",           40),
            ]
            conn.executemany(
                "INSERT INTO fason_durum (ad, renk, ikon, sira) VALUES (?, ?, ?, ?)",
                varsayilan_durumlar
            )
            print(f"[Fason DB] Varsayilan 4 durum eklendi")

        conn.commit()
        print(f"[Fason DB] Tablolar hazir: {DB_YOL}")
    except Exception as e:
        print(f"[Fason DB] Init hatasi: {e}")
    finally:
        conn.close()


# ═════════════════════════════════════════════════
# FİRMA ENDPOINT'LERİ
# ═════════════════════════════════════════════════

@fason_bp.route("/api/fason/firmalar", methods=["GET"])
def api_fason_firmalar():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    tumu = request.args.get("tumu") == "1"
    conn = get_db()
    try:
        if tumu:
            rows = conn.execute("SELECT * FROM fason_firma ORDER BY ad").fetchall()
        else:
            rows = conn.execute("SELECT * FROM fason_firma WHERE aktif = 1 ORDER BY ad").fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


@fason_bp.route("/api/fason/firma-ekle", methods=["POST"])
def api_fason_firma_ekle():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    try:
        d = request.get_json() or {}
        ad = (d.get("ad") or "").strip()
        if not ad:
            return jsonify({"durum": "hata", "mesaj": "Firma adı zorunlu"}), 400
        if len(ad) > 100:
            return jsonify({"durum": "hata", "mesaj": "Firma adı çok uzun"}), 400
        conn = get_db()
        try:
            var = conn.execute("SELECT id FROM fason_firma WHERE LOWER(ad) = LOWER(?)", (ad,)).fetchone()
            if var:
                return jsonify({"durum": "hata", "mesaj": "Bu firma zaten kayıtlı", "id": var["id"]}), 400
            cur = conn.execute("INSERT INTO fason_firma (ad) VALUES (?)", (ad,))
            conn.commit()
            return jsonify({"durum": "ok", "id": cur.lastrowid, "ad": ad, "mesaj": f"{ad} eklendi"})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


@fason_bp.route("/api/fason/firma-sil/<int:fid>", methods=["DELETE"])
def api_fason_firma_sil(fid):
    if session.get("rol") != "admin" and session.get("kullanici") != "admin":
        return jsonify({"durum": "hata", "mesaj": "Sadece admin"}), 403
    try:
        conn = get_db()
        try:
            c = conn.execute("SELECT COUNT(*) FROM fason_irsaliye WHERE firma_id = ?", (fid,)).fetchone()[0]
            if c > 0:
                conn.execute("UPDATE fason_firma SET aktif = 0 WHERE id = ?", (fid,))
                conn.commit()
                return jsonify({"durum": "ok", "mesaj": f"Pasifleştirildi ({c} irsaliyede kullanılıyor)"})
            else:
                conn.execute("DELETE FROM fason_firma WHERE id = ?", (fid,))
                conn.commit()
                return jsonify({"durum": "ok", "mesaj": "Silindi"})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


# ═════════════════════════════════════════════════
# İRSALİYE ENDPOINT'LERİ
# ═════════════════════════════════════════════════

@fason_bp.route("/api/fason/liste", methods=["GET"])
def api_fason_liste():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT
                i.id, i.irsaliye_no, i.aciklama,
                i.giren_kullanici, i.girilme_tarihi,
                i.durum_id, i.durum_notu,
                i.durum_guncelleyen, i.durum_guncelleme_tarihi,
                i.irsaliye_tarihi, i.stok_miktari, i.giris_miktari,
                i.otis_stok, i.otis_giris, i.toplam_fiyat,
                i.firma_id,
                d.ad AS durum_ad, d.renk AS durum_renk, d.ikon AS durum_ikon,
                f.ad AS firma_ad
            FROM fason_irsaliye i
            LEFT JOIN fason_durum d ON i.durum_id = d.id
            LEFT JOIN fason_firma f ON i.firma_id = f.id
            ORDER BY i.id DESC
        """).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


@fason_bp.route("/api/fason/ekle", methods=["POST"])
def api_fason_ekle():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    try:
        d = request.get_json() or {}
        irsaliye_no = (d.get("irsaliye_no") or "").strip()
        firma_id = d.get("firma_id")
        aciklama = (d.get("aciklama") or "").strip()

        if not irsaliye_no:
            return jsonify({"durum": "hata", "mesaj": "İrsaliye No zorunlu"}), 400
        if len(irsaliye_no) > 50:
            return jsonify({"durum": "hata", "mesaj": "İrsaliye No çok uzun"}), 400
        if not firma_id:
            return jsonify({"durum": "hata", "mesaj": "Firma seçiniz"}), 400
        try:
            firma_id = int(firma_id)
        except:
            return jsonify({"durum": "hata", "mesaj": "Geçersiz firma"}), 400

        conn = get_db()
        try:
            fv = conn.execute("SELECT id FROM fason_firma WHERE id = ?", (firma_id,)).fetchone()
            if not fv:
                return jsonify({"durum": "hata", "mesaj": "Firma bulunamadı"}), 400

            mevcut = conn.execute(
                "SELECT id FROM fason_irsaliye WHERE irsaliye_no = ?",
                (irsaliye_no,)
            ).fetchone()

            cur = conn.execute("""
                INSERT INTO fason_irsaliye
                    (irsaliye_no, aciklama, giren_kullanici, firma_id)
                VALUES (?, ?, ?, ?)
            """, (irsaliye_no, aciklama, session.get("kullanici", "-"), firma_id))
            conn.commit()

            return jsonify({
                "durum": "ok",
                "id": cur.lastrowid,
                "mesaj": f"İrsaliye eklendi: {irsaliye_no}",
                "uyari": ("Bu irsaliye numarası daha önce girilmiş" if mevcut else None)
            })
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


@fason_bp.route("/api/fason/durum-guncelle/<int:irs_id>", methods=["POST"])
def api_fason_durum_guncelle(irs_id):
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    try:
        d = request.get_json() or {}
        durum_id = d.get("durum_id")
        durum_notu = (d.get("durum_notu") or "").strip()

        if durum_id is not None and durum_id != "":
            try:
                durum_id = int(durum_id)
            except:
                return jsonify({"durum": "hata", "mesaj": "Geçersiz durum"}), 400
        else:
            durum_id = None

        conn = get_db()
        try:
            simdi = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute("""
                UPDATE fason_irsaliye
                SET durum_id = ?, durum_notu = ?,
                    durum_guncelleyen = ?, durum_guncelleme_tarihi = ?
                WHERE id = ?
            """, (durum_id, durum_notu, session.get("kullanici", "-"), simdi, irs_id))
            conn.commit()

            if conn.total_changes == 0:
                return jsonify({"durum": "hata", "mesaj": "İrsaliye bulunamadı"}), 404
            return jsonify({"durum": "ok", "mesaj": "Durum güncellendi"})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


@fason_bp.route("/api/fason/sil/<int:irs_id>", methods=["DELETE"])
def api_fason_sil(irs_id):
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    try:
        conn = get_db()
        try:
            rec = conn.execute("SELECT giren_kullanici FROM fason_irsaliye WHERE id = ?", (irs_id,)).fetchone()
            if not rec:
                return jsonify({"durum": "hata", "mesaj": "Bulunamadı"}), 404
            kullanici = session.get("kullanici")
            rol = session.get("rol")
            if kullanici != "admin" and rol != "admin" and rec["giren_kullanici"] != kullanici:
                return jsonify({"durum": "hata", "mesaj": "Sadece admin veya kaydı giren silebilir"}), 403
            conn.execute("DELETE FROM fason_irsaliye WHERE id = ?", (irs_id,))
            conn.commit()
            return jsonify({"durum": "ok", "mesaj": "Silindi"})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


# ═════════════════════════════════════════════════
# DURUM YÖNETİMİ
# ═════════════════════════════════════════════════

@fason_bp.route("/api/fason/durumlar", methods=["GET"])
def api_fason_durumlar():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    tumu = request.args.get("tumu") == "1"
    conn = get_db()
    try:
        if tumu:
            rows = conn.execute("SELECT * FROM fason_durum ORDER BY sira, id").fetchall()
        else:
            rows = conn.execute("SELECT * FROM fason_durum WHERE aktif = 1 ORDER BY sira, id").fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


@fason_bp.route("/api/fason/durum-ekle", methods=["POST"])
def api_fason_durum_ekle():
    if session.get("rol") != "admin" and session.get("kullanici") != "admin":
        return jsonify({"durum": "hata", "mesaj": "Sadece admin"}), 403
    try:
        d = request.get_json() or {}
        ad = (d.get("ad") or "").strip()
        renk = (d.get("renk") or "text3").strip()
        ikon = (d.get("ikon") or "fa-circle").strip()
        sira = int(d.get("sira") or 100)
        if not ad:
            return jsonify({"durum": "hata", "mesaj": "Ad zorunlu"}), 400
        conn = get_db()
        try:
            var = conn.execute("SELECT id FROM fason_durum WHERE ad = ?", (ad,)).fetchone()
            if var:
                return jsonify({"durum": "hata", "mesaj": "Bu ad zaten var"}), 400
            cur = conn.execute(
                "INSERT INTO fason_durum (ad, renk, ikon, sira) VALUES (?, ?, ?, ?)",
                (ad, renk, ikon, sira)
            )
            conn.commit()
            return jsonify({"durum": "ok", "id": cur.lastrowid, "mesaj": f"{ad} eklendi"})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


@fason_bp.route("/api/fason/durum-guncelle-tanim/<int:did>", methods=["POST"])
def api_fason_durum_guncelle_tanim(did):
    if session.get("rol") != "admin" and session.get("kullanici") != "admin":
        return jsonify({"durum": "hata", "mesaj": "Sadece admin"}), 403
    try:
        d = request.get_json() or {}
        conn = get_db()
        try:
            mevcut = conn.execute("SELECT * FROM fason_durum WHERE id = ?", (did,)).fetchone()
            if not mevcut:
                return jsonify({"durum": "hata", "mesaj": "Bulunamadı"}), 404
            ad   = (d.get("ad") or mevcut["ad"]).strip()
            renk = (d.get("renk") or mevcut["renk"]).strip()
            ikon = (d.get("ikon") or mevcut["ikon"]).strip()
            sira = int(d.get("sira") if d.get("sira") is not None else mevcut["sira"])
            aktif = 1 if d.get("aktif") in (True, 1, "1") else (0 if d.get("aktif") in (False, 0, "0") else mevcut["aktif"])
            conn.execute(
                "UPDATE fason_durum SET ad = ?, renk = ?, ikon = ?, sira = ?, aktif = ? WHERE id = ?",
                (ad, renk, ikon, sira, aktif, did)
            )
            conn.commit()
            return jsonify({"durum": "ok", "mesaj": "Güncellendi"})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


@fason_bp.route("/api/fason/durum-sil/<int:did>", methods=["DELETE"])
def api_fason_durum_sil(did):
    if session.get("rol") != "admin" and session.get("kullanici") != "admin":
        return jsonify({"durum": "hata", "mesaj": "Sadece admin"}), 403
    try:
        conn = get_db()
        try:
            c = conn.execute("SELECT COUNT(*) FROM fason_irsaliye WHERE durum_id = ?", (did,)).fetchone()[0]
            if c > 0:
                conn.execute("UPDATE fason_durum SET aktif = 0 WHERE id = ?", (did,))
                conn.commit()
                return jsonify({"durum": "ok", "mesaj": f"Pasifleştirildi ({c} irsaliyede kullanılıyor)"})
            else:
                conn.execute("DELETE FROM fason_durum WHERE id = ?", (did,))
                conn.commit()
                return jsonify({"durum": "ok", "mesaj": "Silindi"})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


# ═════════════════════════════════════════════════
# EXCEL EXPORT
# ═════════════════════════════════════════════════

@fason_bp.route("/api/fason/export", methods=["GET", "POST"])
def api_fason_export():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        conn = get_db()
        rows = conn.execute("""
            SELECT
                i.id, i.irsaliye_no, i.aciklama,
                i.giren_kullanici, i.girilme_tarihi,
                i.durum_notu, i.durum_guncelleyen, i.durum_guncelleme_tarihi,
                i.irsaliye_tarihi, i.stok_miktari, i.giris_miktari,
                i.otis_stok, i.otis_giris, i.toplam_fiyat,
                d.ad AS durum_ad,
                f.ad AS firma_ad
            FROM fason_irsaliye i
            LEFT JOIN fason_durum d ON i.durum_id = d.id
            LEFT JOIN fason_firma f ON i.firma_id = f.id
            ORDER BY i.id DESC
        """).fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Fason İrsaliyeleri"

        h_font = Font(bold=True, color="FFFFFF", size=11)
        h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        kilit_fill = PatternFill("solid", fgColor="374151")
        edit_fill = PatternFill("solid", fgColor="065F46")
        border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB")
        )

        # 16 kolon (Firma eklendi)
        basliklar = [
            ("ID",                    "kilit"),   # A
            ("İrsaliye No",           "kilit"),   # B
            ("Firma",                 "kilit"),   # C - YENİ
            ("Açıklama",              "kilit"),   # D
            ("Durum",                 "kilit"),   # E
            ("Durum Notu",            "kilit"),   # F
            ("Giren Kullanıcı",       "kilit"),   # G
            ("Girilme Tarihi",        "kilit"),   # H
            ("Durum Güncelleyen",     "kilit"),   # I
            ("Durum Güncel. Tarihi",  "kilit"),   # J
            ("İrsaliye Tarihi",       "edit"),    # K
            ("Stok Miktarı",          "edit"),    # L
            ("Giriş Miktarı",         "edit"),    # M
            ("OTIS Stok",             "edit"),    # N
            ("OTIS Giriş",            "edit"),    # O
            ("Toplam Fiyat (USD)",    "edit"),    # P
        ]

        ws.merge_cells("A1:J1")
        ws["A1"] = "SABİT SÜTUNLAR (değiştirmeyin)"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=10)
        ws["A1"].fill = kilit_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("K1:P1")
        ws["K1"] = "DÜZENLENEBİLİR (import ile güncellenir)"
        ws["K1"].font = Font(bold=True, color="FFFFFF", size=10)
        ws["K1"].fill = edit_fill
        ws["K1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for col_idx, (baslik, tip) in enumerate(basliklar, start=1):
            cell = ws.cell(row=2, column=col_idx, value=baslik)
            cell.font = h_font
            cell.fill = kilit_fill if tip == "kilit" else edit_fill
            cell.alignment = h_align
            cell.border = border
        ws.row_dimensions[2].height = 32

        genisliker = [6, 18, 22, 22, 22, 25, 14, 16, 16, 18, 14, 12, 12, 12, 12, 16]
        for i, w in enumerate(genisliker, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB")
        )

        for idx, r in enumerate(rows, start=3):
            values = [
                r["id"],
                r["irsaliye_no"],
                r["firma_ad"] or "",
                r["aciklama"] or "",
                r["durum_ad"] or "",
                r["durum_notu"] or "",
                r["giren_kullanici"],
                r["girilme_tarihi"] or "",
                r["durum_guncelleyen"] or "",
                r["durum_guncelleme_tarihi"] or "",
                r["irsaliye_tarihi"] or "",
                r["stok_miktari"],
                r["giris_miktari"],
                r["otis_stok"],
                r["otis_giris"],
                r["toplam_fiyat"],
            ]
            for c_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=idx, column=c_idx, value=v)
                cell.border = thin_border
                if c_idx in (1, 12, 13, 14, 15):
                    cell.alignment = Alignment(horizontal="right")
                    if c_idx != 1 and v is not None:
                        cell.number_format = "#,##0.00"
                elif c_idx == 16 and v is not None:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '"$"#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            if idx % 2 == 0:
                for c_idx in range(1, len(values) + 1):
                    ws.cell(row=idx, column=c_idx).fill = PatternFill("solid", fgColor="F9FAFB")

        ws.freeze_panes = "D3"

        klasor = os.path.join(_fason_export_klasor(), "exports", "excel")
        os.makedirs(klasor, exist_ok=True)
        dosya_adi = f"fason_irsaliyeleri_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        yol = os.path.join(klasor, dosya_adi)
        wb.save(yol)

        try:
            os.startfile(klasor)
        except Exception:
            pass

        return jsonify({
            "durum": "ok",
            "yol": yol,
            "dosya": dosya_adi,
            "kayit": len(rows),
            "mesaj": f"{len(rows)} kayıt Excel'e aktarıldı"
        })
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


# ═════════════════════════════════════════════════
# EXCEL IMPORT — Sadece 6 kolon güncellenir
# ═════════════════════════════════════════════════

@fason_bp.route("/api/fason/import", methods=["POST"])
def api_fason_import():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    if session.get("kullanici") != "admin" and session.get("rol") != "admin":
        return jsonify({"durum": "hata", "mesaj": "Sadece admin"}), 403
    if "dosya" not in request.files:
        return jsonify({"durum": "hata", "mesaj": "Dosya yok"}), 400
    dosya = request.files["dosya"]
    if not dosya.filename:
        return jsonify({"durum": "hata", "mesaj": "Dosya adı boş"}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(dosya, data_only=True)
        ws = wb.active

        h_row = None
        for row_idx in range(1, 5):
            row = [str(c.value or "").strip().lower() for c in ws[row_idx]]
            if "id" in row or "i̇d" in row:
                h_row = row_idx
                break

        if h_row is None:
            return jsonify({"durum": "hata", "mesaj": "Başlık satırında 'ID' bulunamadı"}), 400

        basliklar_raw = [str(c.value or "").strip() for c in ws[h_row]]
        def bul(anahtar_liste):
            for i, b in enumerate(basliklar_raw):
                bl = b.lower()
                for k in anahtar_liste:
                    if k in bl:
                        return i + 1
            return -1

        col_id       = bul(["id"])
        col_irs_tar  = bul(["i̇rsaliye tarihi", "irsaliye tarihi", "irs tarihi"])
        col_stok     = bul(["stok miktarı", "stok miktari"])
        col_giris    = bul(["giriş miktarı", "giris miktari"])
        col_otis_st  = bul(["otis stok"])
        col_otis_gi  = bul(["otis giriş", "otis giris"])
        col_fiyat    = bul(["toplam fiyat", "fiyat"])

        if col_id < 0:
            return jsonify({"durum": "hata", "mesaj": "ID sütunu bulunamadı"}), 400

        eslesenler = {
            "irsaliye_tarihi": col_irs_tar,
            "stok_miktari":    col_stok,
            "giris_miktari":   col_giris,
            "otis_stok":       col_otis_st,
            "otis_giris":      col_otis_gi,
            "toplam_fiyat":    col_fiyat,
        }
        bulunan_kolonlar = [k for k, v in eslesenler.items() if v > 0]
        if not bulunan_kolonlar:
            return jsonify({
                "durum": "hata",
                "mesaj": "Güncellenebilir sütun bulunamadı. Başlıklar: " + ", ".join(basliklar_raw[:20])
            }), 400

        guncellenen = 0
        atlanan = 0
        hatalar = []
        conn = get_db()
        try:
            for row_idx in range(h_row + 1, ws.max_row + 1):
                id_val = ws.cell(row=row_idx, column=col_id).value
                if id_val is None or id_val == "":
                    continue
                try:
                    irs_id = int(id_val)
                except:
                    atlanan += 1
                    continue

                var = conn.execute("SELECT id FROM fason_irsaliye WHERE id = ?", (irs_id,)).fetchone()
                if not var:
                    atlanan += 1
                    hatalar.append(f"Satır {row_idx}: ID {irs_id} bulunamadı")
                    continue

                degerler = {}
                for kolon_ad, col_idx in eslesenler.items():
                    if col_idx < 0:
                        continue
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if kolon_ad == "irsaliye_tarihi":
                        if v is None or v == "":
                            degerler[kolon_ad] = None
                        elif hasattr(v, "strftime"):
                            degerler[kolon_ad] = v.strftime("%d.%m.%Y")
                        else:
                            degerler[kolon_ad] = str(v).strip()
                    else:
                        if v is None or v == "":
                            degerler[kolon_ad] = None
                        else:
                            try:
                                s = str(v).replace(".", "").replace(",", ".") if isinstance(v, str) else v
                                degerler[kolon_ad] = float(s)
                            except:
                                degerler[kolon_ad] = None

                set_parts = []
                vals = []
                for k, val in degerler.items():
                    set_parts.append(f"{k} = ?")
                    vals.append(val)

                if set_parts:
                    vals.append(irs_id)
                    conn.execute(
                        f"UPDATE fason_irsaliye SET {', '.join(set_parts)} WHERE id = ?",
                        vals
                    )
                    guncellenen += 1
            conn.commit()
        finally:
            conn.close()

        return jsonify({
            "durum": "ok",
            "guncellenen": guncellenen,
            "atlanan": atlanan,
            "kolonlar": bulunan_kolonlar,
            "hatalar": hatalar[:20],
            "mesaj": f"{guncellenen} kayıt güncellendi" + (f", {atlanan} atlandı" if atlanan else "")
        })
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


# ═════════════════════════════════════════════════
# EXCEL'DEN TOPLU EKLEME — Firma zorunlu (form-data)
# ═════════════════════════════════════════════════

@fason_bp.route("/api/fason/toplu-ekle", methods=["POST"])
def api_fason_toplu_ekle():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    if "dosya" not in request.files:
        return jsonify({"durum": "hata", "mesaj": "Dosya yok"}), 400
    dosya = request.files["dosya"]
    if not dosya.filename:
        return jsonify({"durum": "hata", "mesaj": "Dosya adı boş"}), 400

    firma_id = request.form.get("firma_id")
    if not firma_id:
        return jsonify({"durum": "hata", "mesaj": "Firma seçiniz"}), 400
    try:
        firma_id = int(firma_id)
    except:
        return jsonify({"durum": "hata", "mesaj": "Geçersiz firma"}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(dosya, data_only=True)
        ws = wb.active

        h_row = None
        for row_idx in range(1, 6):
            row = [str(c.value or "").strip().lower() for c in ws[row_idx]]
            if any("irsaliye" in v or "sipariş" in v or "no" == v.strip() for v in row):
                h_row = row_idx
                break

        veri_baslangic = h_row + 1 if h_row else 1
        col_irs = 1
        col_aciklama = -1

        if h_row:
            basliklar = [str(c.value or "").strip().lower() for c in ws[h_row]]
            for i, b in enumerate(basliklar):
                if "irsaliye" in b or "no" == b.strip() or "sipariş" in b:
                    col_irs = i + 1
                elif "açıklama" in b or "aciklama" in b or "not" in b:
                    col_aciklama = i + 1
                    break

        conn = get_db()
        fv = conn.execute("SELECT id FROM fason_firma WHERE id = ?", (firma_id,)).fetchone()
        if not fv:
            conn.close()
            return jsonify({"durum": "hata", "mesaj": "Firma bulunamadı"}), 400

        eklenen = 0
        mukerrer = 0
        atlanan = 0
        atlanan_sebep = []
        kullanici = session.get("kullanici", "-")

        try:
            for row_idx in range(veri_baslangic, ws.max_row + 1):
                irs_val = ws.cell(row=row_idx, column=col_irs).value
                if irs_val is None:
                    continue
                irs_no = str(irs_val).strip()
                if not irs_no:
                    continue
                if isinstance(irs_val, float) and irs_val.is_integer():
                    irs_no = str(int(irs_val))
                if len(irs_no) > 50:
                    atlanan += 1
                    atlanan_sebep.append(f"Satır {row_idx}: {irs_no[:30]}... (çok uzun)")
                    continue

                aciklama = ""
                if col_aciklama > 0:
                    ac_val = ws.cell(row=row_idx, column=col_aciklama).value
                    if ac_val is not None:
                        aciklama = str(ac_val).strip()[:200]

                var = conn.execute(
                    "SELECT id FROM fason_irsaliye WHERE irsaliye_no = ?",
                    (irs_no,)
                ).fetchone()
                if var:
                    mukerrer += 1

                conn.execute("""
                    INSERT INTO fason_irsaliye
                        (irsaliye_no, aciklama, giren_kullanici, firma_id)
                    VALUES (?, ?, ?, ?)
                """, (irs_no, aciklama, kullanici, firma_id))
                eklenen += 1

            conn.commit()
        finally:
            conn.close()

        mesaj = f"{eklenen} irsaliye eklendi"
        if mukerrer:
            mesaj += f" ({mukerrer} tanesi daha önce vardı)"
        if atlanan:
            mesaj += f", {atlanan} satır atlandı"

        return jsonify({
            "durum": "ok",
            "eklenen": eklenen,
            "mukerrer": mukerrer,
            "atlanan": atlanan,
            "atlanan_sebep": atlanan_sebep[:20],
            "mesaj": mesaj
        })
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500