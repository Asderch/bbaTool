# -*- coding: utf-8 -*-
"""
hammadde_db.py — BBA Hammadde Takip Modülü
OTIS 'Tüm Gelen Malzemeler' export'unu işler, irsaliye/AK bazlı gruplar,
stok değişimini takip eder. Ayrı veritabanı (hammadde.db) — Fason/Sevkiyat'ı kilitlemez.
"""

import os
import sys
import sqlite3
import re
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, session

hammadde_bp = Blueprint("hammadde", __name__)

ORTAK_KLASOR = r"K:\Warehouse\Yeşilovacık\12_Paylaşım Klasörü\01-BBA\bba-tool"


def _db_klasor_bul():
    if os.path.isdir(ORTAK_KLASOR):
        return ORTAK_KLASOR
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


DB_KLASOR = _db_klasor_bul()
DB_YOL = os.path.join(DB_KLASOR, "hammadde.db")
SEVKIYAT_DB_YOL = os.path.join(DB_KLASOR, "sevkiyat.db")

# Hammadde sayılacak mal grupları (senin onayladığın liste)
HEDEF_MAL_GRUPLARI = {
    "Boru Çelik", "Boru Bağlantı Parçaları", "Profiller", "Saclar",
    "Nervürlü Demir", "Köşebent", "Düz Demir", "Çelik Hasır", "Bağlantı Elemanları"
}


def get_db():
    conn = sqlite3.connect(DB_YOL)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def log_kaydet(islem, detay="", ilgili_id=None, ilgili_ad=""):
    """Hammadde işlemlerini merkezi İşlem Geçmişi'ne (sevkiyat.db → islem_log) yazar."""
    try:
        conn = sqlite3.connect(SEVKIYAT_DB_YOL, timeout=10)
        conn.execute(
            "INSERT INTO islem_log (modul,islem,detay,ilgili_id,ilgili_ad,yapan,yapan_ad,tarih) VALUES (?,?,?,?,?,?,?,?)",
            ("Hammadde", islem, detay, ilgili_id, ilgili_ad,
             session.get("kullanici", "sistem"), session.get("ad", "Sistem"),
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[Hammadde log] hata: {e}")


def init_hammadde_db():
    conn = get_db()
    try:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS hammadde_irsaliye (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                irsaliye_no         TEXT NOT NULL UNIQUE,
                irsaliye_tarihi     TEXT,
                teslim_tarihi       TEXT,
                tedarikci           TEXT DEFAULT '',
                ilk_eklenme_tarihi  TEXT DEFAULT (datetime('now','localtime')),
                son_guncelleme_tarihi TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE INDEX IF NOT EXISTS ix_ham_irs_no ON hammadde_irsaliye(irsaliye_no);

            CREATE TABLE IF NOT EXISTS hammadde_kalem (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                irsaliye_id         INTEGER NOT NULL,
                malzeme_tanim       TEXT NOT NULL,
                mal_grubu           TEXT DEFAULT '',
                dokum_no            TEXT DEFAULT '',
                sap_kodu            TEXT DEFAULT '',
                birim               TEXT DEFAULT '',
                giris_miktari       REAL,
                stok_miktari        REAL,
                miktar_ikincil      REAL,
                sertifika_no        TEXT DEFAULT '',
                sertifika_tarihi    TEXT,
                siparis_no          TEXT DEFAULT '',
                siparis_tarihi      TEXT,
                siparis_veren       TEXT DEFAULT '',
                imalat_siparis_no   TEXT DEFAULT '',
                proje_no            TEXT DEFAULT '',
                depo                TEXT DEFAULT '',
                lokasyon            TEXT DEFAULT '',
                aciklama            TEXT DEFAULT '',
                imalat_tipi         TEXT DEFAULT '',
                uretim_tipi         TEXT DEFAULT '',
                musteri             TEXT DEFAULT '',
                tz_no               TEXT DEFAULT '',
                ilk_eklenme_tarihi  TEXT DEFAULT (datetime('now','localtime')),
                stok_son_degisim_tarihi TEXT,
                FOREIGN KEY (irsaliye_id) REFERENCES hammadde_irsaliye(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS ix_ham_kalem_irs   ON hammadde_kalem(irsaliye_id);
            CREATE INDEX IF NOT EXISTS ix_ham_kalem_ak    ON hammadde_kalem(siparis_no);
            CREATE INDEX IF NOT EXISTS ix_ham_kalem_proje ON hammadde_kalem(proje_no);

            CREATE TABLE IF NOT EXISTS hammadde_durum_log (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                kalem_id    INTEGER NOT NULL,
                durum       TEXT NOT NULL,
                not_metni   TEXT DEFAULT '',
                kullanici   TEXT DEFAULT '',
                tarih       TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (kalem_id) REFERENCES hammadde_kalem(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS ix_ham_durum_kalem ON hammadde_durum_log(kalem_id);

            CREATE TABLE IF NOT EXISTS hammadde_import_log (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                dosya_adi       TEXT,
                toplam_satir    INTEGER,
                filtrelenen     INTEGER,
                yeni_kalem      INTEGER,
                guncellenen     INTEGER,
                yeni_irsaliye   INTEGER,
                kullanici       TEXT DEFAULT '',
                tarih           TEXT DEFAULT (datetime('now','localtime'))
            );
        """)
        # SAP karşılaştırma alanları (sonradan eklendi)
        for kolon, tip in [
            ("sap_giris_miktari", "REAL"),
            ("sap_stok_miktari", "REAL"),
            ("sap_kontrol_notu", "TEXT DEFAULT ''"),
            ("sap_kontrol_tarihi", "TEXT"),
            ("sap_kontrol_kullanici", "TEXT DEFAULT ''"),
        ]:
            try:
                conn.execute(f"ALTER TABLE hammadde_kalem ADD COLUMN {kolon} {tip}")
            except Exception:
                pass
        try:
            conn.execute("ALTER TABLE hammadde_durum_log ADD COLUMN irsaliye_id INTEGER")
        except Exception:
            pass
        # kalem_id artık zorunlu değil (durum irsaliye bazlı hale geldi) — NOT NULL kısıtlamasını kaldırmak için tabloyu yeniden kur
        kolon_bilgisi = conn.execute("PRAGMA table_info(hammadde_durum_log)").fetchall()
        kalem_id_notnull = any(k[1] == "kalem_id" and k[3] == 1 for k in kolon_bilgisi)
        if kalem_id_notnull:
            conn.executescript("""
                CREATE TABLE hammadde_durum_log_yeni (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    kalem_id    INTEGER,
                    irsaliye_id INTEGER,
                    durum       TEXT NOT NULL,
                    not_metni   TEXT DEFAULT '',
                    kullanici   TEXT DEFAULT '',
                    tarih       TEXT DEFAULT (datetime('now','localtime'))
                );
                INSERT INTO hammadde_durum_log_yeni (id, kalem_id, irsaliye_id, durum, not_metni, kullanici, tarih)
                    SELECT id, kalem_id, irsaliye_id, durum, not_metni, kullanici, tarih FROM hammadde_durum_log;
                DROP TABLE hammadde_durum_log;
                ALTER TABLE hammadde_durum_log_yeni RENAME TO hammadde_durum_log;
                CREATE INDEX IF NOT EXISTS ix_ham_durum_kalem ON hammadde_durum_log(kalem_id);
                CREATE INDEX IF NOT EXISTS ix_ham_durum_irs ON hammadde_durum_log(irsaliye_id);
            """)
            print("[Hammadde DB] hammadde_durum_log tablosu yeniden kuruldu (kalem_id artık zorunlu değil)")        
        try:
            conn.execute("CREATE INDEX IF NOT EXISTS ix_ham_durum_irs ON hammadde_durum_log(irsaliye_id)")
        except Exception:
            pass 
        conn.commit()
    finally:
        conn.close()


# ═════════════════════════════════════════════════
# YARDIMCI DÖNÜŞÜMLER
# ═════════════════════════════════════════════════

def _sayi(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return None


def _tarih_donustur(v):
    """Excel tarih serisi (int/float) ya da metin -> 'YYYY-MM-DD' string."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        try:
            taban = datetime(1899, 12, 30)
            return (taban + timedelta(days=float(v))).strftime("%Y-%m-%d")
        except Exception:
            return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _metin(v):
    if v is None:
        return ""
    return str(v).strip()


def _giris_miktari_hesapla(birim, mkg, madt):
    """Birim KG ise KG miktarını, değilse ADET miktarını 'giriş miktarı' olarak alır."""
    b = (birim or "").strip().upper()
    if b == "KG":
        return _sayi(mkg)
    m = _sayi(madt)
    return m if m is not None else _sayi(mkg)


def _kalem_anahtari(malzeme_tanim, giris_miktari):
    """Eşleştirme anahtarı: malzeme tanım + giriş miktarı (irsaliye zaten grup seviyesinde ayrı)."""
    mt = (malzeme_tanim or "").strip().upper()
    gm = round(giris_miktari, 3) if giris_miktari is not None else None
    return (mt, gm)


# ═════════════════════════════════════════════════
# IMPORT — OTIS "Tüm Gelen Malzemeler" export'u
# ═════════════════════════════════════════════════

@hammadde_bp.route("/api/hammadde/import", methods=["POST"])
def api_hammadde_import():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    if "dosya" not in request.files:
        return jsonify({"durum": "hata", "mesaj": "Dosya yok"}), 400
    dosya = request.files["dosya"]
    if not dosya.filename:
        return jsonify({"durum": "hata", "mesaj": "Dosya adı boş"}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(dosya, data_only=True, read_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            hdr_row = next(rows_iter)
        except StopIteration:
            return jsonify({"durum": "hata", "mesaj": "Dosya boş"}), 400

        hdr = [(_metin(h)).upper() for h in hdr_row]

        def kol(*adaylar):
            for a in adaylar:
                if a in hdr:
                    return hdr.index(a)
            return -1

        c_stok      = kol("STOK")
        c_irs_no    = kol("İRSALİYE NO", "IRSALIYE NO")
        c_irs_tar   = kol("İRSALİYE TARİHİ", "IRSALIYE TARIHI")
        c_teslim    = kol("TESLİM TARİHİ", "TESLIM TARIHI")
        c_tedarikci = kol("TEDARİKÇİ", "TEDARIKCI")
        c_mal_grubu = kol("MAL GRUBU")
        c_tanim     = kol("MALZEME TANIM")
        c_birim     = kol("BİRİM", "BIRIM")
        c_dokum     = kol("DÖKÜM NO", "DOKUM NO")
        c_sap       = kol("SAP KODU")
        c_mkg       = kol("MİKTAR(KG)", "MIKTAR(KG)")
        c_madt      = kol("MİKTAR(ADT)", "MIKTAR(ADT)")
        c_mikincil  = kol("MİKTAR İKİNCİL", "MIKTAR IKINCIL")
        c_sert_no   = kol("SERTİFİKA NO", "SERTIFIKA NO")
        c_sert_tar  = kol("SERTİFİKA TARİHİ", "SERTIFIKA TARIHI")
        c_sip_no    = kol("SİPARİŞ NO", "SIPARIS NO")
        c_sip_tar   = kol("SİPARİŞ TARİHİ", "SIPARIS TARIHI")
        c_sip_veren = kol("SİPARİŞ VEREN", "SIPARIS VEREN")
        c_imalat_sp = kol("İMALAT SİPARİŞ NO", "IMALAT SIPARIS NO")
        c_proje     = kol("PROJE NO")
        c_depo      = kol("DEPO")
        c_lokasyon  = kol("LOKASYON")
        c_aciklama  = kol("AÇIKLAMA", "ACIKLAMA")
        c_imalat_tp = kol("İMALAT TİPİ", "IMALAT TIPI")
        c_uretim_tp = kol("ÜRETİM TİPİ", "URETIM TIPI")
        c_musteri   = kol("MÜŞTERİ", "MUSTERI")
        c_tz        = kol("TZ NO")

        if c_irs_no < 0 or c_tanim < 0 or c_mal_grubu < 0:
            return jsonify({"durum": "hata", "mesaj": "Beklenen kolonlar bulunamadı (İRSALİYE NO / MALZEME TANIM / MAL GRUBU)"}), 400

        def al(row, idx):
            if idx < 0 or idx >= len(row):
                return None
            return row[idx]

        conn = get_db()
        toplam_satir = 0
        filtrelenen = 0
        yeni_kalem = 0
        guncellenen = 0
        yeni_irsaliye = 0

        # irsaliye_no -> {id, mevcut_kalemler: {anahtar: [id, id, ...]}}
        irsaliye_cache = {}

        try:
            for row in rows_iter:
                toplam_satir += 1
                mal_grubu = _metin(al(row, c_mal_grubu))
                if mal_grubu not in HEDEF_MAL_GRUPLARI:
                    continue
                irsaliye_no = _metin(al(row, c_irs_no))
                malzeme_tanim = _metin(al(row, c_tanim))
                if not irsaliye_no or not malzeme_tanim:
                    continue
                filtrelenen += 1

                # --- İrsaliye başlığı ---
                if irsaliye_no not in irsaliye_cache:
                    r = conn.execute("SELECT id FROM hammadde_irsaliye WHERE irsaliye_no = ?", (irsaliye_no,)).fetchone()
                    if r:
                        irs_id = r["id"]
                    else:
                        cur = conn.execute("""
                            INSERT INTO hammadde_irsaliye (irsaliye_no, irsaliye_tarihi, teslim_tarihi, tedarikci)
                            VALUES (?, ?, ?, ?)
                        """, (
                            irsaliye_no,
                            _tarih_donustur(al(row, c_irs_tar)),
                            _tarih_donustur(al(row, c_teslim)),
                            _metin(al(row, c_tedarikci))
                        ))
                        irs_id = cur.lastrowid
                        yeni_irsaliye += 1

                    # bu irsaliyenin mevcut kalemlerini anahtara göre kuyruğa al
                    mevcut = {}
                    for kr in conn.execute("SELECT id, malzeme_tanim, giris_miktari FROM hammadde_kalem WHERE irsaliye_id = ? ORDER BY id", (irs_id,)):
                        k = _kalem_anahtari(kr["malzeme_tanim"], kr["giris_miktari"])
                        mevcut.setdefault(k, []).append(kr["id"])

                    irsaliye_cache[irsaliye_no] = {"id": irs_id, "kuyruk": mevcut}

                irs_info = irsaliye_cache[irsaliye_no]
                irs_id = irs_info["id"]

                birim = _metin(al(row, c_birim))
                giris_miktari = _giris_miktari_hesapla(birim, al(row, c_mkg), al(row, c_madt))
                stok_miktari = _sayi(al(row, c_stok))
                anahtar = _kalem_anahtari(malzeme_tanim, giris_miktari)

                kuyruk = irs_info["kuyruk"].get(anahtar, [])
                alanlar = dict(
                    mal_grubu=mal_grubu,
                    dokum_no=_metin(al(row, c_dokum)),
                    sap_kodu=_metin(al(row, c_sap)),
                    birim=birim,
                    stok_miktari=stok_miktari,
                    miktar_ikincil=_sayi(al(row, c_mikincil)),
                    sertifika_no=_metin(al(row, c_sert_no)),
                    sertifika_tarihi=_tarih_donustur(al(row, c_sert_tar)),
                    siparis_no=_metin(al(row, c_sip_no)),
                    siparis_tarihi=_tarih_donustur(al(row, c_sip_tar)),
                    siparis_veren=_metin(al(row, c_sip_veren)),
                    imalat_siparis_no=_metin(al(row, c_imalat_sp)),
                    proje_no=_metin(al(row, c_proje)),
                    depo=_metin(al(row, c_depo)),
                    lokasyon=_metin(al(row, c_lokasyon)),
                    aciklama=_metin(al(row, c_aciklama)),
                    imalat_tipi=_metin(al(row, c_imalat_tp)),
                    uretim_tipi=_metin(al(row, c_uretim_tp)),
                    musteri=_metin(al(row, c_musteri)),
                    tz_no=_metin(al(row, c_tz)),
                )

                if kuyruk:
                    kalem_id = kuyruk.pop(0)
                    eski = conn.execute("SELECT stok_miktari FROM hammadde_kalem WHERE id = ?", (kalem_id,)).fetchone()
                    stok_degisti = (eski is None) or (eski["stok_miktari"] != stok_miktari)
                    set_parts = ", ".join(f"{k} = ?" for k in alanlar)
                    vals = list(alanlar.values())
                    if stok_degisti:
                        set_parts += ", stok_son_degisim_tarihi = datetime('now','localtime')"
                    vals.append(kalem_id)
                    conn.execute(f"UPDATE hammadde_kalem SET {set_parts} WHERE id = ?", vals)
                    guncellenen += 1
                else:
                    conn.execute("""
                        INSERT INTO hammadde_kalem
                            (irsaliye_id, malzeme_tanim, giris_miktari, mal_grubu, dokum_no, sap_kodu,
                             birim, stok_miktari, miktar_ikincil, sertifika_no, sertifika_tarihi,
                             siparis_no, siparis_tarihi, siparis_veren, imalat_siparis_no, proje_no,
                             depo, lokasyon, aciklama, imalat_tipi, uretim_tipi, musteri, tz_no,
                             stok_son_degisim_tarihi)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))
                    """, (
                        irs_id, malzeme_tanim, giris_miktari,
                        alanlar["mal_grubu"], alanlar["dokum_no"], alanlar["sap_kodu"],
                        alanlar["birim"], alanlar["stok_miktari"], alanlar["miktar_ikincil"],
                        alanlar["sertifika_no"], alanlar["sertifika_tarihi"],
                        alanlar["siparis_no"], alanlar["siparis_tarihi"], alanlar["siparis_veren"],
                        alanlar["imalat_siparis_no"], alanlar["proje_no"], alanlar["depo"],
                        alanlar["lokasyon"], alanlar["aciklama"], alanlar["imalat_tipi"],
                        alanlar["uretim_tipi"], alanlar["musteri"], alanlar["tz_no"]
                    ))
                    yeni_kalem += 1

            conn.execute("UPDATE hammadde_irsaliye SET son_guncelleme_tarihi = datetime('now','localtime') WHERE irsaliye_no IN ({})".format(
                ",".join("?" * len(irsaliye_cache))
            ), list(irsaliye_cache.keys()) if irsaliye_cache else [])

            conn.commit()
        finally:
            conn.close()

        conn2 = sqlite3.connect(DB_YOL)
        conn2.execute("""
            INSERT INTO hammadde_import_log (dosya_adi, toplam_satir, filtrelenen, yeni_kalem, guncellenen, yeni_irsaliye, kullanici)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (dosya.filename, toplam_satir, filtrelenen, yeni_kalem, guncellenen, yeni_irsaliye, session.get("kullanici", "-")))
        conn2.commit()
        conn2.close()

        log_kaydet(
            "Import",
            f"{dosya.filename}: {yeni_kalem} yeni kalem, {guncellenen} güncellendi, {yeni_irsaliye} yeni irsaliye",
            None, dosya.filename
        )

        return jsonify({
            "durum": "ok",
            "toplam_satir": toplam_satir,
            "filtrelenen": filtrelenen,
            "yeni_kalem": yeni_kalem,
            "guncellenen": guncellenen,
            "yeni_irsaliye": yeni_irsaliye,
            "mesaj": f"{filtrelenen} hammadde satırı işlendi — {yeni_kalem} yeni, {guncellenen} güncellendi ({yeni_irsaliye} yeni irsaliye)"
        })
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


# ═════════════════════════════════════════════════
# LİSTELEME — İrsaliye kartları / AK listesi / detaylar
# ═════════════════════════════════════════════════

@hammadde_bp.route("/api/hammadde/irsaliyeler", methods=["GET"])
def api_hammadde_irsaliyeler():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    tedarikci = request.args.get("tedarikci", "").strip()
    tarih_bas = request.args.get("tarih_bas", "").strip()
    tarih_bit = request.args.get("tarih_bit", "").strip()

    sorgu = """
        SELECT
            i.id, i.irsaliye_no, i.irsaliye_tarihi, i.teslim_tarihi, i.tedarikci,
            i.son_guncelleme_tarihi,
            COUNT(k.id) AS kalem_sayisi,
            SUM(CASE WHEN k.stok_miktari > 0 THEN 1 ELSE 0 END) AS stokta_kalan,
            GROUP_CONCAT(DISTINCT NULLIF(k.siparis_no, '')) AS ak_listesi,
            (SELECT d.durum FROM hammadde_durum_log d WHERE d.irsaliye_id = i.id ORDER BY d.id DESC LIMIT 1) AS son_durum,
            (SELECT d.tarih FROM hammadde_durum_log d WHERE d.irsaliye_id = i.id ORDER BY d.id DESC LIMIT 1) AS son_durum_tarihi,
            (SELECT d.not_metni FROM hammadde_durum_log d WHERE d.irsaliye_id = i.id ORDER BY d.id DESC LIMIT 1) AS son_durum_notu
        FROM hammadde_irsaliye i
        LEFT JOIN hammadde_kalem k ON k.irsaliye_id = i.id
        WHERE 1=1
    """
    params = []
    if tedarikci:
        sorgu += " AND i.tedarikci = ?"
        params.append(tedarikci)
    if tarih_bas:
        sorgu += " AND i.irsaliye_tarihi >= ?"
        params.append(tarih_bas)
    if tarih_bit:
        sorgu += " AND i.irsaliye_tarihi <= ?"
        params.append(tarih_bit)
    sorgu += " GROUP BY i.id ORDER BY i.irsaliye_tarihi DESC, i.id DESC"

    conn = get_db()
    try:
        rows = conn.execute(sorgu, params).fetchall()
        sonuc = []
        irs_idler = []
        for r in rows:
            d = dict(r)
            d["ak_listesi"] = sorted(set((d["ak_listesi"] or "").split(","))) if d["ak_listesi"] else []
            d["durum_tarihleri"] = {}
            sonuc.append(d)
            irs_idler.append(d["id"])

        if irs_idler:
            ph = ",".join("?" * len(irs_idler))
            durum_rows = conn.execute(f"""
                SELECT irsaliye_id, durum, MAX(tarih) AS en_son
                FROM hammadde_durum_log
                WHERE irsaliye_id IN ({ph})
                GROUP BY irsaliye_id, durum
            """, irs_idler).fetchall()
            harita = {}
            for dr in durum_rows:
                harita.setdefault(dr["irsaliye_id"], {})[dr["durum"]] = dr["en_son"]
            for d in sonuc:
                d["durum_tarihleri"] = harita.get(d["id"], {})

        return jsonify(sonuc)
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


@hammadde_bp.route("/api/hammadde/irsaliye/<irsaliye_no>", methods=["GET"])
def api_hammadde_irsaliye_detay(irsaliye_no):
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    conn = get_db()
    try:
        irs = conn.execute("SELECT * FROM hammadde_irsaliye WHERE irsaliye_no = ?", (irsaliye_no,)).fetchone()
        if not irs:
            return jsonify({"durum": "hata", "mesaj": "İrsaliye bulunamadı"}), 404
        kalemler = conn.execute("""
            SELECT * FROM hammadde_kalem WHERE irsaliye_id = ? ORDER BY id
        """, (irs["id"],)).fetchall()
        return jsonify({
            "durum": "ok",
            "irsaliye": dict(irs),
            "kalemler": [dict(k) for k in kalemler]
        })
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


@hammadde_bp.route("/api/hammadde/ak-listesi", methods=["GET"])
def api_hammadde_ak_listesi():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT
                k.siparis_no,
                k.siparis_tarihi,
                k.siparis_veren,
                COUNT(*) AS kalem_sayisi,
                COUNT(DISTINCT k.irsaliye_id) AS irsaliye_sayisi,
                SUM(CASE WHEN k.stok_miktari > 0 THEN 1 ELSE 0 END) AS stokta_kalan,
                GROUP_CONCAT(DISTINCT NULLIF(k.proje_no, '')) AS proje_listesi
            FROM hammadde_kalem k
            WHERE k.siparis_no IS NOT NULL AND k.siparis_no != ''
            GROUP BY k.siparis_no
            ORDER BY k.siparis_tarihi DESC
        """).fetchall()
        sonuc = []
        for r in rows:
            d = dict(r)
            d["proje_listesi"] = sorted(set((d["proje_listesi"] or "").split(","))) if d["proje_listesi"] else []
            sonuc.append(d)
        return jsonify(sonuc)
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


@hammadde_bp.route("/api/hammadde/ak/<path:ak_no>", methods=["GET"])
def api_hammadde_ak_detay(ak_no):
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    conn = get_db()
    try:
        kalemler = conn.execute("""
            SELECT k.*, i.irsaliye_no, i.irsaliye_tarihi, i.tedarikci
            FROM hammadde_kalem k
            JOIN hammadde_irsaliye i ON i.id = k.irsaliye_id
            WHERE k.siparis_no = ?
            ORDER BY i.irsaliye_tarihi DESC, k.id
        """, (ak_no,)).fetchall()
        return jsonify({"durum": "ok", "ak_no": ak_no, "kalemler": [dict(k) for k in kalemler]})
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


# ═════════════════════════════════════════════════
# DURUM GEÇMİŞİ (SAT/SAS/SAP işleme/nakil vb. — genişletilebilir)
# ═════════════════════════════════════════════════

@hammadde_bp.route("/api/hammadde/irsaliye-durum-ekle/<int:irsaliye_id>", methods=["POST"])
def api_hammadde_irsaliye_durum_ekle(irsaliye_id):
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    d = request.get_json(silent=True) or {}
    durum = _metin(d.get("durum"))
    not_metni = _metin(d.get("not"))
    if not durum:
        return jsonify({"durum": "hata", "mesaj": "Durum adı zorunlu"}), 400
    conn = get_db()
    try:
        irs = conn.execute("SELECT id FROM hammadde_irsaliye WHERE id = ?", (irsaliye_id,)).fetchone()
        if not irs:
            return jsonify({"durum": "hata", "mesaj": "İrsaliye bulunamadı"}), 404
        conn.execute(
            "INSERT INTO hammadde_durum_log (irsaliye_id, durum, not_metni, kullanici) VALUES (?, ?, ?, ?)",
            (irsaliye_id, durum, not_metni, session.get("kullanici", "-"))
        )
        conn.commit()
        return jsonify({"durum": "ok", "mesaj": "Durum eklendi"})
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


@hammadde_bp.route("/api/hammadde/irsaliye-durum-gecmisi/<int:irsaliye_id>", methods=["GET"])
def api_hammadde_irsaliye_durum_gecmisi(irsaliye_id):
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM hammadde_durum_log WHERE irsaliye_id = ? ORDER BY id DESC", (irsaliye_id,)
        ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()

@hammadde_bp.route("/api/hammadde/tedarikciler", methods=["GET"])
def api_hammadde_tedarikciler():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT DISTINCT tedarikci FROM hammadde_irsaliye
            WHERE tedarikci IS NOT NULL AND tedarikci != ''
            ORDER BY tedarikci
        """).fetchall()
        return jsonify([r["tedarikci"] for r in rows])
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


# ═════════════════════════════════════════════════
# TÜM KALEMLER — filtreli düz liste (export/import için)
# ═════════════════════════════════════════════════

@hammadde_bp.route("/api/hammadde/kalemler", methods=["GET"])
def api_hammadde_kalemler():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    tedarikci = request.args.get("tedarikci", "").strip()
    tarih_bas = request.args.get("tarih_bas", "").strip()
    tarih_bit = request.args.get("tarih_bit", "").strip()

    sorgu = """
        SELECT k.*, i.irsaliye_no, i.irsaliye_tarihi, i.tedarikci
        FROM hammadde_kalem k
        JOIN hammadde_irsaliye i ON i.id = k.irsaliye_id
        WHERE 1=1
    """
    params = []
    if tedarikci:
        sorgu += " AND i.tedarikci = ?"
        params.append(tedarikci)
    if tarih_bas:
        sorgu += " AND i.irsaliye_tarihi >= ?"
        params.append(tarih_bas)
    if tarih_bit:
        sorgu += " AND i.irsaliye_tarihi <= ?"
        params.append(tarih_bit)
    sorgu += " ORDER BY i.irsaliye_tarihi DESC, k.id"

    conn = get_db()
    try:
        rows = conn.execute(sorgu, params).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()


# ═════════════════════════════════════════════════
# EXCEL EXPORT — kilitli (OTIS) + düzenlenebilir (SAP) sütunlar
# ═════════════════════════════════════════════════

@hammadde_bp.route("/api/hammadde/export", methods=["GET"])
def api_hammadde_export():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    tedarikci = request.args.get("tedarikci", "").strip()
    tarih_bas = request.args.get("tarih_bas", "").strip()
    tarih_bit = request.args.get("tarih_bit", "").strip()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        sorgu = """
            SELECT k.*, i.irsaliye_no, i.irsaliye_tarihi, i.tedarikci
            FROM hammadde_kalem k
            JOIN hammadde_irsaliye i ON i.id = k.irsaliye_id
            WHERE 1=1
        """
        params = []
        if tedarikci:
            sorgu += " AND i.tedarikci = ?"
            params.append(tedarikci)
        if tarih_bas:
            sorgu += " AND i.irsaliye_tarihi >= ?"
            params.append(tarih_bas)
        if tarih_bit:
            sorgu += " AND i.irsaliye_tarihi <= ?"
            params.append(tarih_bit)
        sorgu += " ORDER BY i.irsaliye_tarihi DESC, k.id"

        conn = get_db()
        rows = conn.execute(sorgu, params).fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Hammadde Kontrol"

        h_font = Font(bold=True, color="FFFFFF", size=11)
        h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        kilit_fill = PatternFill("solid", fgColor="374151")
        edit_fill = PatternFill("solid", fgColor="065F46")
        border = Border(
            left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB")
        )

        basliklar = [
            ("ID",                  "kilit"),
            ("İrsaliye No",         "kilit"),
            ("İrsaliye Tarihi",     "kilit"),
            ("Tedarikçi",           "kilit"),
            ("Malzeme Tanım",       "kilit"),
            ("Mal Grubu",           "kilit"),
            ("Döküm No",            "kilit"),
            ("SAP Kodu (OTIS)",     "kilit"),
            ("Proje No",            "kilit"),
            ("OTIS Giriş Miktarı",  "kilit"),
            ("OTIS Stok Miktarı",   "kilit"),
            ("SAP Giriş Miktarı",   "edit"),
            ("SAP Stok Miktarı",    "edit"),
            ("SAP Kontrol Notu",    "edit"),
        ]

        ws.merge_cells("A1:K1")
        ws["A1"] = "SABİT SÜTUNLAR — OTIS export verisi (değiştirmeyin)"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=10)
        ws["A1"].fill = kilit_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("L1:N1")
        ws["L1"] = "SAP'TAN KONTROL EDİP DOLDURUN"
        ws["L1"].font = Font(bold=True, color="FFFFFF", size=10)
        ws["L1"].fill = edit_fill
        ws["L1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for col_idx, (baslik, tip) in enumerate(basliklar, start=1):
            cell = ws.cell(row=2, column=col_idx, value=baslik)
            cell.font = h_font
            cell.fill = kilit_fill if tip == "kilit" else edit_fill
            cell.alignment = h_align
            cell.border = border
        ws.row_dimensions[2].height = 34

        genislikler = [6, 18, 14, 16, 34, 16, 14, 16, 12, 14, 14, 14, 14, 24]
        for i, w in enumerate(genislikler, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"), right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"), bottom=Side(style="thin", color="E5E7EB")
        )

        for idx, r in enumerate(rows, start=3):
            values = [
                r["id"], r["irsaliye_no"], r["irsaliye_tarihi"] or "", r["tedarikci"] or "",
                r["malzeme_tanim"], r["mal_grubu"] or "", r["dokum_no"] or "", r["sap_kodu"] or "",
                r["proje_no"] or "", r["giris_miktari"], r["stok_miktari"],
                r["sap_giris_miktari"], r["sap_stok_miktari"], r["sap_kontrol_notu"] or "",
            ]
            for c_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=idx, column=c_idx, value=v)
                cell.border = thin_border
                if c_idx in (1, 10, 11, 12, 13):
                    cell.alignment = Alignment(horizontal="right")
                    if c_idx != 1 and v is not None:
                        cell.number_format = "#,##0.00"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            if idx % 2 == 0:
                for c_idx in range(1, len(values) + 1):
                    ws.cell(row=idx, column=c_idx).fill = PatternFill("solid", fgColor="F9FAFB")

        ws.freeze_panes = "E3"

        klasor = os.path.join(DB_KLASOR, "exports", "hammadde")
        os.makedirs(klasor, exist_ok=True)
        dosya_adi = f"hammadde_kontrol_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        yol = os.path.join(klasor, dosya_adi)
        wb.save(yol)
        try:
            os.startfile(klasor)
        except Exception:
            pass

        return jsonify({"durum": "ok", "yol": yol, "dosya": dosya_adi, "kayit": len(rows),
                         "mesaj": f"{len(rows)} kayıt Excel'e aktarıldı"})
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500


# ═════════════════════════════════════════════════
# EXCEL IMPORT — SAP karşılaştırma sütunlarını geri yükle
# ═════════════════════════════════════════════════

@hammadde_bp.route("/api/hammadde/sap-import", methods=["POST"])
def api_hammadde_sap_import():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    if "dosya" not in request.files:
        return jsonify({"durum": "hata", "mesaj": "Dosya yok"}), 400
    dosya = request.files["dosya"]
    if not dosya.filename:
        return jsonify({"durum": "hata", "mesaj": "Dosya adı boş"}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(dosya, data_only=True)
        ws = wb.active

        h_row = None
        for row_idx in range(1, 5):
            row = [str(c.value or "").strip().lower() for c in ws[row_idx]]
            if "id" in row:
                h_row = row_idx
                break
        if h_row is None:
            return jsonify({"durum": "hata", "mesaj": "Başlık satırında 'ID' bulunamadı"}), 400

        basliklar = [str(c.value or "").strip().lower() for c in ws[h_row]]
        def bul(anahtarlar):
            for i, b in enumerate(basliklar):
                if any(a in b for a in anahtarlar):
                    return i + 1
            return -1

        col_id    = bul(["id"])
        col_sgir  = bul(["sap giriş", "sap giris"])
        col_sstok = bul(["sap stok"])
        col_not   = bul(["sap kontrol notu", "kontrol notu"])

        if col_id < 0:
            return jsonify({"durum": "hata", "mesaj": "ID sütunu bulunamadı"}), 400

        guncellenen = 0
        degisiklikler = []
        conn = get_db()
        try:
            for row_idx in range(h_row + 1, ws.max_row + 1):
                id_val = ws.cell(row=row_idx, column=col_id).value
                if id_val is None or id_val == "":
                    continue
                try:
                    kalem_id = int(id_val)
                except Exception:
                    continue

                mevcut = conn.execute("SELECT * FROM hammadde_kalem WHERE id = ?", (kalem_id,)).fetchone()
                if not mevcut:
                    continue

                def sayi(v):
                    if v is None or v == "": return None
                    try: return float(v)
                    except Exception: return None

                sap_giris = sayi(ws.cell(row=row_idx, column=col_sgir).value) if col_sgir > 0 else mevcut["sap_giris_miktari"]
                sap_stok  = sayi(ws.cell(row=row_idx, column=col_sstok).value) if col_sstok > 0 else mevcut["sap_stok_miktari"]
                notu      = str(ws.cell(row=row_idx, column=col_not).value or "").strip() if col_not > 0 else mevcut["sap_kontrol_notu"]

                degisti = (sap_giris != mevcut["sap_giris_miktari"]) or (sap_stok != mevcut["sap_stok_miktari"]) or (notu != (mevcut["sap_kontrol_notu"] or ""))
                if not degisti:
                    continue

                conn.execute("""
                    UPDATE hammadde_kalem
                    SET sap_giris_miktari = ?, sap_stok_miktari = ?, sap_kontrol_notu = ?,
                        sap_kontrol_tarihi = datetime('now','localtime'), sap_kontrol_kullanici = ?
                    WHERE id = ?
                """, (sap_giris, sap_stok, notu, session.get("kullanici", "-"), kalem_id))
                guncellenen += 1

                fark_var = False
                if sap_stok is not None and mevcut["stok_miktari"] is not None and abs(sap_stok - mevcut["stok_miktari"]) > 0.01:
                    fark_var = True
                degisiklikler.append({
                    "malzeme_tanim": mevcut["malzeme_tanim"],
                    "otis_stok": mevcut["stok_miktari"], "sap_stok": sap_stok,
                    "fark_var": fark_var
                })
            conn.commit()
        finally:
            conn.close()

        log_kaydet("SAP Karşılaştırma Import",
                    f"{dosya.filename}: {guncellenen} kayıt güncellendi", None, dosya.filename)

        return jsonify({
            "durum": "ok", "guncellenen": guncellenen, "degisiklikler": degisiklikler[:100],
            "mesaj": f"{guncellenen} kayıt güncellendi"
        })
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    
@hammadde_bp.route("/api/hammadde/temizle", methods=["POST"])
def api_hammadde_temizle():
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    if session.get("kullanici") != "admin":
        return jsonify({"durum": "hata", "mesaj": "Sadece admin"}), 403
    conn = get_db()
    try:
        conn.execute("DELETE FROM hammadde_durum_log")
        conn.execute("DELETE FROM hammadde_kalem")
        conn.execute("DELETE FROM hammadde_irsaliye")
        conn.execute("DELETE FROM hammadde_import_log")
        conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('hammadde_durum_log','hammadde_kalem','hammadde_irsaliye','hammadde_import_log')")
        conn.commit()
        log_kaydet("Tüm Veriyi Temizleme", "Hammadde modülündeki tüm irsaliye/kalem/durum verisi silindi", None, "")
        return jsonify({"durum": "ok", "mesaj": "Tüm hammadde verisi temizlendi"})
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()

@hammadde_bp.route("/api/hammadde/kalem-sap-guncelle/<int:kalem_id>", methods=["POST"])
def api_hammadde_kalem_sap_guncelle(kalem_id):
    if not session.get("kullanici"):
        return jsonify({"durum": "hata", "mesaj": "Giriş gerekli"}), 401
    d = request.get_json(silent=True) or {}
    conn = get_db()
    try:
        mevcut = conn.execute("SELECT * FROM hammadde_kalem WHERE id = ?", (kalem_id,)).fetchone()
        if not mevcut:
            return jsonify({"durum": "hata", "mesaj": "Kalem bulunamadı"}), 404

        sap_giris = mevcut["sap_giris_miktari"]
        sap_stok = mevcut["sap_stok_miktari"]
        notu = mevcut["sap_kontrol_notu"]

        if "sap_giris_miktari" in d:
            sap_giris = _sayi(d.get("sap_giris_miktari"))
        if "sap_stok_miktari" in d:
            sap_stok = _sayi(d.get("sap_stok_miktari"))
        if "sap_kontrol_notu" in d:
            notu = _metin(d.get("sap_kontrol_notu"))

        conn.execute("""
            UPDATE hammadde_kalem
            SET sap_giris_miktari = ?, sap_stok_miktari = ?, sap_kontrol_notu = ?,
                sap_kontrol_tarihi = datetime('now','localtime'), sap_kontrol_kullanici = ?
            WHERE id = ?
        """, (sap_giris, sap_stok, notu, session.get("kullanici", "-"), kalem_id))
        conn.commit()
        return jsonify({"durum": "ok", "mesaj": "Güncellendi"})
    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        conn.close()