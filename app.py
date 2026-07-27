from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from agirlik_db import agirlik_bp, init_agirlik_db
from mb52_backend import mb52_bp, mb52_init_db
from malzeme_kontrol import malzeme_bp, init_malzeme_db

# ═══ KULLANICI YÖNETİMİ ═══
from kullanici_db import (
    kullanici_bp,
    kullanici_dosyasini_hazirla,
    sifre_dogrula,
    kullanici_bilgi as kullanici_bilgi_al,
    VARSAYILAN_ROLLER as ROLLER_YENI,
)

from versiyon_db import (
    versiyon_bp,
    versiyon_dosyasini_hazirla,
    guncel_mi,
)

import threading
import webview
import time
import os
import base64
import sys
import sqlite3
import json
from datetime import datetime, date, timedelta

# ----------------------
# ORTAK DATABASE YOLU
# ----------------------

#K:\Warehouse\Yeşilovacık\12_Paylaşım Klasörü\01-BBA\bba-tool

def get_db_klasor():
    """Ağ klasörü varsa onu kullan, yoksa exe'nin yanına yaz"""
    ag_klasor = r"K:\Warehouse\Yeşilovacık\12_Paylaşım Klasörü\01-BBA\bba-tool"
    if os.path.exists(ag_klasor):
        return ag_klasor
    # Ağ yoksa exe'nin / script'in bulunduğu klasör
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

DB_KLASOR = get_db_klasor()
DB_DOSYA = "sevkiyat.db"

def get_base_path():
    if getattr(sys, 'frozen', False): return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

def get_export_path():
    if getattr(sys, 'frozen', False): return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

base = get_base_path()
app = Flask(__name__, template_folder=os.path.join(base,"templates"), static_folder=os.path.join(base,"assets"), static_url_path="/assets")
app.secret_key = "ambar-data-gizli-anahtar-2026"
app.register_blueprint(agirlik_bp)
app.register_blueprint(mb52_bp)
app.register_blueprint(malzeme_bp)
app.register_blueprint(kullanici_bp)
app.register_blueprint(versiyon_bp)
versiyon_dosyasini_hazirla()

# ═══ Kullanıcı JSON dosyasını hazırla (yoksa oluşturur) ═══
kullanici_dosyasini_hazirla()

from functools import wraps

APP_VERSION = "4.2"
APP_ADI     = "Warehouse Data"      # Sidebar logo başlığı için
APP_PREP    = "Berkcan Burak Akar"  # Footer için

@app.context_processor
def inject_globals():
    """Tüm Jinja template'lerinde otomatik erişilebilir değişkenler"""
    return {
        "APP_VERSION": APP_VERSION,
        "APP_ADI":     APP_ADI,
        "APP_PREP":    APP_PREP,
    }

# ═══════════════════════════════════════════════
# ROLLER artık kullanici_db.py içinde (VARSAYILAN_ROLLER)
# ═══════════════════════════════════════════════

def kullanici_rol():
    return session.get("rol")

def kullanici_yetkileri():
    # ← Değişti: ROLLER yerine ROLLER_YENI (kullanici_db.py'den)
    return ROLLER_YENI.get(kullanici_rol(), {})

def yetki_var_mi(yetki):
    return kullanici_yetkileri().get(yetki, False)

def yetki_gerekli(yetki):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not giris_yapildi_mi():
                return jsonify({"durum": "hata", "mesaj": "Oturum yok"}), 401

            if not yetki_var_mi(yetki):
                return jsonify({"durum": "hata", "mesaj": "Bu işlem için yetkiniz yok."}), 403

            return f(*args, **kwargs)
        return wrapper
    return decorator
# ═══════════════════════════════════════════════
# KULLANICILAR artık kullanici_db.py içinde (JSON tabanlı, ortak klasörde)
# İlk kullanıcılar ILK_KULLANICILAR sabitinde tanımlı.
# ═══════════════════════════════════════════════

# ═══════════════════════════════════════════════
# SÜRÜM KİLİDİ MIDDLEWARE
# Eski sürüm ise sadece /guncelle ve /api/versiyon/* açık kalır
# ═══════════════════════════════════════════════
@app.before_request
def surum_kilit_kontrolu():
    # Guncelleme ekranı ve static dosyalar her zaman açık
    if request.path.startswith("/guncelle") or \
       request.path.startswith("/api/versiyon/") or \
       request.path.startswith("/assets/") or \
       request.path.startswith("/static/") or \
       request.path == "/logout" or \
       request.path.startswith("/favicon"):
        return None

    # Sürüm güncel değilse zorla /guncelle'ye yönlendir
    if not guncel_mi():
        return redirect(url_for("guncelle_sayfasi"))


def giris_yapildi_mi(): return session.get("kullanici") is not None


# ═══════════════════════════════════════════════
# TROLL MIDDLEWARE 😄
# "yasakli" rolündeki kullanıcılar her tıklamada rickroll'e gider
# ═══════════════════════════════════════════════
@app.before_request
def yasakli_troll_kontrolu():
    # Bu path'ler açık kalmalı (yoksa çıkış bile yapamaz)
    if request.path.startswith("/rickroll") or \
       request.path.startswith("/assets/") or \
       request.path.startswith("/static/") or \
       request.path == "/logout" or \
       request.path == "/login" or \
       request.path.startswith("/favicon"):
        return None

    # Kullanıcı yasakli rolündeyse zorla rickroll'e yönlendir
    if session.get("rol") == "yasakli":
        return redirect(url_for("rickroll_sayfasi"))

# ----------------------
# SQLite
# ----------------------
def get_db():

    os.makedirs(DB_KLASOR, exist_ok=True)

    db_yol = os.path.join(DB_KLASOR, DB_DOSYA)

    conn = sqlite3.connect(db_yol, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row

    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")

    return conn

def veritabani_olustur():
    conn=get_db()
    try: conn.execute("SELECT planlanan_miktar FROM sevkiyat_kalem LIMIT 1")
    except:
        conn.execute("DROP TABLE IF EXISTS sevkiyat_hareket")
        conn.execute("DROP TABLE IF EXISTS sevkiyat_kalem")
        conn.execute("DROP TABLE IF EXISTS sevkiyat_plan")
    conn.execute("""CREATE TABLE IF NOT EXISTS sevkiyat_plan (
        id INTEGER PRIMARY KEY AUTOINCREMENT, plan_adi TEXT NOT NULL, plan_tipi TEXT NOT NULL,
        baslangic TEXT NOT NULL, bitis TEXT NOT NULL, olusturan TEXT NOT NULL,
        olusturan_ad TEXT NOT NULL, tarih TEXT NOT NULL, durum TEXT DEFAULT 'Aktif')""")
    conn.execute("""CREATE TABLE IF NOT EXISTS sevkiyat_kalem (
        id INTEGER PRIMARY KEY AUTOINCREMENT, plan_id INTEGER NOT NULL,
        yuklenici_firma TEXT, siparis_no TEXT, mal_grubu TEXT, malzeme_tanimi TEXT NOT NULL,
        planlanan_miktar REAL NOT NULL, gonderilen_miktar REAL DEFAULT 0, birim TEXT DEFAULT 'KG',
        tir_plaka TEXT, durum TEXT DEFAULT 'Bekliyor',
        devreden_plan_id INTEGER, devreden_plan_adi TEXT,
        hedef_plan_id INTEGER, hedef_plan_adi TEXT,
        not_ TEXT,
        FOREIGN KEY (plan_id) REFERENCES sevkiyat_plan(id))""")
    conn.execute("""CREATE TABLE IF NOT EXISTS sevkiyat_hareket (
        id INTEGER PRIMARY KEY AUTOINCREMENT, kalem_id INTEGER NOT NULL,
        islem TEXT NOT NULL, miktar REAL NOT NULL, tir_plaka TEXT,
        yapan TEXT, yapan_ad TEXT, tarih TEXT NOT NULL,
        detay TEXT,
        FOREIGN KEY (kalem_id) REFERENCES sevkiyat_kalem(id))""")

    # detay kolonu yoksa ekle (mevcut db için)
    try:
        conn.execute("SELECT detay FROM sevkiyat_hareket LIMIT 1")
    except:
        try:
            conn.execute("ALTER TABLE sevkiyat_hareket ADD COLUMN detay TEXT")
        except:
            pass

    # hedef_plan kolonları (devreden kalem nereye gitti)
    for col in ["hedef_plan_id", "hedef_plan_adi"]:
        try:
            conn.execute(f"ALTER TABLE sevkiyat_kalem ADD COLUMN {col} TEXT")
        except:
            pass

    # ---- PERSONEL TABLOSU ----
    conn.execute("""CREATE TABLE IF NOT EXISTS personel (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sicil_no TEXT,
        ad_soyad TEXT NOT NULL,
        departman TEXT DEFAULT 'Ambar',
        pozisyon TEXT,
        yaka TEXT DEFAULT 'Mavi',
        lokasyon TEXT,
        ise_giris TEXT,
        cinsiyet TEXT,
        toplam_izin INTEGER DEFAULT 14,
        kullanilan_izin REAL DEFAULT 0,
        aktif INTEGER DEFAULT 1,
        not_ TEXT
    )""")

    # Mevcut DB için yeni kolonları ekle
    for col, tip in [("sicil_no", "TEXT"), ("cinsiyet", "TEXT")]:
        try:
            conn.execute(f"ALTER TABLE personel ADD COLUMN {col} {tip}")
        except:
            pass

    conn.execute("""CREATE TABLE IF NOT EXISTS izin_kayit (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        personel_id INTEGER NOT NULL,
        baslangic TEXT NOT NULL,
        bitis TEXT NOT NULL,
        gun_sayisi REAL NOT NULL,
        izin_turu TEXT DEFAULT 'Yıllık İzin',
        aciklama TEXT,
        ekleyen TEXT,
        ekleyen_ad TEXT,
        tarih TEXT NOT NULL,
        FOREIGN KEY (personel_id) REFERENCES personel(id)
    )""")

    # ---- İŞLEM LOG TABLOSU ----
    conn.execute("""CREATE TABLE IF NOT EXISTS islem_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        modul TEXT NOT NULL,
        islem TEXT NOT NULL,
        detay TEXT,
        ilgili_id INTEGER,
        ilgili_ad TEXT,
        yapan TEXT,
        yapan_ad TEXT,
        tarih TEXT NOT NULL
    )""")

    # ---- PERFORMANS İNDEXLERİ ----
    conn.execute("CREATE INDEX IF NOT EXISTS idx_kalem_plan_id ON sevkiyat_kalem(plan_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_kalem_durum ON sevkiyat_kalem(durum)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_kalem_mal_grubu ON sevkiyat_kalem(mal_grubu)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_kalem_plan_durum ON sevkiyat_kalem(plan_id, durum)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_hareket_kalem_id ON sevkiyat_hareket(kalem_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_hareket_islem ON sevkiyat_hareket(islem)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_plan_tipi ON sevkiyat_plan(plan_tipi)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_plan_bitis ON sevkiyat_plan(bitis)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_plan_durum ON sevkiyat_plan(durum)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_izin_personel ON izin_kayit(personel_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_log_modul ON islem_log(modul)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_log_tarih ON islem_log(tarih)")

    conn.commit(); conn.close()


def log_kaydet(modul, islem, detay="", ilgili_id=None, ilgili_ad=""):
    """Merkezi log fonksiyonu"""
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO islem_log (modul,islem,detay,ilgili_id,ilgili_ad,yapan,yapan_ad,tarih) VALUES (?,?,?,?,?,?,?,?)",
            (modul, islem, detay, ilgili_id, ilgili_ad,
             session.get("kullanici","sistem"), session.get("ad","Sistem"),
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
        conn.close()
    except:
        pass


def mal_grubu_kategorileri_olustur():
    """Varsayılan ana kategorileri oluştur (yoksa)"""
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS ana_mal_grubu (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        kategori_adi TEXT NOT NULL,
        renk TEXT DEFAULT '#1565c0'
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS mal_grubu_esleme (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mal_grubu TEXT NOT NULL UNIQUE,
        ana_kategori_id INTEGER,
        FOREIGN KEY (ana_kategori_id) REFERENCES ana_mal_grubu(id)
    )""")

    conn.execute("CREATE INDEX IF NOT EXISTS idx_esleme_mg ON mal_grubu_esleme(mal_grubu)")

    mevcut = conn.execute("SELECT COUNT(*) as c FROM ana_mal_grubu").fetchone()["c"]
    if mevcut == 0:
        conn.execute("INSERT INTO ana_mal_grubu (kategori_adi, renk) VALUES (?, ?)", ("N.Demir", "#c62828"))
        conn.execute("INSERT INTO ana_mal_grubu (kategori_adi, renk) VALUES (?, ?)", ("Hammadde", "#e67e22"))
        conn.execute("INSERT INTO ana_mal_grubu (kategori_adi, renk) VALUES (?, ?)", ("Hazır Ürün", "#217346"))



    conn.commit()
    conn.close()

# ---- ROUTES ----
@app.route("/splash")
def splash(): return render_template("splash.html")

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method=="POST":
        d = request.json
        k = d.get("kullanici","").strip().lower()
        s = d.get("sifre","").strip()

        # JSON tabanlı doğrulama (kullanici_db.py)
        if not sifre_dogrula(k, s):
            return jsonify({"durum":"hata","mesaj":"Kullanıcı adı veya şifre hatalı."}), 401

        bilgi = kullanici_bilgi_al(k)
        if not bilgi:
            return jsonify({"durum":"hata","mesaj":"Kullanıcı bulunamadı."}), 401

        session["kullanici"] = k
        session["ad"]        = bilgi["ad"]
        session["rol"]       = bilgi["rol"]
        return jsonify({"durum":"ok", "ad":bilgi["ad"], "rol":bilgi["rol"]})

    return render_template("login.html")

@app.route("/logout")
def logout(): session.clear(); return redirect(url_for("login"))

@app.route("/")
def index():
    if not giris_yapildi_mi(): return redirect(url_for("login"))
    return render_template("index.html", ad=session.get("ad"))

@app.route("/agirlik-hesaplama")
def hesaplama():
    if not giris_yapildi_mi(): return redirect(url_for("login"))
    return render_template("agirlik-hesaplama.html")

@app.route("/personel")
def personel_sayfa():
    if not giris_yapildi_mi(): return redirect(url_for("login"))
    if kullanici_rol() == "sayim": return redirect(url_for("index"))
    return render_template("personel.html")

@app.route("/islem-gecmisi")
def islem_gecmisi_sayfa():
    if not giris_yapildi_mi(): return redirect(url_for("login"))
    if not yetki_var_mi("panel_gor"): return redirect(url_for("index"))
    return render_template("islem-gecmisi.html")

@app.route("/nakil-dashboard")
def nakil_dashboard_sayfa():
    if not giris_yapildi_mi(): return redirect(url_for("login"))
    return render_template("nakil-dashboard.html")

@app.route("/saha_doluluk")
def saha_doluluk_sayfa():
    if not giris_yapildi_mi(): return redirect(url_for("login"))
    return render_template("saha-doluluk.html")

@app.route("/fark-raporu")
def fark_raporu():
    return render_template("fark-raporu.html")

@app.route("/api/nakil-dashboard/export-html", methods=["POST"])
def nakil_dashboard_export():
    """NAK verisini alıp self-contained HTML dosyası oluştur"""
    if not giris_yapildi_mi():
        return jsonify({"durum":"hata","mesaj":"Yetkisiz"}), 401
    try:
        import json as json_mod
        d = request.json
        nak_data = d.get("nak", [])
        if not nak_data:
            return jsonify({"durum":"hata","mesaj":"Veri yok"}), 400

        nak_json = json_mod.dumps(nak_data, ensure_ascii=False)

        template_path = os.path.join(app.template_folder, "nakil-dashboard.html")
        with open(template_path, "r", encoding="utf-8") as f:
            html = f.read()

        # 1) Sidebar kaldır
        start_marker = '<!-- APP SIDEBAR -->'
        end_marker = '<!-- /APP SIDEBAR -->'
        s_idx = html.find(start_marker)
        e_idx = html.find(end_marker)
        if s_idx >= 0 and e_idx >= 0:
            html = html[:s_idx] + html[e_idx + len(end_marker):]

        # 2) margin-left kaldır
        html = html.replace('margin-left:220px', 'margin-left:0')

        # 3) Dış dosya referanslarını kaldır
        html = html.replace('<link rel="stylesheet" href="/assets/css/style.css">', '')
        html = html.replace('<script src="/assets/js/theme.js"></script>', '')
        html = html.replace('<script src="/assets/js/popup.js" defer></script>', '')

        # 4) nd-main margin fix
        html = html.replace('class="nd-main"', 'class="nd-main" style="margin-left:0"')

        # 5) Butonları gizle
        html = html.replace('id="temaToggleSidebar"', 'id="temaToggleSidebar" style="display:none"')
        html = html.replace('id="temaToggleTopbar"', 'id="temaToggleTopbar" style="display:none"')
        html = html.replace('class="nd-btn green"', 'class="nd-btn green" style="display:none"')
        html = html.replace('id="uploadZone"', 'id="uploadZone" style="display:none"')

        # 6) Dark tema inline CSS ekle
        dark_css = """
<style>
:root {
  --bg:#0e1116;--bg-content:#13161c;--bg-white:#1a1e27;--bg-card:#1a1e27;
  --bg-hover:#21262f;--border:#282e3a;--border-light:#21262f;
  --text:#cdd1da;--text2:#7e8698;--text3:#4e5568;
  --blue:#5b8def;--blue-dim:#1a2540;--green:#3dbc8e;--green-dim:#142b22;
  --amber:#c9922e;--amber-dim:#2a2010;--red:#c75050;--red-dim:#2a1515;
  --purple:#8070b8;--purple-dim:#1e1a30;--teal:#4a9e94;--teal-dim:#142826;
  --pink:#b05580;--pink-dim:#281520;
  --radius:8px;--radius-lg:12px;
  --shadow-sm:0 1px 3px rgba(0,0,0,.06);--shadow:0 4px 12px rgba(0,0,0,.08);
  --transition:.15s ease;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter','Segoe UI',sans-serif;background:var(--bg);color:var(--text);font-size:13px;-webkit-font-smoothing:antialiased}
.sidebar{display:none!important}
table{width:100%;border-collapse:collapse;font-size:12px}
th{background:var(--bg-hover);color:var(--text3);padding:8px 10px;text-align:left;font-size:10px;font-weight:600;border-bottom:1px solid var(--border)}
td{padding:8px 10px;border-bottom:1px solid var(--border);color:var(--text)}
tr:hover td{background:var(--bg-hover)}
input,select{background:var(--bg-hover);border:1px solid var(--border);color:var(--text);border-radius:var(--radius);padding:6px 10px;font-size:12px}
button{font-family:inherit}
</style>
"""
        html = html.replace('<head>', '<head>' + dark_css)

        # 7) Veri embed + dummy fonksiyonlar
        embed_script = f'''RAW=[];
NAK={nak_json};
FIL=[...NAK];
setTimeout(function(){{
  popF();
  applyFilters();
  showPage("overview",document.querySelector(".nd-tab"));
}}, 500);
function temaToggle(){{}}
function toast(){{}}
function bildirim(){{return Promise.resolve(true)}}
function onayla(){{return Promise.resolve(true)}}'''
        html = html.replace('/* embedded data placeholder */', embed_script)

        klasor = os.path.join(get_export_path(), "exports", "html")
        os.makedirs(klasor, exist_ok=True)
        dosya_adi = "nakil_dashboard_" + datetime.now().strftime("%Y%m%d_%H%M") + ".html"
        yol = os.path.join(klasor, dosya_adi)

        with open(yol, "w", encoding="utf-8") as f:
            f.write(html)

        try:
            os.startfile(klasor)
        except:
            pass

        return jsonify({"durum":"ok","yol":yol,"dosya":dosya_adi,"kayit":len(nak_data)})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/islem-log/liste")
@yetki_gerekli("panel_gor")
def islem_log_liste():
    try:
        modul = request.args.get("modul")
        kullanici = request.args.get("kullanici")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        limit = int(request.args.get("limit", 200))

        conn = get_db()
        sql = "SELECT * FROM islem_log WHERE 1=1"
        params = []

        if modul:
            sql += " AND modul=?"
            params.append(modul)
        if kullanici:
            sql += " AND yapan_ad LIKE ?"
            params.append(f"%{kullanici}%")
        if start_date:
            sql += " AND tarih>=?"
            params.append(start_date)
        if end_date:
            sql += " AND tarih<=?"
            params.append(end_date + " 23:59:59")

        sql += " ORDER BY id DESC LIMIT ?"
        params.append(limit)

        loglar = conn.execute(sql, params).fetchall()
        conn.close()

        return jsonify([{
            "id": l["id"],
            "modul": l["modul"],
            "islem": l["islem"],
            "detay": l["detay"],
            "ilgili_id": l["ilgili_id"],
            "ilgili_ad": l["ilgili_ad"],
            "yapan_ad": l["yapan_ad"],
            "tarih": l["tarih"]
        } for l in loglar])
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

# ---- MAL GRUBU KATEGORİ API ----
@app.route("/api/mal-grubu/kategoriler")
def mal_grubu_kategoriler():
    if not giris_yapildi_mi(): return jsonify({"durum":"hata"}), 401
    try:
        conn = get_db()
        kategoriler = conn.execute("SELECT * FROM ana_mal_grubu ORDER BY id").fetchall()
        eslesmeler = conn.execute("SELECT * FROM mal_grubu_esleme ORDER BY mal_grubu").fetchall()

        # Tüm benzersiz mal gruplarını al
        tum_mg = conn.execute("""
            SELECT DISTINCT mal_grubu FROM sevkiyat_kalem
            WHERE mal_grubu IS NOT NULL AND mal_grubu != ''
            ORDER BY mal_grubu
        """).fetchall()
        conn.close()

        esleme_dict = {e["mal_grubu"]: e["ana_kategori_id"] for e in eslesmeler}

        return jsonify({
            "kategoriler": [{"id": k["id"], "kategori_adi": k["kategori_adi"], "renk": k["renk"]} for k in kategoriler],
            "eslesmeler": esleme_dict,
            "mal_gruplari": [m["mal_grubu"] for m in tum_mg]
        })
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/mal-grubu/kategori-ekle", methods=["POST"])
@yetki_gerekli("plan_olustur")
def mal_grubu_kategori_ekle():
    try:
        d = request.json
        conn = get_db()
        conn.execute("INSERT INTO ana_mal_grubu (kategori_adi, renk) VALUES (?,?)",
                     (d["kategori_adi"], d.get("renk", "#1565c0")))
        conn.commit()
        conn.close()
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/mal-grubu/kategori-sil", methods=["POST"])
@yetki_gerekli("plan_sil")
def mal_grubu_kategori_sil():
    try:
        kid = request.json["id"]
        conn = get_db()
        conn.execute("DELETE FROM mal_grubu_esleme WHERE ana_kategori_id=?", (kid,))
        conn.execute("DELETE FROM ana_mal_grubu WHERE id=?", (kid,))
        conn.commit()
        conn.close()
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/mal-grubu/esle", methods=["POST"])
@yetki_gerekli("plan_olustur")
def mal_grubu_esle():
    """Mal gruplarını ana kategorilere ata. Body: {eslesmeler: {mal_grubu: kategori_id, ...}}"""
    try:
        d = request.json
        eslesmeler = d.get("eslesmeler", {})
        conn = get_db()

        for mg, kid in eslesmeler.items():
            if kid is None or kid == "":
                conn.execute("DELETE FROM mal_grubu_esleme WHERE mal_grubu=?", (mg,))
            else:
                mevcut = conn.execute("SELECT id FROM mal_grubu_esleme WHERE mal_grubu=?", (mg,)).fetchone()
                if mevcut:
                    conn.execute("UPDATE mal_grubu_esleme SET ana_kategori_id=? WHERE mal_grubu=?", (int(kid), mg))
                else:
                    conn.execute("INSERT INTO mal_grubu_esleme (mal_grubu, ana_kategori_id) VALUES (?,?)", (mg, int(kid)))

        conn.commit()
        conn.close()
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

# ---- PERSONEL API ----
@app.route("/api/personel/liste")
def personel_liste():
    if kullanici_rol() == "sayim": return jsonify({"durum":"hata","mesaj":"Yetkiniz yok"}), 403
    if not giris_yapildi_mi(): return jsonify({"durum":"hata"}), 401
    try:
        conn = get_db()
        personeller = conn.execute("SELECT * FROM personel ORDER BY id").fetchall()

        result = []
        bugun = date.today()

        for p in personeller:
            # İzin kayıtlarını al
            izinler = conn.execute(
                "SELECT * FROM izin_kayit WHERE personel_id=? ORDER BY baslangic DESC",
                (p["id"],)
            ).fetchall()

            kullanilan = sum(i["gun_sayisi"] for i in izinler)
            toplam_izin = p["toplam_izin"] if p["toplam_izin"] is not None else 14
            kalan_izin = max(toplam_izin - kullanilan, 0)

            # Aktif izinde mi kontrol
            izinde = False
            aktif_izin = None
            for i in izinler:
                try:
                    iz_bas = datetime.strptime(i["baslangic"], "%Y-%m-%d").date()
                    iz_bit = datetime.strptime(i["bitis"], "%Y-%m-%d").date()
                    if iz_bas <= bugun <= iz_bit:
                        izinde = True
                        aktif_izin = {
                            "baslangic": i["baslangic"],
                            "bitis": i["bitis"],
                            "gun_sayisi": i["gun_sayisi"],
                            "izin_turu": i["izin_turu"]
                        }
                        break
                except:
                    pass

            result.append({
                "id": p["id"],
                "ad_soyad": p["ad_soyad"],
                "sicil_no": p["sicil_no"],
                "departman": p["departman"],
                "pozisyon": p["pozisyon"],
                "yaka": p["yaka"],
                "lokasyon": p["lokasyon"],
                "ise_giris": p["ise_giris"],
                "toplam_izin": toplam_izin,
                "kullanilan_izin": kullanilan,
                "kalan_izin": kalan_izin,
                "aktif": p["aktif"],
                "not_": p["not_"],
                "izinde": izinde,
                "aktif_izin": aktif_izin,
                "izinler": [{
                    "id": i["id"],
                    "baslangic": i["baslangic"],
                    "bitis": i["bitis"],
                    "gun_sayisi": i["gun_sayisi"],
                    "izin_turu": i["izin_turu"],
                    "aciklama": i["aciklama"],
                    "ekleyen_ad": i["ekleyen_ad"],
                    "tarih": i["tarih"]
                } for i in izinler]
            })

        conn.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/personel/ekle", methods=["POST"])
@yetki_gerekli("plan_olustur")
def personel_ekle():
    try:
        d = request.json
        conn = get_db()
        conn.execute("""
            INSERT INTO personel (
                ad_soyad, sicil_no, departman, pozisyon, yaka, lokasyon, ise_giris, toplam_izin, not_
            )
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            d["ad_soyad"],
            d.get("sicil_no", ""),
            d.get("departman", "Ambar"),
            d.get("pozisyon", ""),
            d.get("yaka", "Mavi"),
            d.get("lokasyon", ""),
            d.get("ise_giris", ""),
            d.get("toplam_izin") if d.get("toplam_izin") is not None else 14,
            d.get("not_", "")
        ))
        conn.commit()
        conn.close()
        log_kaydet("Personel", "Personel Ekleme", f"{d['ad_soyad']} — {d.get('pozisyon','')}", None, d["ad_soyad"])
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/personel/guncelle", methods=["POST"])
@yetki_gerekli("plan_olustur")
def personel_guncelle():
    try:
        d = request.json
        conn = get_db()
        conn.execute("""
            UPDATE personel
            SET ad_soyad=?, sicil_no=?, departman=?, pozisyon=?, yaka=?, lokasyon=?,
                ise_giris=?, toplam_izin=?, not_=?
            WHERE id=?
        """, (
            d["ad_soyad"],
            d.get("sicil_no", ""),
            d.get("departman", "Ambar"),
            d.get("pozisyon", ""),
            d.get("yaka", "Mavi"),
            d.get("lokasyon", ""),
            d.get("ise_giris", ""),
            d.get("toplam_izin") if d.get("toplam_izin") is not None else 14,
            d.get("not_", ""),
            d["id"]
        ))
        conn.commit()
        conn.close()
        log_kaydet("Personel", "Personel Güncelleme", f"{d['ad_soyad']} güncellendi", d["id"], d["ad_soyad"])
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/personel/sil", methods=["POST"])
@yetki_gerekli("plan_sil")
def personel_sil():
    try:
        pid = request.json["id"]
        conn = get_db()
        p = conn.execute("SELECT ad_soyad FROM personel WHERE id=?", (pid,)).fetchone()
        p_ad = p["ad_soyad"] if p else f"#{pid}"
        conn.execute("DELETE FROM izin_kayit WHERE personel_id=?", (pid,))
        conn.execute("DELETE FROM personel WHERE id=?", (pid,))
        conn.commit()
        conn.close()
        log_kaydet("Personel", "Personel Silme", f"{p_ad} silindi", pid, p_ad)
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/personel/izin-ekle", methods=["POST"])
@yetki_gerekli("plan_olustur")
def personel_izin_ekle():
    try:
        d = request.json
        conn = get_db()
        conn.execute("""
            INSERT INTO izin_kayit (personel_id, baslangic, bitis, gun_sayisi, izin_turu, aciklama, ekleyen, ekleyen_ad, tarih)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (d["personel_id"], d["baslangic"], d["bitis"], float(d["gun_sayisi"]),
              d.get("izin_turu","Yıllık İzin"), d.get("aciklama",""),
              session["kullanici"], session["ad"],
              datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        conn.close()
        log_kaydet("Personel", "İzin Ekleme", f"Personel #{d['personel_id']}: {d['baslangic']} — {d['bitis']} ({d['gun_sayisi']} gün, {d.get('izin_turu','Yıllık İzin')})", d["personel_id"])
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/personel/izin-sil", methods=["POST"])
@yetki_gerekli("plan_sil")
def personel_izin_sil():
    try:
        conn = get_db()
        conn.execute("DELETE FROM izin_kayit WHERE id=?", (request.json["id"],))
        conn.commit()
        conn.close()
        log_kaydet("Personel", "İzin Silme", f"İzin kaydı #{request.json['id']} silindi")
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500


@app.route("/api/raporlar")
@yetki_gerekli("rapor_gor")
def api_raporlar():
    def dl(k, t):
        l = []
        if os.path.exists(k):
            for a in sorted(os.listdir(k), reverse=True):
                y = os.path.join(k, a)
                if os.path.isfile(y):
                    l.append({
                        "ad": a,
                        "yol": y,
                        "tip": t,
                        "boyut": os.path.getsize(y),
                        "tarih": datetime.fromtimestamp(os.path.getmtime(y)).strftime("%d.%m.%Y %H:%M")
                    })
        return l

    tum = dl(os.path.join(get_export_path(), "exports", "pdf"), "pdf") + \
          dl(os.path.join(get_export_path(), "exports", "excel"), "excel")

    tum.sort(key=lambda x: x["tarih"], reverse=True)
    return jsonify(tum)

@app.route("/sevkiyat-planlar")
def sevkiyat_planlar():
    if not giris_yapildi_mi(): return redirect(url_for("login"))
    return render_template("sevkiyat-planlar.html")

@app.route("/sevkiyat-detay/<int:plan_id>")
def sevkiyat_detay(plan_id):
    if not giris_yapildi_mi(): return redirect(url_for("login"))
    return render_template("sevkiyat-detay.html")

@app.route("/sevkiyat-dashboard")
def sevkiyat_dashboard():
    if not giris_yapildi_mi():
        return redirect(url_for("login"))

    if not yetki_var_mi("panel_gor"):
        return redirect(url_for("index"))

    return render_template("sevkiyat-dashboard.html")


@app.route("/api/kullanici-bilgi")
def kullanici_bilgi():
    if not giris_yapildi_mi():
        return jsonify({"durum":"hata"}), 401

    # JSON'dan TAZE oku - admin biri rolü değiştirdiyse anında yansır
    username = session["kullanici"]
    bilgi = kullanici_bilgi_al(username)

    # Kullanıcı silinmişse session'ı temizle
    if not bilgi:
        session.clear()
        return jsonify({"durum":"hata","mesaj":"Kullanıcı silinmiş"}), 401

    # Session'ı da güncelle (rol değişimleri senkron olsun)
    session["ad"]  = bilgi["ad"]
    session["rol"] = bilgi["rol"]

    return jsonify({
        "kullanici": username,
        "ad":        bilgi["ad"],
        "rol":       bilgi["rol"],
        "yetkiler":  bilgi["yetkiler"]
    })


# ═══════════════════════════════════════════════
# KULLANICI YÖNETİMİ SAYFASI (sadece admin)
# ═══════════════════════════════════════════════
@app.route("/kullanici-yonetimi")
def kullanici_yonetimi_sayfa():
    if not giris_yapildi_mi():
        return redirect(url_for("login"))
    # Sadece "admin" kullanıcı adı ile giriş yapan görebilir
    if session.get("kullanici") != "admin":
        return redirect(url_for("index"))
    return render_template("kullanici-yonetimi.html")

# Güncelleme kilit ekranı (herkese açık — kimlik doğrulama bile gerektirmez)
@app.route("/guncelle")
def guncelle_sayfasi():
    return render_template("guncelle.html")


# Sürüm yönetimi sayfası — sadece admin kullanıcı adı
@app.route("/surum-yonetimi")
def surum_yonetimi_sayfa():
    if not giris_yapildi_mi():
        return redirect(url_for("login"))
    if session.get("kullanici") != "admin":
        return redirect(url_for("index"))
    return render_template("surum-yonetimi.html")


# ---- SEVKİYAT PLAN API ----
@app.route("/api/sevkiyat/plan-yeni", methods=["POST"])
@yetki_gerekli("plan_olustur")
def sevkiyat_plan_yeni():
    try:
        d = request.json
        conn = get_db()
        conn.execute(
            "INSERT INTO sevkiyat_plan (plan_adi,plan_tipi,baslangic,bitis,olusturan,olusturan_ad,tarih) VALUES (?,?,?,?,?,?,?)",
            (
                d["plan_adi"],
                d["plan_tipi"],
                d["baslangic"],
                d["bitis"],
                session["kullanici"],
                session["ad"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            )
        )
        conn.commit()
        pid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.close()
        log_kaydet("Sevkiyat", "Plan Oluşturma", f"Plan: {d['plan_adi']} ({d['plan_tipi']})", pid, d["plan_adi"])
        return jsonify({"durum":"ok","id":pid})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/sevkiyat/plan-liste")
def sevkiyat_plan_liste():
    if not giris_yapildi_mi():
        return jsonify({"durum": "hata"}), 401

    try:
        conn = get_db()
        plans = conn.execute("SELECT * FROM sevkiyat_plan ORDER BY id DESC").fetchall()

        result = []

        for p in plans:
            toplam = conn.execute("""
                SELECT 
                    COUNT(*) as c,
                    SUM(planlanan_miktar) as pm,
                    SUM(gonderilen_miktar) as gm
                FROM sevkiyat_kalem
                WHERE plan_id=?
            """, (p["id"],)).fetchone()

            durum_sayac = conn.execute("""
                SELECT
                    durum,
                    COUNT(*) as adet,
                    SUM(
                        CASE
                            WHEN planlanan_miktar - gonderilen_miktar > 0
                            THEN planlanan_miktar - gonderilen_miktar
                            ELSE 0
                        END
                    ) as kg
                FROM sevkiyat_kalem
                WHERE plan_id=?
                GROUP BY durum
            """, (p["id"],)).fetchall()

            ozet = {
                r["durum"]: {
                    "adet": r["adet"] or 0,
                    "kg": r["kg"] or 0
                }
                for r in durum_sayac
            }

            toplam_kalem = toplam["c"] or 0
            pm = toplam["pm"] or 0
            gercek_gm = toplam["gm"] or 0

            gosterilen_gonderilen = min(gercek_gm, pm)
            fazla_gonderim = max(gercek_gm - pm, 0)
            kalan_miktar = max(pm - gosterilen_gonderilen, 0)

            bekleyen = ozet.get("Bekliyor", {}).get("adet", 0)
            kismi = ozet.get("Kısmi", {}).get("adet", 0)

            plan_durum = p["durum"]

            if plan_durum != "Kapatıldı":
                if toplam_kalem > 0 and bekleyen == 0 and kismi == 0:
                    plan_durum = "Tamamlandı"
                else:
                    plan_durum = "Aktif"

            yuzde = 0
            if pm > 0:
                yuzde = round(min((gosterilen_gonderilen / pm) * 100, 100), 1)

            result.append({
                "id": p["id"],
                "plan_adi": p["plan_adi"],
                "plan_tipi": p["plan_tipi"],
                "baslangic": p["baslangic"],
                "bitis": p["bitis"],
                "olusturan_ad": p["olusturan_ad"],
                "tarih": p["tarih"],
                "durum": plan_durum,
                "toplam_kalem": toplam_kalem,
                "planlanan_miktar": pm,
                "gonderilen_miktar": gosterilen_gonderilen,
                "gercek_gonderilen_miktar": gercek_gm,
                "fazla_gonderim": fazla_gonderim,
                "kalan_miktar": kalan_miktar,
                "yuzde": yuzde,
                "ozet": ozet
            })

        conn.close()
        return jsonify(result)

    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500   

@app.route("/api/sevkiyat/plan-detay/<int:plan_id>")
def sevkiyat_plan_detay(plan_id):
    if not giris_yapildi_mi():
        return jsonify({"durum":"hata"}), 401

    try:
        conn = get_db()
        p = conn.execute("SELECT * FROM sevkiyat_plan WHERE id=?", (plan_id,)).fetchone()

        if not p:
            conn.close()
            return jsonify({"durum":"hata","mesaj":"Plan bulunamadı"}), 404

        kalemler = conn.execute(
            "SELECT * FROM sevkiyat_kalem WHERE plan_id=? ORDER BY id",
            (plan_id,)
        ).fetchall()

        kalem_liste = []

        for k in kalemler:
            hareketler = conn.execute(
                "SELECT * FROM sevkiyat_hareket WHERE kalem_id=? ORDER BY id",
                (k["id"],)
            ).fetchall()

            planlanan = k["planlanan_miktar"] or 0
            gercek_gonderilen = k["gonderilen_miktar"] or 0

            gonderilen_gosterim = min(gercek_gonderilen, planlanan)
            fazla = max(gercek_gonderilen - planlanan, 0)
            kalan = max(planlanan - gonderilen_gosterim, 0)

            # Hedef plan bilgisini al (varsa direkt, yoksa otomatik bul)
            hedef_pid = k["hedef_plan_id"] if "hedef_plan_id" in k.keys() else None
            hedef_padi = k["hedef_plan_adi"] if "hedef_plan_adi" in k.keys() else None

            # Eğer durum Devreden ama hedef boşsa, yeni plandaki devir kaleminden bul
            if k["durum"] == "Devreden" and not hedef_pid:
                hedef_kalem = conn.execute("""
                    SELECT sk.plan_id, sp.plan_adi FROM sevkiyat_kalem sk
                    JOIN sevkiyat_plan sp ON sk.plan_id = sp.id
                    WHERE sk.devreden_plan_id=? AND sk.malzeme_tanimi=? AND sk.plan_id!=?
                    ORDER BY sk.id DESC LIMIT 1
                """, (k["plan_id"], k["malzeme_tanimi"], k["plan_id"])).fetchone()
                if hedef_kalem:
                    hedef_pid = hedef_kalem["plan_id"]
                    hedef_padi = hedef_kalem["plan_adi"]

            kalem_liste.append({
                "id": k["id"],
                "yuklenici_firma": k["yuklenici_firma"],
                "siparis_no": k["siparis_no"],
                "mal_grubu": k["mal_grubu"],
                "malzeme_tanimi": k["malzeme_tanimi"],

                "planlanan_miktar": planlanan,
                "gonderilen_miktar": gercek_gonderilen,
                "gonderilen_gosterim": gonderilen_gosterim,
                "kalan_miktar": kalan,
                "fazla_miktar": fazla,

                "birim": k["birim"],
                "tir_plaka": k["tir_plaka"],
                "durum": k["durum"],
                "devreden_plan_id": k["devreden_plan_id"],
                "devreden_plan_adi": k["devreden_plan_adi"],
                "hedef_plan_id": hedef_pid,
                "hedef_plan_adi": hedef_padi,
                "not_": k["not_"],

                "hareketler": [
                    {
                        "id": h["id"],
                        "islem": h["islem"],
                        "miktar": h["miktar"],
                        "tir_plaka": h["tir_plaka"],
                        "yapan_ad": h["yapan_ad"],
                        "tarih": h["tarih"],
                        "detay": h["detay"] if "detay" in h.keys() else ""
                    }
                    for h in hareketler
                ]
            })

        conn.close()

        return jsonify({
            "plan": {
                "id": p["id"],
                "plan_adi": p["plan_adi"],
                "plan_tipi": p["plan_tipi"],
                "baslangic": p["baslangic"],
                "bitis": p["bitis"],
                "olusturan_ad": p["olusturan_ad"],
                "tarih": p["tarih"],
                "durum": p["durum"]
            },
            "kalemler": kalem_liste
        })

    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/sevkiyat/plan-sil", methods=["POST"])
@yetki_gerekli("plan_sil")
def sevkiyat_plan_sil():
    try:
        pid = request.json["id"]
        conn = get_db()
        kalem_ids = conn.execute("SELECT id FROM sevkiyat_kalem WHERE plan_id=?", (pid,)).fetchall()

        for ki in kalem_ids:
            conn.execute("DELETE FROM sevkiyat_hareket WHERE kalem_id=?", (ki["id"],))

        conn.execute("DELETE FROM sevkiyat_kalem WHERE plan_id=?", (pid,))
        plan = conn.execute("SELECT plan_adi FROM sevkiyat_plan WHERE id=?", (pid,)).fetchone()
        plan_adi = plan["plan_adi"] if plan else f"Plan #{pid}"
        conn.execute("DELETE FROM sevkiyat_plan WHERE id=?", (pid,))
        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Plan Silme", f"Plan silindi: {plan_adi} (ID: {pid})", pid, plan_adi)
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500


@app.route("/api/sevkiyat/plan-kapat", methods=["POST"])
@yetki_gerekli("plan_kapat")
def sevkiyat_plan_kapat():
    try:
        conn = get_db()
        conn.execute("UPDATE sevkiyat_plan SET durum='Kapatıldı' WHERE id=?", (request.json["id"],))
        conn.commit()
        conn.close()
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

# ---- KALEM API ----
@app.route("/api/sevkiyat/kalem-ekle", methods=["POST"])
@yetki_gerekli("kalem_ekle")
def sevkiyat_kalem_ekle():
    try:
        d = request.json
        conn = get_db()
        conn.execute(
            "INSERT INTO sevkiyat_kalem (plan_id,yuklenici_firma,siparis_no,mal_grubu,malzeme_tanimi,planlanan_miktar,birim) VALUES (?,?,?,?,?,?,?)",
            (
                d["plan_id"],
                d.get("yuklenici_firma",""),
                d.get("siparis_no",""),
                d.get("mal_grubu",""),
                d["malzeme_tanimi"],
                d["miktar"],
                d.get("birim","KG")
            )
        )
        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Kalem Ekleme", f"{d['malzeme_tanimi']} - {d['miktar']} {d.get('birim','KG')}", d["plan_id"], d["malzeme_tanimi"])
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/sevkiyat/kalem-toplu-ekle", methods=["POST"])
@yetki_gerekli("kalem_toplu_ekle")
def sevkiyat_kalem_toplu_ekle():
    try:
        d = request.json
        pid = d["plan_id"]
        conn = get_db()

        for k in d["kalemler"]:
            conn.execute(
                "INSERT INTO sevkiyat_kalem (plan_id,yuklenici_firma,siparis_no,mal_grubu,malzeme_tanimi,planlanan_miktar,birim) VALUES (?,?,?,?,?,?,?)",
                (
                    pid,
                    k.get("yuklenici_firma",""),
                    k.get("siparis_no",""),
                    k.get("mal_grubu",""),
                    k["malzeme_tanimi"],
                    k["miktar"],
                    k.get("birim","KG")
                )
            )

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Toplu Kalem Ekleme", f"{len(d['kalemler'])} kalem eklendi (Plan ID: {pid})", pid)
        return jsonify({"durum":"ok","eklenen":len(d["kalemler"])})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/sevkiyat/kalem-sil", methods=["POST"])
@yetki_gerekli("kalem_sil")
def sevkiyat_kalem_sil():
    try:
        kid = request.json["id"]
        conn = get_db()

        kalem = conn.execute("SELECT * FROM sevkiyat_kalem WHERE id=?", (kid,)).fetchone()
        kalem_bilgi = f"{kalem['malzeme_tanimi']} ({kalem['planlanan_miktar']} {kalem['birim']})" if kalem else f"Kalem #{kid}"

        # Silme logunu plan bazında kaydet (kalem silinecek ama plan_id'den ilk kalemi referans al)
        plan_id = kalem["plan_id"] if kalem else None

        conn.execute("DELETE FROM sevkiyat_hareket WHERE kalem_id=?", (kid,))
        conn.execute("DELETE FROM sevkiyat_kalem WHERE id=?", (kid,))

        # Silme kaydını hareket tablosuna yaz (kalem_id=0 çünkü kalem silindi)
        if plan_id:
            conn.execute(
                "INSERT INTO sevkiyat_hareket (kalem_id,islem,miktar,tir_plaka,yapan,yapan_ad,tarih,detay) VALUES (?,?,?,?,?,?,?,?)",
                (0, "silme", 0, "", session["kullanici"], session["ad"],
                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                 f"Kalem silindi: {kalem_bilgi} (Plan ID: {plan_id})")
            )

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Kalem Silme", f"Kalem silindi: {kalem_bilgi}", plan_id, kalem_bilgi)
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/sevkiyat/kalem-guncelle", methods=["POST"])
@yetki_gerekli("kalem_guncelle")
def sevkiyat_kalem_guncelle():
    try:
        d = request.json
        conn = get_db()

        eski = conn.execute("SELECT * FROM sevkiyat_kalem WHERE id=?", (d["id"],)).fetchone()

        conn.execute("""
            UPDATE sevkiyat_kalem
            SET yuklenici_firma=?,
                siparis_no=?,
                mal_grubu=?,
                malzeme_tanimi=?,
                planlanan_miktar=?
            WHERE id=?
        """, (
            d["yuklenici_firma"],
            d["siparis_no"],
            d["mal_grubu"],
            d["malzeme_tanimi"],
            d["planlanan_miktar"],
            d["id"]
        ))

        # Değişiklikleri logla
        degisiklikler = []
        if eski:
            if eski["malzeme_tanimi"] != d["malzeme_tanimi"]:
                degisiklikler.append(f"Malzeme: {eski['malzeme_tanimi']} → {d['malzeme_tanimi']}")
            if float(eski["planlanan_miktar"]) != float(d["planlanan_miktar"]):
                degisiklikler.append(f"Miktar: {eski['planlanan_miktar']} → {d['planlanan_miktar']}")
            if (eski["yuklenici_firma"] or "") != d["yuklenici_firma"]:
                degisiklikler.append(f"Yüklenici: {eski['yuklenici_firma'] or '-'} → {d['yuklenici_firma'] or '-'}")
            if (eski["mal_grubu"] or "") != d["mal_grubu"]:
                degisiklikler.append(f"Mal Grubu: {eski['mal_grubu'] or '-'} → {d['mal_grubu'] or '-'}")
            if (eski["siparis_no"] or "") != d["siparis_no"]:
                degisiklikler.append(f"Sipariş: {eski['siparis_no'] or '-'} → {d['siparis_no'] or '-'}")

        detay_metin = "Kalem düzenlendi" + (": " + ", ".join(degisiklikler) if degisiklikler else "")

        conn.execute(
            "INSERT INTO sevkiyat_hareket (kalem_id,islem,miktar,tir_plaka,yapan,yapan_ad,tarih,detay) VALUES (?,?,?,?,?,?,?,?)",
            (d["id"], "duzenleme", 0, "", session["kullanici"], session["ad"],
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"), detay_metin)
        )

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Kalem Düzenleme", detay_metin, d["id"], d["malzeme_tanimi"])
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/sevkiyat/kalem-gonder", methods=["POST"])
@yetki_gerekli("kalem_gonder")
def sevkiyat_kalem_gonder():
    try:
        d = request.json
        kid = d["id"]
        miktar = float(d["miktar"])
        tir_plaka = d.get("tir_plaka","")

        conn = get_db()
        kalem = conn.execute("SELECT * FROM sevkiyat_kalem WHERE id=?", (kid,)).fetchone()

        if not kalem:
            conn.close()
            return jsonify({"durum":"hata","mesaj":"Kalem bulunamadı"}), 404

        if miktar <= 0:
            conn.close()
            return jsonify({"durum":"hata","mesaj":"Miktar 0'dan büyük olmalı."}), 400

        yeni_gonderilen = kalem["gonderilen_miktar"] + miktar
        yeni_durum = "Gönderildi" if yeni_gonderilen >= kalem["planlanan_miktar"] else "Kısmi"

        conn.execute(
            "UPDATE sevkiyat_kalem SET gonderilen_miktar=?, durum=?, tir_plaka=COALESCE(NULLIF(?,''),tir_plaka) WHERE id=?",
            (yeni_gonderilen, yeni_durum, tir_plaka, kid)
        )

        conn.execute(
            "INSERT INTO sevkiyat_hareket (kalem_id,islem,miktar,tir_plaka,yapan,yapan_ad,tarih,detay) VALUES (?,?,?,?,?,?,?,?)",
            (
                kid,
                "gonder",
                miktar,
                tir_plaka,
                session["kullanici"],
                session["ad"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                f"{miktar} {kalem['birim']} gönderildi" + (f" — Plaka: {tir_plaka}" if tir_plaka else "")
            )
        )

        planlanan = kalem["planlanan_miktar"] or 0
        gonderilen_gosterim = min(yeni_gonderilen, planlanan)
        
        kalan = max(planlanan - gonderilen_gosterim, 0)
        fazla = max(yeni_gonderilen - planlanan, 0)

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Gönderim", f"{miktar} {kalem['birim']} gönderildi" + (f" — Plaka: {tir_plaka}" if tir_plaka else ""), kid, kalem["malzeme_tanimi"])

        return jsonify({
            "durum":"ok",
            "yeni_durum":yeni_durum,
            "gonderilen":yeni_gonderilen,
            "kalan":kalan,
            "fazla":fazla
        })
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/sevkiyat/kalem-toplu-devret", methods=["POST"])
@yetki_gerekli("toplu_devret")
def sevkiyat_kalem_toplu_devret():
    try:
        d = request.json
        kalem_ids = d["kalem_ids"]
        hedef = d["hedef_plan_id"]

        conn = get_db()
        adet = 0

        for kid in kalem_ids:
            kalem = conn.execute("SELECT * FROM sevkiyat_kalem WHERE id=?", (kid,)).fetchone()

            if not kalem:
                continue

            kalan = max(kalem["planlanan_miktar"] - kalem["gonderilen_miktar"], 0)

            if kalan <= 0:
                continue

            kaynak = conn.execute(
                "SELECT plan_adi FROM sevkiyat_plan WHERE id=?",
                (kalem["plan_id"],)
            ).fetchone()

            hedef_plan = conn.execute(
                "SELECT plan_adi FROM sevkiyat_plan WHERE id=?",
                (hedef,)
            ).fetchone()

            kaynak_adi = kaynak["plan_adi"] if kaynak else f"Plan #{kalem['plan_id']}"
            hedef_adi = hedef_plan["plan_adi"] if hedef_plan else f"Plan #{hedef}"

            if kalem["gonderilen_miktar"] > 0:
                conn.execute(
                    "UPDATE sevkiyat_kalem SET planlanan_miktar=?, durum='Gönderildi', hedef_plan_id=?, hedef_plan_adi=? WHERE id=?",
                    (kalem["gonderilen_miktar"], hedef, hedef_adi, kid)
                )
            else:
                conn.execute(
                    "UPDATE sevkiyat_kalem SET durum='Devreden', hedef_plan_id=?, hedef_plan_adi=? WHERE id=?",
                    (hedef, hedef_adi, kid)
                )

            conn.execute("""
                INSERT INTO sevkiyat_kalem
                (
                    plan_id,
                    yuklenici_firma,
                    siparis_no,
                    mal_grubu,
                    malzeme_tanimi,
                    planlanan_miktar,
                    birim,
                    devreden_plan_id,
                    devreden_plan_adi
                )
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (
                hedef,
                kalem["yuklenici_firma"],
                kalem["siparis_no"],
                kalem["mal_grubu"],
                kalem["malzeme_tanimi"],
                kalan,
                kalem["birim"],
                kalem["plan_id"],
                kaynak_adi
            ))

            adet += 1

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Toplu Devir", f"{adet} kalem devredildi (Hedef plan: {hedef})", hedef)
        return jsonify({"durum":"ok","adet":adet})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500
    

@app.route("/api/sevkiyat/kalem-sifirla", methods=["POST"])
@yetki_gerekli("kalem_sifirla")
def sevkiyat_kalem_sifirla():
    try:
        kid = request.json["id"]
        conn = get_db()
        kalem = conn.execute("SELECT * FROM sevkiyat_kalem WHERE id=?", (kid,)).fetchone()
        eski_gonderilen = kalem["gonderilen_miktar"] if kalem else 0

        conn.execute("UPDATE sevkiyat_kalem SET gonderilen_miktar=0, durum='Bekliyor', tir_plaka='' WHERE id=?", (kid,))
        conn.execute("DELETE FROM sevkiyat_hareket WHERE kalem_id=?", (kid,))

        conn.execute(
            "INSERT INTO sevkiyat_hareket (kalem_id,islem,miktar,tir_plaka,yapan,yapan_ad,tarih,detay) VALUES (?,?,?,?,?,?,?,?)",
            (kid, "sifirla", eski_gonderilen, "", session["kullanici"], session["ad"],
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             f"Gönderim sıfırlandı (eski: {eski_gonderilen})")
        )

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Gönderim Sıfırlama", f"Kalem #{kid} sıfırlandı (eski: {eski_gonderilen})", kid)
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500
    

@app.route("/api/sevkiyat/kalem-kalan-iptal", methods=["POST"])
@yetki_gerekli("kalem_gonder")
def sevkiyat_kalem_kalan_iptal():
    """Kalan miktarı iptal eder: planlanan_miktar = gonderilen_miktar yaparak kalemi tamamlar"""
    try:
        kid = request.json["id"]
        conn = get_db()
        kalem = conn.execute("SELECT * FROM sevkiyat_kalem WHERE id=?", (kid,)).fetchone()

        if not kalem:
            conn.close()
            return jsonify({"durum":"hata","mesaj":"Kalem bulunamadı"}), 404

        gonderilen = kalem["gonderilen_miktar"] or 0
        planlanan = kalem["planlanan_miktar"] or 0
        kalan = max(planlanan - gonderilen, 0)

        if kalan <= 0:
            conn.close()
            return jsonify({"durum":"hata","mesaj":"Bu kalemde iptal edilecek kalan yok."}), 400

        # planlanan_miktar = gonderilen_miktar yap, durum = Gönderildi
        yeni_planlanan = gonderilen if gonderilen > 0 else 0
        yeni_durum = "Gönderildi" if gonderilen > 0 else "Bekliyor"

        conn.execute(
            "UPDATE sevkiyat_kalem SET planlanan_miktar=?, durum=? WHERE id=?",
            (yeni_planlanan, yeni_durum, kid)
        )

        conn.execute(
            "INSERT INTO sevkiyat_hareket (kalem_id,islem,miktar,tir_plaka,yapan,yapan_ad,tarih,detay) VALUES (?,?,?,?,?,?,?,?)",
            (kid, "kalan_iptal", kalan, "", session["kullanici"], session["ad"],
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             f"Kalan miktar iptal edildi: {kalan} {kalem['birim']} (Planlanan: {planlanan} → {yeni_planlanan})")
        )

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Kalan İptal", f"{kalan} {kalem['birim']} iptal edildi (Planlanan: {planlanan} → {yeni_planlanan})", kid, kalem["malzeme_tanimi"])
        return jsonify({"durum":"ok","iptal_edilen":kalan})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500


@app.route("/api/sevkiyat/toplu-gonder", methods=["POST"])
@yetki_gerekli("kalem_gonder")
def sevkiyat_toplu_gonder():
    """Toplu gönderim: birden fazla kaleme tek seferde gönderim yap"""
    try:
        d = request.json
        tir_plaka = d.get("tir_plaka", "")
        gonderimler = d.get("gonderimler", [])  # [{id, miktar}, ...]

        if not gonderimler:
            return jsonify({"durum":"hata","mesaj":"Gönderilecek kalem yok."}), 400

        conn = get_db()
        basarili = 0
        hatalar = []

        for g in gonderimler:
            kid = g["id"]
            miktar = float(g.get("miktar", 0))

            if miktar <= 0:
                continue

            kalem = conn.execute("SELECT * FROM sevkiyat_kalem WHERE id=?", (kid,)).fetchone()
            if not kalem:
                hatalar.append(f"Kalem #{kid} bulunamadı")
                continue

            yeni_gonderilen = kalem["gonderilen_miktar"] + miktar
            yeni_durum = "Gönderildi" if yeni_gonderilen >= kalem["planlanan_miktar"] else "Kısmi"

            conn.execute(
                "UPDATE sevkiyat_kalem SET gonderilen_miktar=?, durum=?, tir_plaka=COALESCE(NULLIF(?,''),tir_plaka) WHERE id=?",
                (yeni_gonderilen, yeni_durum, tir_plaka, kid)
            )

            conn.execute(
                "INSERT INTO sevkiyat_hareket (kalem_id,islem,miktar,tir_plaka,yapan,yapan_ad,tarih,detay) VALUES (?,?,?,?,?,?,?,?)",
                (kid, "gonder", miktar, tir_plaka, session["kullanici"], session["ad"],
                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                 f"Toplu gönderim: {miktar} {kalem['birim']}" + (f" — Plaka: {tir_plaka}" if tir_plaka else ""))
            )

            basarili += 1

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Toplu Gönderim", f"{basarili} kalem gönderildi" + (f" — Plaka: {tir_plaka}" if tir_plaka else ""))
        return jsonify({"durum":"ok","basarili":basarili,"hatalar":hatalar})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500


@app.route("/api/sevkiyat/toplu-kalan-iptal", methods=["POST"])
@yetki_gerekli("kalem_gonder")
def sevkiyat_toplu_kalan_iptal():
    """Seçilen kalemlerin kalan miktarlarını toplu iptal et"""
    try:
        d = request.json
        kalem_ids = d.get("kalem_ids", [])

        if not kalem_ids:
            return jsonify({"durum":"hata","mesaj":"Kalem seçilmedi."}), 400

        conn = get_db()
        basarili = 0

        for kid in kalem_ids:
            kalem = conn.execute("SELECT * FROM sevkiyat_kalem WHERE id=?", (kid,)).fetchone()
            if not kalem:
                continue

            gonderilen = kalem["gonderilen_miktar"] or 0
            planlanan = kalem["planlanan_miktar"] or 0
            kalan = max(planlanan - gonderilen, 0)

            if kalan <= 0:
                continue

            # planlanan = gönderilen yap, durum = Gönderildi
            yeni_planlanan = gonderilen if gonderilen > 0 else 0
            yeni_durum = "Gönderildi" if gonderilen > 0 else "Bekliyor"

            conn.execute(
                "UPDATE sevkiyat_kalem SET planlanan_miktar=?, durum=? WHERE id=?",
                (yeni_planlanan, yeni_durum, kid)
            )

            conn.execute(
                "INSERT INTO sevkiyat_hareket (kalem_id,islem,miktar,tir_plaka,yapan,yapan_ad,tarih,detay) VALUES (?,?,?,?,?,?,?,?)",
                (kid, "kalan_iptal", kalan, "", session["kullanici"], session["ad"],
                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                 f"Toplu kalan iptal: {kalan} {kalem['birim']} (Planlanan: {planlanan} → {yeni_planlanan})")
            )

            basarili += 1

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Toplu Kalan İptal", f"{basarili} kalemin kalanı iptal edildi")
        return jsonify({"durum":"ok","basarili":basarili})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500


@app.route("/api/sevkiyat/hareket-guncelle", methods=["POST"])
@yetki_gerekli("kalem_guncelle")
def sevkiyat_hareket_guncelle():
    """Bir gönderim hareketinin miktarını düzenle"""
    try:
        d = request.json
        hareket_id = d["hareket_id"]
        yeni_miktar = float(d["yeni_miktar"])

        if yeni_miktar < 0:
            return jsonify({"durum":"hata","mesaj":"Miktar negatif olamaz."}), 400

        conn = get_db()
        hareket = conn.execute("SELECT * FROM sevkiyat_hareket WHERE id=?", (hareket_id,)).fetchone()

        if not hareket:
            conn.close()
            return jsonify({"durum":"hata","mesaj":"Hareket bulunamadı"}), 404

        if hareket["islem"] != "gonder":
            conn.close()
            return jsonify({"durum":"hata","mesaj":"Sadece gönderim hareketleri düzenlenebilir."}), 400

        eski_miktar = hareket["miktar"]
        fark = yeni_miktar - eski_miktar
        kalem_id = hareket["kalem_id"]

        kalem = conn.execute("SELECT * FROM sevkiyat_kalem WHERE id=?", (kalem_id,)).fetchone()
        if not kalem:
            conn.close()
            return jsonify({"durum":"hata","mesaj":"İlgili kalem bulunamadı"}), 404

        # Hareket miktarını güncelle
        conn.execute("UPDATE sevkiyat_hareket SET miktar=? WHERE id=?", (yeni_miktar, hareket_id))

        # Kalem gönderilen miktarını güncelle
        yeni_gonderilen = max(kalem["gonderilen_miktar"] + fark, 0)
        yeni_durum = "Gönderildi" if yeni_gonderilen >= kalem["planlanan_miktar"] else ("Kısmi" if yeni_gonderilen > 0 else "Bekliyor")

        conn.execute(
            "UPDATE sevkiyat_kalem SET gonderilen_miktar=?, durum=? WHERE id=?",
            (yeni_gonderilen, yeni_durum, kalem_id)
        )

        # Log
        conn.execute(
            "INSERT INTO sevkiyat_hareket (kalem_id,islem,miktar,tir_plaka,yapan,yapan_ad,tarih,detay) VALUES (?,?,?,?,?,?,?,?)",
            (kalem_id, "duzenleme", 0, "", session["kullanici"], session["ad"],
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             f"Gönderim düzenlendi: {eski_miktar} → {yeni_miktar} {kalem['birim']} (Hareket #{hareket_id})")
        )

        conn.commit()
        conn.close()
        log_kaydet("Sevkiyat", "Gönderim Düzenleme", f"Hareket #{hareket_id}: {eski_miktar} → {yeni_miktar} {kalem['birim']}", kalem_id, kalem["malzeme_tanimi"])
        return jsonify({"durum":"ok","eski_miktar":eski_miktar,"yeni_miktar":yeni_miktar})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500


# ---- DASHBOARD API ----
@app.route("/api/sevkiyat/dashboard")
@yetki_gerekli("panel_gor")
def sevkiyat_dashboard_api():
    try:
        filtre = request.args.get("filter")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        plan_tipi = request.args.get("plan_tipi")  # plan tipi filtresi
        mal_gruplari = request.args.get("mal_gruplari")  # virgülle ayrılmış mal grupları
        
        today = date.today()
        
        if filtre == "today":
            start = today
            end = today
            
        elif filtre == "week":
            start = today - timedelta(days=today.weekday())
            end = start + timedelta(days=6)
            
        elif filtre == "lastweek":
            this_week_start = today - timedelta(days=today.weekday())
            start = this_week_start - timedelta(days=7)
            end = this_week_start - timedelta(days=1)

        elif filtre == "month":
            start = date(today.year, today.month, 1)
            if today.month == 12:
                end = date(today.year, 12, 31)
            else:
                end = date(today.year, today.month + 1, 1) - timedelta(days=1)

        elif filtre == "lastmonth":
            first_this = date(today.year, today.month, 1)
            end = first_this - timedelta(days=1)
            start = date(end.year, end.month, 1)
            
        elif filtre == "year":
            start = date(today.year, 1, 1)
            end = date(today.year, 12, 31)
            
        elif filtre == "range" and start_date and end_date:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            
        else:
            start = None
            end = None

        conn = get_db()

        if start and end:
            sql = "SELECT * FROM sevkiyat_plan WHERE bitis BETWEEN ? AND ?"
            params = [start, end]
            if plan_tipi:
                sql += " AND plan_tipi=?"
                params.append(plan_tipi)
            sql += " ORDER BY id DESC"
            planlar = conn.execute(sql, params).fetchall()
        else:
            if plan_tipi:
                planlar = conn.execute(
                    "SELECT * FROM sevkiyat_plan WHERE plan_tipi=? ORDER BY id DESC",
                    (plan_tipi,)
                ).fetchall()
            else:
                planlar = conn.execute(
                    "SELECT * FROM sevkiyat_plan ORDER BY id DESC"
                ).fetchall()

        plist = []

        # Mal grubu filtresi hazırla
        mg_filtre_sql = ""
        mg_params = []
        if mal_gruplari:
            mg_list = [m.strip() for m in mal_gruplari.split(",") if m.strip()]
            if mg_list:
                placeholders = ",".join(["?" for _ in mg_list])
                mg_filtre_sql = f" AND mal_grubu IN ({placeholders})"
                mg_params = mg_list

        for p in planlar:
            t = conn.execute(
                "SELECT COUNT(*) as c, SUM(planlanan_miktar) as pm, SUM(gonderilen_miktar) as gm FROM sevkiyat_kalem WHERE plan_id=?" + mg_filtre_sql,
                [p["id"]] + mg_params
            ).fetchone()

            ds = conn.execute("""
                SELECT durum, COUNT(*) as adet, SUM(planlanan_miktar - gonderilen_miktar) as kg
                FROM sevkiyat_kalem
                WHERE plan_id=?""" + mg_filtre_sql + """
                GROUP BY durum
            """, [p["id"]] + mg_params).fetchall()

            ozet = {
                r["durum"]: {
                    "adet": r["adet"],
                    "kg": max(r["kg"] or 0, 0)
                }
                for r in ds
            }

            pm = t["pm"] or 0
            gercek_gm = t["gm"] or 0
            
            gm = min(gercek_gm, pm)
            fazla_gm = max(gercek_gm - pm, 0)

            # Devreden kalemlerin planlanan miktarını düş (artık bu planda değiller)
            devreden_kg = ozet.get("Devreden", {}).get("kg", 0)
            km = max(pm - gm - devreden_kg, 0)

            # Yüzde: devreden hariç planlanan üzerinden hesapla
            efektif_planlanan = max(pm - devreden_kg, 0)
            yuzde = round(min(gm / efektif_planlanan * 100, 100), 1) if efektif_planlanan > 0 else (100 if pm > 0 else 0)

            try:
                kalan_gun = (datetime.strptime(p["bitis"], "%Y-%m-%d").date() - date.today()).days
            except:
                kalan_gun = None

            plist.append({
                "id": p["id"],
                "plan_adi": p["plan_adi"],
                "plan_tipi": p["plan_tipi"],
                "baslangic": p["baslangic"],
                "bitis": p["bitis"],
                "durum": p["durum"],
                "toplam_kalem": t["c"],
                "planlanan_kg": pm,
                "gercek_gonderilen_kg": gercek_gm,
                "fazla_gonderim_kg": fazla_gm,
                "gonderilen_kg": gm,
                "kalan_kg": km,
                "yuzde": yuzde,
                "kalan_gun": kalan_gun,
                "bekleyen": ozet.get("Bekliyor", {}).get("adet", 0),
                "kismi": ozet.get("Kısmi", {}).get("adet", 0),
                "gonderildi": ozet.get("Gönderildi", {}).get("adet", 0),
                "devreden": ozet.get("Devreden", {}).get("adet", 0),
                "ozet": ozet
            })

        bugun = date.today().strftime("%Y-%m-%d")

        geciken = conn.execute("""
            SELECT k.*, p.plan_adi, p.bitis
            FROM sevkiyat_kalem k
            JOIN sevkiyat_plan p ON k.plan_id=p.id
            WHERE p.bitis<? AND k.durum NOT IN ('Gönderildi','Devreden')
            ORDER BY p.bitis
        """, (bugun,)).fetchall()

        # En çok sevk edilen mal grupları - FİLTRELİ
        # Sadece filtrelenmiş planlardaki kalemleri say
        plan_id_list = [p["id"] for p in plist] if plist else []

        if plan_id_list:
            ph = ",".join(["?" for _ in plan_id_list])
            mg_sql = f"""
                SELECT mal_grubu, SUM(gonderilen_miktar) as t, birim
                FROM sevkiyat_kalem
                WHERE gonderilen_miktar>0
                AND mal_grubu IS NOT NULL AND mal_grubu!=''
                AND plan_id IN ({ph})
            """
            mg_sql_params = list(plan_id_list)

            if mg_filtre_sql:
                mg_sql += mg_filtre_sql
                mg_sql_params.extend(mg_params)

            mg_sql += " GROUP BY mal_grubu ORDER BY t DESC LIMIT 15"
            top_mg = conn.execute(mg_sql, mg_sql_params).fetchall()
        else:
            top_mg = []

        # Kategori eşlemelerini al (top_mg'ye kategori adı eklemek için)
        esleme_rows = conn.execute("SELECT * FROM mal_grubu_esleme").fetchall()
        kat_rows = conn.execute("SELECT * FROM ana_mal_grubu").fetchall()
        esleme_mg_dict = {e["mal_grubu"]: e["ana_kategori_id"] for e in esleme_rows}
        kat_id_dict = {k["id"]: {"ad": k["kategori_adi"], "renk": k["renk"]} for k in kat_rows}

        conn.close()

        # Tüm benzersiz mal gruplarını al (filtre seçenekleri için)
        conn2 = get_db()
        tum_mal_gruplari = conn2.execute("""
            SELECT DISTINCT mal_grubu FROM sevkiyat_kalem
            WHERE mal_grubu IS NOT NULL AND mal_grubu != ''
            ORDER BY mal_grubu
        """).fetchall()

        # Kategori eşlemelerini al
        kategoriler = conn2.execute("SELECT * FROM ana_mal_grubu ORDER BY id").fetchall()
        eslesmeler = conn2.execute("SELECT * FROM mal_grubu_esleme").fetchall()
        conn2.close()

        mg_listesi = [m["mal_grubu"] for m in tum_mal_gruplari]
        esleme_dict = {e["mal_grubu"]: e["ana_kategori_id"] for e in eslesmeler}
        kat_listesi = [{"id": k["id"], "kategori_adi": k["kategori_adi"], "renk": k["renk"]} for k in kategoriler]

        toplam_planlanan = sum(p["planlanan_kg"] for p in plist)
        toplam_gonderilen = sum(p["gonderilen_kg"] for p in plist)
        toplam_kalem = sum(p["toplam_kalem"] for p in plist)

        bekleyen_adet = sum(p["bekleyen"] for p in plist)
        devreden_adet = sum(p["devreden"] for p in plist)

        devreden_kg = sum(
            p["ozet"].get("Devreden", {}).get("kg", 0)
            for p in plist
        )

        bekleyen_kg = sum(
            p["ozet"].get("Bekliyor", {}).get("kg", 0) +
            p["ozet"].get("Kısmi", {}).get("kg", 0)
            for p in plist
        )

        # Toplam planlanan = tüm kalemlerin toplamı (devreden dahil, çünkü sevk edilecek)
        efektif_planlanan = toplam_planlanan

        return jsonify({
            "planlar": plist,
            "genel": {
                "toplam_kalem": toplam_kalem,
                "planlanan_kg": efektif_planlanan,
                "gonderilen_kg": toplam_gonderilen,
                "bekleyen_adet": bekleyen_adet,
                "bekleyen_kg": bekleyen_kg,
                "devreden_adet": devreden_adet,
                "devreden_kg": devreden_kg
            },
            "mal_gruplari": mg_listesi,
            "kategoriler": kat_listesi,
            "eslesmeler": esleme_dict,
            "geciken": [{
                "id": g["id"],
                "plan_adi": g["plan_adi"],
                "bitis": g["bitis"],
                "malzeme_tanimi": g["malzeme_tanimi"],
                "planlanan": g["planlanan_miktar"],
                "gonderilen": g["gonderilen_miktar"],
                "kalan": max(g["planlanan_miktar"] - g["gonderilen_miktar"], 0),
                "birim": g["birim"],
                "durum": g["durum"],
                "yuklenici_firma": g["yuklenici_firma"]
            } for g in geciken],
            "top_mal_grubu": [{
                "mal_grubu": t["mal_grubu"],
                "toplam": t["t"],
                "birim": t["birim"],
                "kategori_id": esleme_mg_dict.get(t["mal_grubu"]),
                "kategori_adi": kat_id_dict.get(esleme_mg_dict.get(t["mal_grubu"]), {}).get("ad", ""),
                "kategori_renk": kat_id_dict.get(esleme_mg_dict.get(t["mal_grubu"]), {}).get("renk", "#666")
            } for t in top_mg]
        })

    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500


# ---- DASHBOARD FİLTRELİ EXCEL EXPORT ----
@app.route("/api/sevkiyat/dashboard-excel", methods=["POST"])
@yetki_gerekli("export_excel")
def dashboard_filtreli_excel():
    """Dashboard'daki aktif filtrelere göre tüm kalemlerin Excel çıktısı"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        d = request.json
        plan_tipi = d.get("plan_tipi")
        mal_gruplari = d.get("mal_gruplari", [])
        filtre_adi = d.get("filtre_adi", "Dashboard Export")
        tarih_filtre = d.get("tarih_filtre")
        start_date_str = d.get("start_date")
        end_date_str = d.get("end_date")

        # Tarih aralığını hesapla (dashboard ile aynı mantık)
        today = date.today()
        start = None
        end = None

        if tarih_filtre == "today":
            start = today; end = today
        elif tarih_filtre == "week":
            start = today - timedelta(days=today.weekday()); end = start + timedelta(days=6)
        elif tarih_filtre == "lastweek":
            this_week_start = today - timedelta(days=today.weekday())
            start = this_week_start - timedelta(days=7); end = this_week_start - timedelta(days=1)
        elif tarih_filtre == "month":
            start = date(today.year, today.month, 1)
            end = date(today.year, today.month + 1, 1) - timedelta(days=1) if today.month < 12 else date(today.year, 12, 31)
        elif tarih_filtre == "lastmonth":
            first_this = date(today.year, today.month, 1)
            end = first_this - timedelta(days=1)
            start = date(end.year, end.month, 1)
        elif tarih_filtre == "year":
            start = date(today.year, 1, 1); end = date(today.year, 12, 31)
        elif tarih_filtre == "range" and start_date_str and end_date_str:
            start = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end = datetime.strptime(end_date_str, "%Y-%m-%d").date()

        conn = get_db()

        # Planları filtrele (tarih + tip)
        sql_plan = "SELECT * FROM sevkiyat_plan WHERE 1=1"
        plan_params = []

        if plan_tipi:
            sql_plan += " AND plan_tipi=?"
            plan_params.append(plan_tipi)

        if start and end:
            sql_plan += " AND bitis BETWEEN ? AND ?"
            plan_params.append(str(start))
            plan_params.append(str(end))

        sql_plan += " ORDER BY id DESC"
        planlar = conn.execute(sql_plan, plan_params).fetchall()

        # Tüm kalemleri topla
        tum_kalemler = []
        for p in planlar:
            sql = "SELECT * FROM sevkiyat_kalem WHERE plan_id=?"
            params = [p["id"]]

            if mal_gruplari:
                placeholders = ",".join(["?" for _ in mal_gruplari])
                sql += f" AND mal_grubu IN ({placeholders})"
                params.extend(mal_gruplari)

            sql += " ORDER BY id"
            kalemler = conn.execute(sql, params).fetchall()

            for k in kalemler:
                tum_kalemler.append({
                    "plan_adi": p["plan_adi"],
                    "plan_tipi": p["plan_tipi"],
                    "yuklenici_firma": k["yuklenici_firma"] or "",
                    "siparis_no": k["siparis_no"] or "",
                    "mal_grubu": k["mal_grubu"] or "",
                    "malzeme_tanimi": k["malzeme_tanimi"] or "",
                    "planlanan": k["planlanan_miktar"] or 0,
                    "gonderilen": k["gonderilen_miktar"] or 0,
                    "kalan": max((k["planlanan_miktar"] or 0) - min(k["gonderilen_miktar"] or 0, k["planlanan_miktar"] or 0), 0),
                    "birim": k["birim"] or "KG",
                    "durum": k["durum"] or ""
                })

        conn.close()

        if not tum_kalemler:
            return jsonify({"durum":"hata","mesaj":"Filtreye uygun kalem bulunamadı."}), 400

        # Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Export"
        ws.sheet_view.showGridLines = False

        L = "012946"; B = "FFFFFF"; G = "F4F6F8"; S = "DDDDDD"
        def kn(r=S): s = Side(style="thin", color=r); return Border(left=s, right=s, top=s, bottom=s)
        def dl(h): return PatternFill("solid", fgColor=h)

        # Başlık
        ws.row_dimensions[1].height = 30
        ws.merge_cells("A1:I1")
        ws.merge_cells("J1:L1")
        for c in "ABCDEFGHIJKL":
            ws[f"{c}1"].fill = dl(L); ws[f"{c}1"].border = kn(L)
        ws["A1"].value = filtre_adi
        ws["A1"].font = Font(bold=True, size=13, color=B, name="Arial")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws["J1"].value = "Tarih: " + datetime.now().strftime("%d.%m.%Y")
        ws["J1"].font = Font(size=10, color=B, name="Arial")
        ws["J1"].alignment = Alignment(horizontal="right", vertical="center")

        # Kolon başlıkları
        basliklar = ["S.No", "Plan", "Yüklenici", "Sipariş No", "Mal Grubu", "Malzeme Tanımı", "Planlanan", "Gönderilen", "Kalan", "Birim", "Durum", "Plan Tipi"]
        ws.row_dimensions[2].height = 6
        ws.row_dimensions[3].height = 22
        for i, b in enumerate(basliklar, 1):
            h = ws.cell(row=3, column=i, value=b)
            h.font = Font(bold=True, size=10, color=B, name="Arial")
            h.fill = dl(L)
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.border = kn(L)

        # Veriler
        for idx, k in enumerate(tum_kalemler):
            r = 4 + idx
            bg = B if idx % 2 == 0 else G
            vals = [idx+1, k["plan_adi"], k["yuklenici_firma"], k["siparis_no"], k["mal_grubu"],
                    k["malzeme_tanimi"], k["planlanan"], k["gonderilen"], k["kalan"], k["birim"], k["durum"], k["plan_tipi"]]
            for col, val in enumerate(vals, 1):
                h = ws.cell(row=r, column=col, value=val)
                h.font = Font(size=9, name="Arial")
                h.fill = dl(bg)
                h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                h.border = kn()

        # Toplamlar
        sr = 4 + len(tum_kalemler) + 1
        tp_ham = sum(k["planlanan"] for k in tum_kalemler)
        tg = sum(min(k["gonderilen"], k["planlanan"]) for k in tum_kalemler)
        td = sum(k["planlanan"] for k in tum_kalemler if k["durum"] == "Devreden")
        tp = max(tp_ham - td, 0)
        tk = max(tp - tg, 0)

        for off, et, dg in [(0, "Toplam Planlanan", f"{tp:,.2f}"), (1, "Toplam Gönderilen", f"{tg:,.2f}"), (2, "Toplam Kalan", f"{tk:,.2f}")]:
            r = sr + off
            ws.merge_cells(f"A{r}:I{r}")
            for c in range(1, 13):
                h = ws.cell(row=r, column=c); h.fill = dl(L); h.border = kn(L)
            ws[f"A{r}"].value = et
            ws[f"A{r}"].font = Font(bold=True, size=10, color=B, name="Arial")
            ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
            ws[f"J{r}"].value = dg
            ws[f"J{r}"].font = Font(bold=True, size=10, color=B, name="Arial")
            ws[f"J{r}"].alignment = Alignment(horizontal="right", vertical="center")

        # Kolon genişlikleri
        widths = {"A":6,"B":25,"C":18,"D":30,"E":16,"F":40,"G":14,"H":14,"I":14,"J":8,"K":12,"L":14}
        for c, w in widths.items():
            ws.column_dimensions[c].width = w

        klasor = os.path.join(get_export_path(), "exports", "excel")
        os.makedirs(klasor, exist_ok=True)
        ad = "dashboard_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx"
        yol = os.path.join(klasor, ad)
        wb.save(yol)
        try: os.startfile(klasor)
        except: pass
        return jsonify({"durum":"ok","yol":yol,"kalem_sayisi":len(tum_kalemler)})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500
### ============================================================
### NAKIL DASHBOARD - PDF EXPORT ROUTE
### Bu kodu app.py'deki nakil-dashboard bölümüne ekle
### ============================================================

@app.route('/api/nakil-dashboard/export-pdf', methods=['POST'])
def nakil_dashboard_export_pdf():
    """Nakil dashboard verilerini PDF olarak export eder."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from collections import defaultdict
        import json, os
        from datetime import datetime

        data = request.get_json()
        nak = data.get('nak', [])
        if not nak:
            return jsonify({'durum': 'hata', 'mesaj': 'Veri yok'})

        # Font ayarı (Türkçe karakter desteği)
        font_path = os.path.join(os.path.dirname(__file__), 'assets', 'fonts', 'DejaVuSans.ttf')
        font_path_bold = os.path.join(os.path.dirname(__file__), 'assets', 'fonts', 'DejaVuSans-Bold.ttf')
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont('TR', font_path))
            pdfmetrics.registerFont(TTFont('TRB', font_path_bold))
            FONT = 'TR'
            FONTB = 'TRB'
        else:
            FONT = 'Helvetica'
            FONTB = 'Helvetica-Bold'

        # Çıktı dosyası
        tarih = datetime.now().strftime('%Y%m%d_%H%M%S')
        dosya_adi = f'nakil_rapor_{tarih}.pdf'
        dosya_yolu = os.path.join(os.path.dirname(__file__), 'exports', dosya_adi)
        os.makedirs(os.path.dirname(dosya_yolu), exist_ok=True)

        doc = SimpleDocTemplate(
            dosya_yolu,
            pagesize=landscape(A4),
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm
        )

        # Stiller
        styles = getSampleStyleSheet()
        s_title = ParagraphStyle('Baslik', fontName=FONTB, fontSize=16, alignment=TA_CENTER, spaceAfter=8)
        s_subtitle = ParagraphStyle('AltBaslik', fontName=FONT, fontSize=9, alignment=TA_CENTER, textColor=colors.grey, spaceAfter=14)
        s_section = ParagraphStyle('Bolum', fontName=FONTB, fontSize=12, spaceBefore=14, spaceAfter=6, textColor=colors.HexColor('#2c3e50'))
        s_info = ParagraphStyle('Bilgi', fontName=FONT, fontSize=8, textColor=colors.grey)
        s_cell = ParagraphStyle('Hucre', fontName=FONT, fontSize=7, leading=9)
        s_cell_b = ParagraphStyle('HucreB', fontName=FONTB, fontSize=7, leading=9)
        s_cell_r = ParagraphStyle('HucreR', fontName=FONT, fontSize=7, leading=9, alignment=TA_RIGHT)

        elements = []

        # --- YARDIMCI FONKSİYONLAR ---
        def fmt_sayi(v):
            """Sayı formatla: 1.234,56"""
            try:
                return f"{v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except:
                return str(v)

        def fmt_kg(v):
            """KG formatla: ton/kg"""
            if v >= 10000:
                return f"{v/1000:,.1f} t".replace(',', 'X').replace('.', ',').replace('X', '.')
            return f"{v:,.0f} KG".replace(',', '.')

        def fmt_tl(v):
            """TL formatla"""
            if v >= 1_000_000:
                return f"{v/1_000_000:,.1f} M".replace(',', 'X').replace('.', ',').replace('X', '.')
            if v >= 1000:
                return f"{v/1000:,.0f} K".replace(',', '.')
            return f"{v:,.0f}".replace(',', '.')

        def tutar_pb_str(rows):
            """Para birimi bazlı toplam değer string"""
            pb_toplam = defaultdict(float)
            for r in rows:
                pb = r.get('paraBirimi', '').strip()
                t = r.get('tutar', 0)
                if t and pb:
                    pb_toplam[pb] += t
            sembol = {'TRY': '₺', 'USD': '$', 'EUR': '€', 'RUB': '₽'}
            parts = []
            for pb in sorted(pb_toplam.keys(), key=lambda x: -pb_toplam[x]):
                s = sembol.get(pb, pb)
                parts.append(f"{fmt_tl(pb_toplam[pb])} {s}")
            return ' | '.join(parts) if parts else '-'

        def ref_say(rows):
            """Benzersiz irsaliye/fiş sayısı"""
            keys = set()
            for r in rows:
                ref = (r.get('referans') or '').strip()
                bn = r.get('belgeNo', '')
                keys.add(ref if ref else bn)
            keys.discard('')
            return len(keys)

        def birim_ozet(rows):
            """Birim bazlı toplam"""
            bg = defaultdict(float)
            for r in rows:
                b = r.get('birim', '')
                if b:
                    bg[b] += r.get('miktar', 0)
            parts = []
            for b, t in sorted(bg.items(), key=lambda x: -x[1]):
                if b == 'KG':
                    parts.append(fmt_kg(t))
                else:
                    parts.append(f"{fmt_sayi(t)} {b}")
            return ' | '.join(parts) if parts else '-'

        def grup(rows, key):
            """Veriyi grupla"""
            g = defaultdict(list)
            for r in rows:
                k = r.get(key, '(Boş)') or '(Boş)'
                g[k].append(r)
            return g

        def make_table(header, data_rows, col_widths=None):
            """Tablo oluştur"""
            tbl_data = [header] + data_rows
            tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), FONTB),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), FONT),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#dee2e6')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]))
            return tbl

        # --- İŞLEM TÜRLERİ ---
        IT = {
            'A67': 'Fason Nakil', 'A69': 'Depo Nakil', '221': 'Mal Çıkışı',
            '101': 'Mal Girişi', 'A19': 'HM Satış', 'A21': 'Fason Satış',
            'A09': 'Ticari BÜ Satış', 'A51': 'BÜ Satış',
            'A01': 'İdari Mazot', 'A07': 'Mek. Mazot'
        }

        # ============================================================
        # SAYFA 1: GENEL BAKIŞ
        # ============================================================
        elements.append(Paragraph('BBA Malzeme Nakil Raporu', s_title))
        elements.append(Paragraph(f'Oluşturma: {datetime.now().strftime("%d.%m.%Y %H:%M")} | {len(nak)} hareket | {ref_say(nak)} irsaliye/fiş', s_subtitle))

        # Genel KPI tablosu
        islem_grp = grup(nak, 'islemTuru')
        kpi_header = ['İşlem Türü', 'Açıklama', 'İrsaliye/Fiş', 'Hareket', 'Miktar']
        kpi_rows = []
        for it_code in sorted(islem_grp.keys()):
            rows = islem_grp[it_code]
            kpi_rows.append([
                it_code,
                IT.get(it_code, it_code),
                str(ref_say(rows)),
                str(len(rows)),
                birim_ozet(rows)
            ])
        kpi_rows.append(['TOPLAM', '', str(ref_say(nak)), str(len(nak)), birim_ozet(nak)])
        elements.append(Paragraph('İşlem Türü Özeti', s_section))
        elements.append(make_table(kpi_header, kpi_rows, col_widths=[40, 80, 60, 50, 150]))
        elements.append(Spacer(1, 10))

        # Kullanıcı özeti
        usr_grp = grup(nak, 'kullanici')
        usr_header = ['Kullanıcı', 'İrsaliye/Fiş', 'Hareket', 'Miktar']
        usr_rows = []
        for u in sorted(usr_grp.keys(), key=lambda x: -len(usr_grp[x])):
            rows = usr_grp[u]
            usr_rows.append([u, str(ref_say(rows)), str(len(rows)), birim_ozet(rows)])
        elements.append(Paragraph('Kullanıcı Özeti', s_section))
        elements.append(make_table(usr_header, usr_rows, col_widths=[80, 60, 50, 200]))

        # ============================================================
        # SAYFA 2: A67 FASON
        # ============================================================
        a67 = [r for r in nak if r.get('islemTuru') == 'A67']
        if a67:
            elements.append(PageBreak())
            elements.append(Paragraph('A67 Fason Nakil', s_section))
            elements.append(Paragraph(f'{ref_say(a67)} irsaliye/fiş | {birim_ozet(a67)}', s_info))
            elements.append(Spacer(1, 6))
            fg = grup(a67, 'firma')
            header = ['Firma', 'İrsaliye/Fiş', 'Hareket', 'Miktar', 'Malzeme']
            rows_t = []
            for f in sorted(fg.keys(), key=lambda x: -sum(r['miktar'] for r in fg[x])):
                rs = fg[f]
                mlz = len(set(r['malzemeKodu'] for r in rs if r.get('malzemeKodu')))
                rows_t.append([f, str(ref_say(rs)), str(len(rs)), birim_ozet(rs), str(mlz)])
            elements.append(make_table(header, rows_t, col_widths=[120, 60, 50, 150, 40]))

        # ============================================================
        # SAYFA 3: A69 DEPO
        # ============================================================
        a69 = [r for r in nak if r.get('islemTuru') == 'A69']
        if a69:
            elements.append(PageBreak())
            elements.append(Paragraph('A69 Depo Nakil', s_section))
            elements.append(Paragraph(f'{ref_say(a69)} irsaliye/fiş | {birim_ozet(a69)}', s_info))
            elements.append(Spacer(1, 6))
            dg = grup(a69, 'kaynakDepo')
            header = ['Kaynak Depo', 'İrsaliye/Fiş', 'Hareket', 'Miktar', 'Malzeme']
            rows_t = []
            for d in sorted(dg.keys(), key=lambda x: -sum(r['miktar'] for r in dg[x])):
                rs = dg[d]
                mlz = len(set(r['malzemeKodu'] for r in rs if r.get('malzemeKodu')))
                rows_t.append([d, str(ref_say(rs)), str(len(rs)), birim_ozet(rs), str(mlz)])
            elements.append(make_table(header, rows_t, col_widths=[80, 60, 50, 150, 40]))

        # ============================================================
        # SAYFA 4: 221 MAL ÇIKIŞI
        # ============================================================
        m221 = [r for r in nak if r.get('islemTuru') == '221']
        if m221:
            elements.append(PageBreak())
            elements.append(Paragraph('221 Mal Çıkışı', s_section))
            elements.append(Paragraph(f'{ref_say(m221)} irsaliye/fiş | {birim_ozet(m221)}', s_info))
            elements.append(Spacer(1, 6))
            dg = grup(m221, 'kaynakDepo')
            header = ['Masraf Yeri (Depo)', 'İrsaliye/Fiş', 'Hareket', 'Miktar', 'Malzeme']
            rows_t = []
            for d in sorted(dg.keys(), key=lambda x: -sum(r['miktar'] for r in dg[x])):
                rs = dg[d]
                mlz = len(set(r['malzemeKodu'] for r in rs if r.get('malzemeKodu')))
                rows_t.append([d, str(ref_say(rs)), str(len(rs)), birim_ozet(rs), str(mlz)])
            elements.append(make_table(header, rows_t, col_widths=[80, 60, 50, 150, 40]))

        # ============================================================
        # SAYFA 5: 101 MAL GİRİŞİ + PARASAL DEĞER
        # ============================================================
        m101 = [r for r in nak if r.get('islemTuru') == '101']
        if m101:
            elements.append(PageBreak())
            elements.append(Paragraph('101 Mal Girişi', s_section))
            elements.append(Paragraph(f'{ref_say(m101)} irsaliye/fiş | {birim_ozet(m101)} | {tutar_pb_str(m101)}', s_info))
            elements.append(Spacer(1, 6))

            # Depo bazlı tonaj + para birimi bazlı parasal değer
            dg = grup(m101, 'girisDepo')
            header = ['Giriş Deposu', 'İrsaliye/Fiş', 'Miktar', 'Parasal Değer', 'Malzeme']
            rows_t = []
            for d in sorted(dg.keys(), key=lambda x: -sum(r.get('tutar', 0) for r in dg[x])):
                rs = dg[d]
                mlz = len(set(r['malzemeKodu'] for r in rs if r.get('malzemeKodu')))
                rows_t.append([d, str(ref_say(rs)), birim_ozet(rs), tutar_pb_str(rs), str(mlz)])
            elements.append(make_table(header, rows_t, col_widths=[80, 60, 120, 100, 40]))

        # ============================================================
        # SAYFA 6: A19 HM SATIŞ
        # ============================================================
        for it_code, it_ad in [('A19', 'HM Satış'), ('A21', 'Fason Satış'), ('A09', 'Ticari BÜ Satış')]:
            it_data = [r for r in nak if r.get('islemTuru') == it_code]
            if it_data:
                elements.append(PageBreak())
                elements.append(Paragraph(f'{it_code} {it_ad}', s_section))
                elements.append(Paragraph(f'{ref_say(it_data)} irsaliye/fiş | {birim_ozet(it_data)}', s_info))
                elements.append(Spacer(1, 6))
                fg = grup(it_data, 'firma')
                header = ['Firma', 'İrsaliye/Fiş', 'Hareket', 'Miktar', 'Malzeme']
                rows_t = []
                for f in sorted(fg.keys(), key=lambda x: -sum(r['miktar'] for r in fg[x])):
                    rs = fg[f]
                    mlz = len(set(r['malzemeKodu'] for r in rs if r.get('malzemeKodu')))
                    rows_t.append([f, str(ref_say(rs)), str(len(rs)), birim_ozet(rs), str(mlz)])
                elements.append(make_table(header, rows_t, col_widths=[120, 60, 50, 150, 40]))

        # ============================================================
        # SAYFA 7: A51 BÜ SATIŞ (mal grubu bazlı)
        # ============================================================
        a51 = [r for r in nak if r.get('islemTuru') == 'A51']
        if a51:
            elements.append(PageBreak())
            elements.append(Paragraph('A51 Bitmiş Ürün Satış', s_section))
            elements.append(Paragraph(f'{ref_say(a51)} irsaliye/fiş | {birim_ozet(a51)}', s_info))
            elements.append(Spacer(1, 6))
            mg = grup(a51, 'malGrubuTanimi')
            header = ['Mal Grubu', 'İrsaliye/Fiş', 'Hareket', 'Miktar', 'Malzeme']
            rows_t = []
            for m in sorted(mg.keys(), key=lambda x: -sum(r['miktar'] for r in mg[x])):
                rs = mg[m]
                mlz = len(set(r['malzemeKodu'] for r in rs if r.get('malzemeKodu')))
                rows_t.append([m, str(ref_say(rs)), str(len(rs)), birim_ozet(rs), str(mlz)])
            elements.append(make_table(header, rows_t, col_widths=[120, 60, 50, 150, 40]))

        # ============================================================
        # SAYFA 8: MAZOT
        # ============================================================
        mazot = [r for r in nak if r.get('islemTuru') in ('A01', 'A07')]
        if mazot:
            elements.append(PageBreak())
            a01 = [r for r in mazot if r['islemTuru'] == 'A01']
            a07 = [r for r in mazot if r['islemTuru'] == 'A07']
            elements.append(Paragraph('Mazot Tüketimi (A01 + A07)', s_section))
            elements.append(Paragraph(f'A01 İdari: {ref_say(a01)} fiş | A07 Mekanik: {ref_say(a07)} fiş | Toplam: {birim_ozet(mazot)}', s_info))
            elements.append(Spacer(1, 6))
            header = ['İşlem', 'Açıklama', 'İrsaliye/Fiş', 'Miktar']
            rows_t = []
            if a01:
                rows_t.append(['A01', 'İdari Mazot', str(ref_say(a01)), birim_ozet(a01)])
            if a07:
                rows_t.append(['A07', 'Mekanik Mazot', str(ref_say(a07)), birim_ozet(a07)])
            elements.append(make_table(header, rows_t, col_widths=[40, 80, 60, 150]))

        # ============================================================
        # SAYFA 9: MAL GRUBU ÖZET
        # ============================================================
        elements.append(PageBreak())
        elements.append(Paragraph('Mal Grubu Özeti', s_section))
        mg_all = grup(nak, 'malGrubuTanimi')
        header = ['Mal Grubu', 'Top.', 'A67', 'A69', '221', '101', 'A19', 'A21', 'A09', 'A51', 'Miktar', 'Mlz']
        rows_t = []
        for mg_name in sorted(mg_all.keys(), key=lambda x: -sum(r['miktar'] for r in mg_all[x])):
            if not mg_name or mg_name == '(Boş)':
                continue
            rs = mg_all[mg_name]
            mlz = len(set(r['malzemeKodu'] for r in rs if r.get('malzemeKodu')))
            row = [
                mg_name,
                str(ref_say(rs)),
                str(ref_say([r for r in rs if r['islemTuru'] == 'A67'])),
                str(ref_say([r for r in rs if r['islemTuru'] == 'A69'])),
                str(ref_say([r for r in rs if r['islemTuru'] == '221'])),
                str(ref_say([r for r in rs if r['islemTuru'] == '101'])),
                str(ref_say([r for r in rs if r['islemTuru'] == 'A19'])),
                str(ref_say([r for r in rs if r['islemTuru'] == 'A21'])),
                str(ref_say([r for r in rs if r['islemTuru'] == 'A09'])),
                str(ref_say([r for r in rs if r['islemTuru'] == 'A51'])),
                birim_ozet(rs),
                str(mlz)
            ]
            rows_t.append(row)
        elements.append(make_table(header, rows_t, col_widths=[90, 30, 30, 30, 30, 30, 30, 30, 30, 30, 80, 30]))

        # PDF OLUŞTUR
        doc.build(elements)

        return jsonify({
            'durum': 'ok',
            'dosya': dosya_yolu,
            'kayit': len(nak)
        })

    except ImportError:
        return jsonify({'durum': 'hata', 'mesaj': 'reportlab kütüphanesi yüklü değil. pip install reportlab'})
    except Exception as e:
        return jsonify({'durum': 'hata', 'mesaj': str(e)})
 

# ---- SEVKİYAT PDF (Türkçe destekli) ----
@app.route("/kaydet/sevkiyat-pdf", methods=["POST"])
@yetki_gerekli("export_pdf")
def kaydet_sevkiyat_pdf():

    if not giris_yapildi_mi():
        return jsonify({"durum": "hata", "mesaj": "Yetkisiz"}), 401

    try:
        from fpdf import FPDF

        d = request.json
        plan_adi = d.get("plan_adi", "Sevkiyat")
        tarih_arali = d.get("tarih_arali", "")
        kalemler = d["kalemler"]
        tarih = datetime.now().strftime("%d.%m.%Y")

        def kisalt(text, limit=24):
            text = str(text or "").strip()
            return text if len(text) <= limit else text[:limit - 3] + "..."

        def iki_satira_sigdir(pdf, text, col_width, max_lines=2):
            text = str(text or "").replace("\r", "").strip()
            if not text:
                return "-"

            kelimeler = text.split()
            satirlar = []
            mevcut = ""

            for kelime in kelimeler:
                deneme = (mevcut + " " + kelime).strip()

                if pdf.get_string_width(deneme) <= (col_width - 2):
                    mevcut = deneme
                else:
                    if mevcut:
                        satirlar.append(mevcut)
                    mevcut = kelime

            if mevcut:
                satirlar.append(mevcut)

            if len(satirlar) <= max_lines:
                return "\n".join(satirlar)

            kisaltilmis = satirlar[:max_lines]
            son = kisaltilmis[-1]

            while pdf.get_string_width(son + "...") > (col_width - 2) and len(son) > 1:
                son = son[:-1]

            kisaltilmis[-1] = son + "..."
            return "\n".join(kisaltilmis)

        def satir_sayisi_hesapla(text):
            text = str(text or "")
            return max(len(text.split("\n")), 1)

        basliklar = [
            "S.No",
            "Yüklenici",
            "Sipariş No",
            "Mal Grubu",
            "Malzeme Tanımı",
            "Planlanan",
            "Gönderilen",
            "Kalan",
            "Birim",
            "Durum",
            "Devir"
        ]

        genislikler = [8, 22, 28, 14, 62, 16, 16, 16, 9, 14, 28]

        class SevkiyatPDF(FPDF):
            def header(self):
                self.set_font(self.fn, "B", 14)
                self.cell(0, 10, self.plan_adi, 0, 1, "C")

                self.set_font(self.fn, "", 8)
                self.cell(0, 5, f"{self.tarih_arali}     Tarih: {self.tarih}", 0, 1, "C")

                self.ln(4)

                self.set_fill_color(1, 41, 70)
                self.set_text_color(255, 255, 255)
                self.set_font(self.fn, "B", 7)

                for i, b in enumerate(self.basliklar):
                    self.cell(self.genislikler[i], 7, b, 1, 0, "C", True)

                self.ln()

                self.set_text_color(0, 0, 0)
                self.set_font(self.fn, "", 7.2)

            def footer(self):
                self.set_y(-10)
                self.set_font(self.fn, "", 7)
                self.cell(0, 5, f"Sayfa {self.page_no()}", 0, 0, "C")

        pdf = SevkiyatPDF(orientation="L", unit="mm", format="A4")

        font_path = os.path.join(get_export_path(), "assets", "fonts", "DejaVuSans.ttf")
        font_path_bold = os.path.join(get_export_path(), "assets", "fonts", "DejaVuSans-Bold.ttf")

        if os.path.exists(font_path) and os.path.exists(font_path_bold):
            pdf.add_font("DejaVu", "", font_path, uni=True)
            pdf.add_font("DejaVu", "B", font_path_bold, uni=True)
            fn = "DejaVu"
        else:
            fn = "Helvetica"

        pdf.fn = fn
        pdf.plan_adi = plan_adi
        pdf.tarih_arali = tarih_arali
        pdf.tarih = tarih
        pdf.basliklar = basliklar
        pdf.genislikler = genislikler

        pdf.add_page()

        # TOPLAMLAR
        tp_ham = sum((k.get("planlanan_miktar", 0) or 0) for k in kalemler)
        tg = sum(min((k.get("gonderilen_miktar", 0) or 0), (k.get("planlanan_miktar", 0) or 0)) for k in kalemler)
        tf = sum(max((k.get("gonderilen_miktar", 0) or 0) - (k.get("planlanan_miktar", 0) or 0), 0) for k in kalemler)
        td = sum((k.get("planlanan_miktar", 0) or 0) for k in kalemler if k.get("durum") == "Devreden")
        tp = max(tp_ham - td, 0)
        tk = max(tp - tg, 0)

        for idx, k in enumerate(kalemler):

            if pdf.get_y() > 175:
                pdf.add_page()

            if idx % 2 == 0:
                pdf.set_fill_color(244, 246, 248)
            else:
                pdf.set_fill_color(255, 255, 255)

            planlanan = k.get("planlanan_miktar", 0) or 0
            gercek_gonderilen = k.get("gonderilen_miktar", 0) or 0
            gonderilen_gosterim = min(gercek_gonderilen, planlanan)
            kalan = max(planlanan - gonderilen_gosterim, 0)

            devreden = k.get("devreden_plan_adi", "") or "-"
            malzeme = k.get("malzeme_tanimi", "") or ""

            malzeme = malzeme.replace("( planından devir )", "")
            malzeme = malzeme.replace("(planından devir)", "")
            malzeme = malzeme.replace("planından devir", "")
            malzeme = malzeme.replace("( )", "")
            malzeme = malzeme.replace("()", "")
            malzeme = malzeme.strip()

            devreden_plan = k.get("devreden_plan_adi", "") or ""
            if devreden_plan and devreden_plan in malzeme:
                malzeme = malzeme.replace(devreden_plan, "").strip()

            malzeme_goster = iki_satira_sigdir(pdf, malzeme, genislikler[4], 2)
            devreden_goster = iki_satira_sigdir(pdf, devreden_plan if devreden_plan else "-", genislikler[10], 2)
            siparis_goster = iki_satira_sigdir(pdf, k.get("siparis_no", "") or "-", genislikler[2], 2)

            vals = [
                str(idx + 1),
                kisalt(k.get("yuklenici_firma", "") or "-", 20),
                siparis_goster,
                kisalt(k.get("mal_grubu", "") or "-", 12),
                malzeme_goster,
                f"{planlanan:,.2f}",
                f"{gonderilen_gosterim:,.2f}",
                f"{kalan:,.2f}",
                k.get("birim", "KG"),
                kisalt(k.get("durum", ""), 12),
                devreden_goster
            ]

            # Satır yüksekliğini hesapla - multi-cell kolonlara göre
            line_height = 5
            max_lines = 1
            for v in vals:
                lines = satir_sayisi_hesapla(v)
                if lines > max_lines:
                    max_lines = lines

            row_height = max(line_height * max_lines, 7)
            y_start = pdf.get_y()
            x_start = pdf.get_x()

            # Önce tüm hücrelerin arka planını ve border'ını çiz
            for i, v in enumerate(vals):
                pdf.set_xy(x_start + sum(genislikler[:i]), y_start)
                pdf.cell(genislikler[i], row_height, "", 1, 0, "C", True)

            # Sonra metinleri yaz
            for i, v in enumerate(vals):
                x_pos = x_start + sum(genislikler[:i])

                if i in (2, 4, 10):
                    # Multi-line: sipariş no, malzeme tanımı, devir
                    pdf.set_xy(x_pos, y_start)
                    pdf.multi_cell(genislikler[i], line_height, str(v), 0, "L")
                else:
                    # Tek satır
                    hizalama = "L" if i == 1 else "C"
                    y_offset = (row_height - line_height) / 2
                    pdf.set_xy(x_pos, y_start + y_offset)
                    pdf.cell(genislikler[i], line_height, str(v), 0, 0, hizalama)

            pdf.set_xy(x_start, y_start + row_height)

        pdf.ln(4)

        pdf.set_font(fn, "B", 9)
        pdf.set_fill_color(1, 41, 70)
        pdf.set_text_color(255, 255, 255)

        kutu_w = 68
        bosluk = 4
        kutu_h = 8

        pdf.cell(kutu_w, kutu_h, f"Toplam Planlanan: {tp:,.2f} KG", 1, 0, "C", True)
        pdf.cell(bosluk, kutu_h, "", 0, 0)
        pdf.cell(kutu_w, kutu_h, f"Toplam Gönderilen: {tg:,.2f} KG", 1, 0, "C", True)
        pdf.cell(bosluk, kutu_h, "", 0, 0)
        pdf.cell(kutu_w, kutu_h, f"Toplam Fazla Gönderim: {tf:,.2f} KG", 1, 0, "C", True)
        pdf.cell(bosluk, kutu_h, "", 0, 1)

        pdf.ln(2)
        pdf.cell(kutu_w, kutu_h, f"Toplam Kalan: {tk:,.2f} KG", 1, 1, "C", True)

        klasor = os.path.join(get_export_path(), "exports", "pdf")
        os.makedirs(klasor, exist_ok=True)

        ad = "sevkiyat_" + datetime.now().strftime("%Y%m%d_%H%M") + ".pdf"
        yol = os.path.join(klasor, ad)

        pdf.output(yol)

        try:
            os.startfile(klasor)
        except:
            pass

        return jsonify({"durum": "ok", "yol": yol})

    except Exception as e:
        return jsonify({
            "durum": "hata",
            "mesaj": str(e)
        }), 500
    

# ---- SEVKİYAT EXCEL KAYDET ----
@app.route("/kaydet/sevkiyat-excel", methods=["POST"])
@yetki_gerekli("export_excel")
def kaydet_sevkiyat_excel():
    if not giris_yapildi_mi(): return jsonify({"durum":"hata","mesaj":"Yetkisiz"}), 401
    try:
        def kisalt_excel(text, limit=30):
            text = str(text or "")
            return text if len(text) <= limit else text[:limit-3] + "..."

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        d=request.json; plan_adi=d.get("plan_adi","Sevkiyat"); kalemler=d["kalemler"]; tarih=datetime.now().strftime("%d.%m.%Y")
        wb=Workbook();ws=wb.active;ws.title="Sevkiyat";ws.sheet_view.showGridLines=False
        L="012946";B="FFFFFF";G="F4F6F8";S="DDDDDD"
        def kn(r=S): s=Side(style="thin",color=r); return Border(left=s,right=s,top=s,bottom=s)
        def dl(h): return PatternFill("solid",fgColor=h)
        for c,w in {"A":8,"B":18,"C":30,"D":16,"E":40,"F":14,"G":14,"H":14,"I":8,"J":14,"K":14,"L":28}.items(): ws.column_dimensions[c].width=w
        ws.row_dimensions[1].height=30;ws.merge_cells("A1:H1");ws.merge_cells("I1:L1")
        for c in "ABCDEFGHIJKL": ws[f"{c}1"].fill=dl(L);ws[f"{c}1"].border=kn(L)
        ws["A1"].value=plan_adi;ws["A1"].font=Font(bold=True,size=13,color=B,name="Arial");ws["A1"].alignment=Alignment(horizontal="left",vertical="center")
        ws["I1"].value="Tarih: "+tarih;ws["I1"].font=Font(size=10,color=B,name="Arial");ws["I1"].alignment=Alignment(horizontal="right",vertical="center")
        ws.row_dimensions[2].height=6
        basliklar=["S.No","Yüklenici Firma","Sipariş No","Mal Grubu","Malzeme Tanımı","Planlanan","Gönderilen","Kalan","Birim","Durum","Tır Plaka","Devir"]
        ws.row_dimensions[3].height=22
        for i,b in enumerate(basliklar,1): h=ws.cell(row=3,column=i,value=b);h.font=Font(bold=True,size=10,color=B,name="Arial");h.fill=dl(L);h.alignment=Alignment(horizontal="center",vertical="center");h.border=kn(L)        
        

        for idx,k in enumerate(kalemler):
            r=4+idx;bg=B if idx%2==0 else G;ws.row_dimensions[r].height=None
            planlanan = k.get("planlanan_miktar", 0) or 0
            gercek_gonderilen = k.get("gonderilen_miktar", 0) or 0
            gonderilen_gosterim = min(gercek_gonderilen, planlanan)
            fazla_gonderim = max(gercek_gonderilen - planlanan, 0)
            kalan = max(planlanan - gonderilen_gosterim, 0)

            # Malzeme'den devir bilgisi temizle
            malzeme = k.get("malzeme_tanimi","") or ""
            malzeme = malzeme.replace("( planından devir )", "").replace("(planından devir)", "").replace("planından devir", "").replace("( )", "").replace("()", "").strip()
            devreden_plan = k.get("devreden_plan_adi","") or ""
            if devreden_plan and devreden_plan in malzeme:
                malzeme = malzeme.replace(devreden_plan, "").strip()
            
            vals = [
    idx+1,
    k.get("yuklenici_firma",""),
    k.get("siparis_no",""),
    k.get("mal_grubu",""),
    malzeme,
    planlanan,
    gonderilen_gosterim,
    kalan,
    k.get("birim","KG"),
    k.get("durum",""),
    k.get("tir_plaka",""),
    devreden_plan if devreden_plan else "-"
]


            for col,val in enumerate(vals,1): h=ws.cell(row=r,column=col,value=val);h.font=Font(size=9,name="Arial");h.fill=dl(bg);h.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);h.border=kn()
        sr=4+len(kalemler)+1;ws.row_dimensions[sr].height=6
        tp_ham = sum((k.get("planlanan_miktar",0) or 0) for k in kalemler)
        tg = sum(min((k.get("gonderilen_miktar",0) or 0), (k.get("planlanan_miktar",0) or 0)) for k in kalemler)
        tf = sum(max((k.get("gonderilen_miktar",0) or 0) - (k.get("planlanan_miktar",0) or 0), 0) for k in kalemler)
        td = sum((k.get("planlanan_miktar",0) or 0) for k in kalemler if k.get("durum") == "Devreden")
        tp = max(tp_ham - td, 0)
        tk = max(tp - tg, 0)
        for off,et,dg in [(1,"Toplam Planlanan",tp),(2,"Toplam Gönderilen",tg),(3,"Toplam Fazla Gönderim",tf),(4,"Toplam Kalan",tk)]:
            r=sr+off;ws.row_dimensions[r].height=22;ws.merge_cells(f"A{r}:J{r}")
            for c in range(1,13): h=ws.cell(row=r,column=c);h.fill=dl(L);h.border=kn(L)
            ws[f"A{r}"].value=et;ws[f"A{r}"].font=Font(bold=True,size=11,color=B,name="Arial");ws[f"A{r}"].alignment=Alignment(horizontal="left",vertical="center")
            ws[f"K{r}"].value=f"{dg:,.0f} KG";ws[f"K{r}"].font=Font(bold=True,size=11,color=B,name="Arial");ws[f"K{r}"].alignment=Alignment(horizontal="right",vertical="center")
        ra=sr+5;ws[f"A{ra}"].value="";ws[f"A{ra}"].font=Font(size=9,color="888888",name="Arial")
        ws[f"L{ra}"].value="Version 3.2";ws[f"L{ra}"].font=Font(size=9,color="888888",name="Arial");ws[f"L{ra}"].alignment=Alignment(horizontal="right")
        
        
        from openpyxl.utils import get_column_letter

# --- OTOMATİK SÜTUN GENİŞLİĞİ ---
        for i, column in enumerate(ws.columns, 1):

            max_length = 0
            column_letter = get_column_letter(i)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

    # normal genişlik
            adjusted_width = max_length + 2

    # kolon bazlı limitler
            if column_letter == "A":
                adjusted_width = min(adjusted_width, 6)

            elif column_letter == "B":
                adjusted_width = min(adjusted_width, 20)

            elif column_letter == "C":
                adjusted_width = min(adjusted_width, 35)

            elif column_letter == "D":
                adjusted_width = min(adjusted_width, 16)

            elif column_letter == "E":
                adjusted_width = min(adjusted_width, 100)  # MALZEME TANIMI GENİŞ

            else:
                adjusted_width = min(adjusted_width, 14)

            ws.column_dimensions[column_letter].width = adjusted_width
        


        klasor=os.path.join(get_export_path(),"exports","excel");os.makedirs(klasor,exist_ok=True)
        ad="sevkiyat_"+datetime.now().strftime("%Y%m%d_%H%M")+".xlsx";yol=os.path.join(klasor,ad);wb.save(yol)
        try: os.startfile(klasor)
        except: pass
        return jsonify({"durum":"ok","yol":yol})
    except Exception as e: return jsonify({"durum":"hata","mesaj":str(e)}), 500

    

# ---- ŞABLON İNDİR ----
@app.route("/api/sevkiyat/sablon-indir", methods=["POST"])
def sevkiyat_sablon_indir():
    if not giris_yapildi_mi(): return jsonify({"durum":"hata","mesaj":"Yetkisiz"}), 401
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb=Workbook();ws=wb.active;ws.title="Sevkiyat"
        L="012946";B="FFFFFF"
        def kn(): s=Side(style="thin",color="DDDDDD"); return Border(left=s,right=s,top=s,bottom=s)
        def dl(h): return PatternFill("solid",fgColor=h)
        for c,w in {"A":25,"B":40,"C":20,"D":50,"E":20,"F":10,"G":12}.items(): ws.column_dimensions[c].width=w
        for col,b in enumerate(["YÜKLENİCİ FİRMA","SİPARİŞ NO","MAL GRUBU","MALZEME TANIMI","MİKTAR","BİRİM","DURUM"],1):
            h=ws.cell(row=1,column=col,value=b);h.font=Font(bold=True,size=11,color=B,name="Arial");h.fill=dl(L);h.alignment=Alignment(horizontal="center",vertical="center");h.border=kn()
        for i,row in enumerate([["DARICIOĞLU","AKU-90VFA-M-TCT-00573.22","PROFİL","UPE 200",500,"KG",""],["ARS ÇELİK","AKU-90VFA-M-ICN-ICI-EM-01135.24","SAC","SAC 10X1500X6000 S275J2 EN10025",30000,"KG",""],["SÜMBÜL ÇELİK","AKU-90VFA-M-ICN-ICI-EM-00732.23","DEMİR DÜZ","DEMİZ DÜZ Ø6 L=6000 S235JR EN10025",120,"KG",""]]):
            for col,val in enumerate(row,1): h=ws.cell(row=2+i,column=col,value=val);h.font=Font(size=10,name="Arial");h.alignment=Alignment(horizontal="center",vertical="center");h.border=kn()
        klasor=os.path.join(get_export_path(),"import");os.makedirs(klasor,exist_ok=True)
        yol=os.path.join(klasor,"sevkiyat_sablon.xlsx");wb.save(yol)
        return jsonify({"durum":"ok","yol":yol})
    except Exception as e: return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/import-klasor-ac", methods=["POST"])
def import_klasor_ac():
    if not giris_yapildi_mi(): return jsonify({"durum":"hata"}), 401
    try: klasor=os.path.join(get_export_path(),"import");os.makedirs(klasor,exist_ok=True);os.startfile(klasor);return jsonify({"durum":"ok"})
    except Exception as e: return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/sevkiyat/import-oku", methods=["POST"])
@yetki_gerekli("import_excel")
def sevkiyat_import_oku():
    if not giris_yapildi_mi(): return jsonify({"durum":"hata"}), 401
    try:
        from openpyxl import load_workbook
        klasor=os.path.join(get_export_path(),"import")
        if not os.path.exists(klasor): return jsonify({"durum":"hata","mesaj":"import klasörü bulunamadı."}), 404
        xlsx_dosyalar=[f for f in os.listdir(klasor) if f.endswith((".xlsx",".xls")) and not f.startswith("~$")]
        if not xlsx_dosyalar: return jsonify({"durum":"hata","mesaj":"import klasöründe Excel dosyası bulunamadı."}), 404
        xlsx_dosyalar.sort(key=lambda f: os.path.getmtime(os.path.join(klasor,f)), reverse=True)
        dosya_adi=xlsx_dosyalar[0]; dosya_yol=os.path.join(klasor,dosya_adi)
        wb=load_workbook(dosya_yol,read_only=True,data_only=True); ws=wb.active
        kalemler=[]
        for row in ws.iter_rows(min_row=2,values_only=True):
            if not row or len(row)<4: continue
            yuklenici=str(row[0] or "").strip(); siparis=str(row[1] or "").strip()
            mal_grubu=str(row[2] or "").strip(); malzeme=str(row[3] or "").strip()
            miktar=0
            try: miktar=float(row[4])
            except: pass
            birim_raw=str(row[5] or "KG").strip().upper() if len(row)>5 else "KG"
            if not malzeme or miktar<=0: continue
            birim="KG"
            if birim_raw in ("ADET","ADT","AD"): birim="Adet"
            elif birim_raw in ("METRE","MT","M"): birim="Metre"
            kalemler.append({"yuklenici_firma":yuklenici,"siparis_no":siparis,"mal_grubu":mal_grubu,"malzeme_tanimi":malzeme,"miktar":miktar,"birim":birim})
        wb.close()
        if not kalemler: return jsonify({"durum":"hata","mesaj":f"'{dosya_adi}' dosyasında geçerli veri bulunamadı."}), 400
        return jsonify({"durum":"ok","dosya":dosya_adi,"kalemler":kalemler})
    except Exception as e: return jsonify({"durum":"hata","mesaj":str(e)}), 500

# ---- MEVCUT RAPOR/DOSYA API ----

@app.route("/api/dosya-ac", methods=["POST"])
def dosya_ac():
    if not giris_yapildi_mi(): return jsonify({"durum":"hata"}), 401
    try: yol=request.json.get("yol");os.startfile(yol) if os.path.exists(yol) else None;return jsonify({"durum":"ok"})
    except Exception as e: return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/dosya-sil", methods=["POST"])
@yetki_gerekli("dosya_sil")
def dosya_sil():
    try:
        yol = request.json.get("yol")
        os.remove(yol) if os.path.exists(yol) else None
        return jsonify({"durum":"ok"})
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route("/api/klasor-ac", methods=["POST"])
def klasor_ac():
    if not giris_yapildi_mi(): return jsonify({"durum":"hata"}), 401
    try: klasor=os.path.join(get_export_path(),"exports",request.json.get("tip"));os.makedirs(klasor,exist_ok=True);os.startfile(klasor);return jsonify({"durum":"ok"})
    except Exception as e: return jsonify({"durum":"hata","mesaj":str(e)}), 500

@app.route('/mutabakat')
def mutabakat():
    return render_template('mutabakat.html')

@app.route("/kaydet/pdf", methods=["POST"])
def kaydet_pdf():
    if not giris_yapildi_mi(): return jsonify({"durum":"hata","mesaj":"Yetkisiz"}), 401
    try:
        d=request.json;klasor=os.path.join(get_export_path(),"exports","pdf");os.makedirs(klasor,exist_ok=True)
        yol=os.path.join(klasor,d["ad"])
        with open(yol,"wb") as f: f.write(base64.b64decode(d["dosya"]))
        try: os.startfile(klasor)
        except: pass
        return jsonify({"durum":"ok","yol":yol})
    except Exception as e: return jsonify({"durum":"hata","mesaj":str(e)}), 500


# ─────────────────────────────────────────────────────────────
#  app.py'deki eski @app.route("/kaydet/excel") fonksiyonunu
#  tamamen SİL, yerine bunu YAPIŞTIR
# ─────────────────────────────────────────────────────────────

@app.route("/kaydet/excel", methods=["POST"])
def kaydet_excel():
    if not giris_yapildi_mi():
        return jsonify({"durum": "hata", "mesaj": "Yetkisiz"}), 401
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        d = request.json
        satirlar  = d["satirlar"]
        toplamKg  = d["toplamKg"]
        toplamTon = d["toplamTon"]
        tarih     = datetime.now().strftime("%d.%m.%Y")

        wb = Workbook()
        ws = wb.active
        ws.title = "Rapor"
        ws.sheet_view.showGridLines = False

        L = "012946"; B = "FFFFFF"; G = "F4F6F8"; S = "DDDDDD"

        def kn(r=S):
            s = Side(style="thin", color=r)
            return Border(left=s, right=s, top=s, bottom=s)

        def dl(h):
            return PatternFill("solid", fgColor=h)

        # 15 kolon: A-O
        basliklar = [
            "Lokasyon", "Sayfa No", "SAP Kodu", "Malzeme", "Kalite", "EN Standardı",
            "Kalınlık / Çap\n(mm)", "Genişlik / Kenar A\n(mm)",
            "Yükseklik / Kenar B\n(mm)", "Et Kalınlığı\n(mm)", "Uzunluk\n(mm)",
            "Ölçü", "Birim Kg", "Adet", "Toplam Kg"
        ]

        genislikler = {
            "A":14, "B":10, "C":16, "D":16, "E":12, "F":14,
            "G":14, "H":16, "I":16, "J":14, "K":12,
            "L":28, "M":12, "N":8, "O":14
        }
        for c, w in genislikler.items():
            ws.column_dimensions[c].width = w

        # Başlık
        ws.row_dimensions[1].height = 30
        ws.merge_cells("A1:K1")
        ws.merge_cells("L1:O1")
        for c in "ABCDEFGHIJKLMNO":
            ws[f"{c}1"].fill = dl(L)
            ws[f"{c}1"].border = kn(L)
        ws["A1"].value = "Malzeme Ağırlık Hesaplama Raporu"
        ws["A1"].font = Font(bold=True, size=13, color=B, name="Arial")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws["L1"].value = "Tarih: " + tarih
        ws["L1"].font = Font(size=10, color=B, name="Arial")
        ws["L1"].alignment = Alignment(horizontal="right", vertical="center")

        # Kolon başlıkları
        ws.row_dimensions[2].height = 6
        ws.row_dimensions[3].height = 32
        for i, bas in enumerate(basliklar, 1):
            h = ws.cell(row=3, column=i, value=bas)
            h.font = Font(bold=True, size=9, color=B, name="Arial")
            h.fill = dl(L)
            h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            h.border = kn(L)

        # Veri satırları
        for idx, s in enumerate(satirlar):
            r = 4 + idx
            bg = B if idx % 2 == 0 else G

            kalinlik_cap = s.get("profil_tipi") or s.get("kalinlik") or s.get("cap") or ""

            vals = [
                s.get("lokasyon", ""),
                s.get("sayfaNo", ""),
                s.get("sapKodu", ""),
                s.get("malzeme", ""),
                s.get("kalite", ""),
                s.get("enStd", ""),
                kalinlik_cap,
                s.get("genislik", ""),
                s.get("yukseklik", ""),
                s.get("et", ""),
                s.get("uzunluk", ""),
                s.get("olcu", ""),
                s.get("birimKg", 0),
                s.get("adet", 0),
                s.get("kg", 0),
            ]

            for col, val in enumerate(vals, 1):
                h = ws.cell(row=r, column=col)

                if col >= 13:  # birimKg, adet, kg
                    try:
                        h.value = float(val) if val else 0
                        h.number_format = '#,##0.00'
                    except:
                        h.value = val
                elif 7 <= col <= 11:  # boyut kolonları
                    try:
                        v = float(val) if val else ""
                        h.value = v if v else ""
                        if v:
                            h.number_format = '#,##0.##'
                    except:
                        h.value = val
                else:
                    h.value = val if val else ""

                h.font = Font(size=9, name="Arial")
                h.fill = dl(bg)
                h.alignment = Alignment(horizontal="center", vertical="center")
                h.border = kn()

        # Toplamlar
        son = 3 + len(satirlar)
        ws.row_dimensions[son + 1].height = 8

        for off, et, dg in [(2, "Toplam Kg", toplamKg), (3, "Toplam Ton", toplamTon)]:
            r = son + off
            ws.row_dimensions[r].height = 22
            ws.merge_cells(f"A{r}:N{r}")
            for c in range(1, 16):
                h = ws.cell(row=r, column=c)
                h.fill = dl(L)
                h.border = kn(L)
            ws[f"A{r}"].value = et
            ws[f"A{r}"].font = Font(bold=True, size=11, color=B, name="Arial")
            ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
            ws[f"O{r}"].value = dg
            ws[f"O{r}"].font = Font(bold=True, size=11, color=B, name="Arial")
            ws[f"O{r}"].alignment = Alignment(horizontal="right", vertical="center")

        ra = son + 5
        ws[f"O{ra}"].value = "Version 3.2"
        ws[f"O{ra}"].font = Font(size=9, color="888888", name="Arial")
        ws[f"O{ra}"].alignment = Alignment(horizontal="right")

        klasor = os.path.join(get_export_path(), "exports", "excel")
        os.makedirs(klasor, exist_ok=True)
        yol = os.path.join(klasor, "agirlik_" + datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx")
        wb.save(yol)
        try:
            os.startfile(klasor)
        except:
            pass
        return jsonify({"durum": "ok", "yol": yol})

    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    
@app.route("/api/mutabakat/export", methods=["POST"])
def mutabakat_export():
    if not giris_yapildi_mi():
        return jsonify({"durum": "hata", "mesaj": "Yetkisiz"}), 401
    try:
        d = request.json
        fname = os.path.basename(d["filename"])          # güvenlik
        klasor = os.path.join(get_export_path(), "exports")
        os.makedirs(klasor, exist_ok=True)
        yol = os.path.join(klasor, fname)
        with open(yol, "wb") as f:
            f.write(base64.b64decode(d["data"]))
        try:
            os.startfile(klasor)                          # bitince klasörü aç
        except:
            pass
        return jsonify({"ok": True, "yol": yol})          # HTML j.ok bekliyor
    except Exception as e:
        return jsonify({"ok": False, "mesaj": str(e)}), 500
    

@app.route("/malzeme-kontrol")
def malzeme_kontrol_sayfa():
    return render_template("malzeme-kontrol.html")

@app.route("/rickroll")
def rickroll_sayfasi():
    return render_template("rickroll.html")

# ---- SERVER ----
def start_server(): app.run(host="127.0.0.1",port=5000)
def open_login(): time.sleep(5);webview.windows[0].load_url("http://127.0.0.1:5000/login")

if __name__=="__main__":
    veritabani_olustur()
    mal_grubu_kategorileri_olustur()
    init_agirlik_db()
    init_malzeme_db()
    mb52_init_db()  
    e=get_export_path()
    os.makedirs(os.path.join(e,"exports","pdf"),exist_ok=True)
    os.makedirs(os.path.join(e,"exports","excel"),exist_ok=True)
    os.makedirs(os.path.join(e,"exports","html"),exist_ok=True)
    os.makedirs(os.path.join(e,"import"),exist_ok=True)
    os.makedirs(os.path.join(e,"assets","fonts"),exist_ok=True)
    server=threading.Thread(target=start_server);server.daemon=True;server.start()
    webview.create_window("Warehouse Data Management","http://127.0.0.1:5000/splash",width=1200,height=800)
    threading.Thread(target=open_login).start();webview.start()