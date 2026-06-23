# agirlik_db.py v2.2 — Lokasyon, Sayfa No, SAP Kodu, Kalite, EN + boyut alanlari
# v2.2: Database lock sorunu giderildi — busy_timeout, synchronous=NORMAL, try/finally

import sqlite3, os, socket
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file

agirlik_bp = Blueprint("agirlik", __name__)

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "agirlik_hesaplama.db")


def get_agirlik_db():
    """
    Lock-safe DB bağlantısı.
    - timeout=30: connect aşamasında 30 sn bekle
    - busy_timeout=30000: yazma kilidi varsa 30 sn bekle (default 0)
    - WAL: okuma yazmayı engellemez
    - synchronous=NORMAL: WAL ile güvenli, daha hızlı commit
    """
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=30000")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA foreign_keys=ON")
    except sqlite3.OperationalError:
        # DB başka biri tarafından locked olsa bile bağlantı sağlanır
        pass
    return conn


def init_agirlik_db():
    conn = get_agirlik_db()
    try:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS hesaplama_gruplari (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                aciklama    TEXT    DEFAULT '',
                toplam_kg   REAL    DEFAULT 0,
                toplam_ton  REAL    DEFAULT 0,
                kalem_sayisi INTEGER DEFAULT 0,
                tarih       TEXT    NOT NULL,
                kullanici   TEXT    DEFAULT '',
                pc_adi      TEXT    DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS hesaplama_satirlari (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                grup_id     INTEGER NOT NULL,
                lokasyon    TEXT    DEFAULT '',
                sayfa_no    TEXT    DEFAULT '',
                sap_kodu    TEXT    DEFAULT '',
                malzeme     TEXT    NOT NULL,
                kalite      TEXT    DEFAULT '',
                en_std      TEXT    DEFAULT '',
                olcu        TEXT    NOT NULL,
                birim_kg    REAL    NOT NULL,
                adet        REAL    NOT NULL,
                toplam_kg   REAL    NOT NULL,
                cap         TEXT    DEFAULT '',
                kalinlik    TEXT    DEFAULT '',
                genislik    TEXT    DEFAULT '',
                yukseklik   TEXT    DEFAULT '',
                et          TEXT    DEFAULT '',
                uzunluk     TEXT    DEFAULT '',
                profil_tipi TEXT    DEFAULT '',
                FOREIGN KEY (grup_id) REFERENCES hesaplama_gruplari(id) ON DELETE CASCADE
            );
        """)
        # Mevcut DB icin yeni kolonlari ekle
        for kolon, tip in [
            ("lokasyon","TEXT DEFAULT ''"), ("sayfa_no","TEXT DEFAULT ''"),
            ("sira_no","INTEGER"),
            ("sap_kodu","TEXT DEFAULT ''"), ("kalite","TEXT DEFAULT ''"),
            ("en_std","TEXT DEFAULT ''"), ("cap","TEXT DEFAULT ''"),
            ("kalinlik","TEXT DEFAULT ''"), ("genislik","TEXT DEFAULT ''"),
            ("yukseklik","TEXT DEFAULT ''"), ("et","TEXT DEFAULT ''"),
            ("uzunluk","TEXT DEFAULT ''"), ("profil_tipi","TEXT DEFAULT ''"),
        ]:
            try: conn.execute(f"ALTER TABLE hesaplama_satirlari ADD COLUMN {kolon} {tip}")
            except: pass
        conn.commit()
    finally:
        conn.close()


def _satir_kolonlari(conn):
    return [r[1] for r in conn.execute("PRAGMA table_info(hesaplama_satirlari)").fetchall()]


def _safe(row, col, kolonlar, default=""):
    if col in kolonlar:
        v = row[col]
        return v if v is not None else default
    return default


# ── Kaydet ──
@agirlik_bp.route("/api/agirlik/kaydet", methods=["POST"])
def agirlik_kaydet():
    conn = None
    try:
        data = request.get_json()
        satirlar = data.get("satirlar", [])
        if not satirlar:
            return jsonify({"durum":"hata","mesaj":"Kaydedilecek satir yok"}), 400

        conn = get_agirlik_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO hesaplama_gruplari (aciklama,toplam_kg,toplam_ton,kalem_sayisi,tarih,kullanici,pc_adi)
            VALUES (?,?,?,?,?,?,?)
        """, (
            data.get("aciklama",""),
            data.get("toplamKg",0),
            data.get("toplamTon",0),
            len(satirlar),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            data.get("kullanici",""),
            socket.gethostname()
        ))
        gid = cur.lastrowid

        # Toplu insert — tek transaction, çok daha hızlı + lock süresini kısaltır
        cur.executemany("""
            INSERT INTO hesaplama_satirlari
            (grup_id, lokasyon, sayfa_no, sira_no, sap_kodu, malzeme, kalite, en_std,
             olcu, birim_kg, adet, toplam_kg,
             cap, kalinlik, genislik, yukseklik, et, uzunluk, profil_tipi)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, [(
            gid,
            s.get("lokasyon",""),
            s.get("sayfaNo",""),
            (int(s["siraNo"]) if s.get("siraNo") not in (None, "", "null") and str(s.get("siraNo")).strip().isdigit() else None),
            s.get("sapKodu",""),
            s.get("malzeme",""),
            s.get("kalite",""),
            s.get("enStd",""),
            s.get("olcu",""),
            s.get("birimKg",0),
            s.get("adet",0),
            s.get("kg",0),
            s.get("cap",""),
            s.get("kalinlik",""),
            s.get("genislik",""),
            s.get("yukseklik",""),
            s.get("et",""),
            s.get("uzunluk",""),
            s.get("profil_tipi",""),
        ) for s in satirlar])

        conn.commit()
        return jsonify({"durum":"ok","mesaj":f"{len(satirlar)} kalem kaydedildi","grup_id":gid})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({"durum":"hata","mesaj":str(e)}), 500
    finally:
        if conn:
            try: conn.close()
            except: pass


# ── Gecmis ──
@agirlik_bp.route("/api/agirlik/gecmis", methods=["GET"])
def agirlik_gecmis():
    conn = None
    try:
        conn = get_agirlik_db()
        rows = conn.execute("SELECT * FROM hesaplama_gruplari ORDER BY id DESC LIMIT 100").fetchall()
        return jsonify([{
            "id":g["id"], "aciklama":g["aciklama"], "toplam_kg":g["toplam_kg"],
            "toplam_ton":g["toplam_ton"], "kalem_sayisi":g["kalem_sayisi"],
            "tarih":g["tarih"], "kullanici":g["kullanici"], "pc_adi":g["pc_adi"]
        } for g in rows])
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500
    finally:
        if conn:
            try: conn.close()
            except: pass


# ── Detay ──
@agirlik_bp.route("/api/agirlik/detay/<int:grup_id>", methods=["GET"])
def agirlik_detay(grup_id):
    conn = None
    try:
        conn = get_agirlik_db()
        g = conn.execute("SELECT * FROM hesaplama_gruplari WHERE id=?", (grup_id,)).fetchone()
        if not g:
            return jsonify({"durum":"hata","mesaj":"Grup bulunamadi"}), 404

        kolonlar = _satir_kolonlari(conn)
        satirlar = conn.execute("SELECT * FROM hesaplama_satirlari WHERE grup_id=? ORDER BY id", (grup_id,)).fetchall()

        result = {
            "id":g["id"], "aciklama":g["aciklama"], "toplam_kg":g["toplam_kg"],
            "toplam_ton":g["toplam_ton"], "kalem_sayisi":g["kalem_sayisi"],
            "tarih":g["tarih"], "kullanici":g["kullanici"], "pc_adi":g["pc_adi"],
            "satirlar": []
        }

        for s in satirlar:
            result["satirlar"].append({
                "id":          s["id"],
                "lokasyon":    _safe(s, "lokasyon", kolonlar),
                "sayfaNo":     _safe(s, "sayfa_no", kolonlar),
                "siraNo":      _safe(s, "sira_no", kolonlar, None),
                "sapKodu":     _safe(s, "sap_kodu", kolonlar),
                "malzeme":     s["malzeme"],
                "kalite":      _safe(s, "kalite", kolonlar),
                "enStd":       _safe(s, "en_std", kolonlar),
                "olcu":        s["olcu"],
                "birimKg":     s["birim_kg"],
                "adet":        s["adet"],
                "kg":          s["toplam_kg"],
                "cap":         _safe(s, "cap", kolonlar),
                "kalinlik":    _safe(s, "kalinlik", kolonlar),
                "genislik":    _safe(s, "genislik", kolonlar),
                "yukseklik":   _safe(s, "yukseklik", kolonlar),
                "et":          _safe(s, "et", kolonlar),
                "uzunluk":     _safe(s, "uzunluk", kolonlar),
                "profil_tipi": _safe(s, "profil_tipi", kolonlar),
            })

        return jsonify(result)
    except Exception as e:
        return jsonify({"durum":"hata","mesaj":str(e)}), 500
    finally:
        if conn:
            try: conn.close()
            except: pass


# ── Sil ──
@agirlik_bp.route("/api/agirlik/sil/<int:grup_id>", methods=["DELETE"])
def agirlik_sil(grup_id):
    conn = None
    try:
        conn = get_agirlik_db()
        conn.execute("DELETE FROM hesaplama_satirlari WHERE grup_id=?", (grup_id,))
        conn.execute("DELETE FROM hesaplama_gruplari WHERE id=?", (grup_id,))
        conn.commit()
        return jsonify({"durum":"ok","mesaj":"Silindi"})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({"durum":"hata","mesaj":str(e)}), 500
    finally:
        if conn:
            try: conn.close()
            except: pass


# ── DB Indir ──
@agirlik_bp.route("/api/agirlik/export-db", methods=["GET"])
def agirlik_export_db():
    if os.path.exists(DB_PATH):
        return send_file(DB_PATH, as_attachment=True, download_name=f"agirlik_{socket.gethostname()}.db")
    return jsonify({"durum":"hata","mesaj":"DB bulunamadi"}), 404


# ── DB Import ──
@agirlik_bp.route("/api/agirlik/import-db", methods=["POST"])
def agirlik_import_db():
    conn_main = None
    conn_imp = None
    tmp = None
    try:
        if "dosya" not in request.files:
            return jsonify({"durum":"hata","mesaj":"Dosya yok"}), 400

        f = request.files["dosya"]
        tmp = os.path.join(os.path.dirname(DB_PATH), "_import_tmp.db")
        f.save(tmp)

        conn_main = get_agirlik_db()
        conn_imp = sqlite3.connect(tmp, timeout=30)
        conn_imp.row_factory = sqlite3.Row

        imp_kolonlar = [r[1] for r in conn_imp.execute("PRAGMA table_info(hesaplama_satirlari)").fetchall()]
        gruplar = conn_imp.execute("SELECT * FROM hesaplama_gruplari ORDER BY id").fetchall()
        imported = 0

        for g in gruplar:
            cur = conn_main.cursor()
            cur.execute("""
                INSERT INTO hesaplama_gruplari (aciklama,toplam_kg,toplam_ton,kalem_sayisi,tarih,kullanici,pc_adi)
                VALUES (?,?,?,?,?,?,?)
            """, (g["aciklama"],g["toplam_kg"],g["toplam_ton"],g["kalem_sayisi"],g["tarih"],g["kullanici"],g["pc_adi"]))
            new_id = cur.lastrowid

            for s in conn_imp.execute("SELECT * FROM hesaplama_satirlari WHERE grup_id=?", (g["id"],)).fetchall():
                def sf(col, d=""): return _safe(s, col, imp_kolonlar, d)
                conn_main.execute("""
                    INSERT INTO hesaplama_satirlari
                    (grup_id,lokasyon,sayfa_no,sira_no,sap_kodu,malzeme,kalite,en_std,olcu,birim_kg,adet,toplam_kg,cap,kalinlik,genislik,yukseklik,et,uzunluk,profil_tipi)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    new_id, sf("lokasyon"), sf("sayfa_no"), sf("sira_no", None), sf("sap_kodu"),
                    s["malzeme"], sf("kalite"), sf("en_std"),
                    s["olcu"], s["birim_kg"], s["adet"], s["toplam_kg"],
                    sf("cap"), sf("kalinlik"), sf("genislik"),
                    sf("yukseklik"), sf("et"), sf("uzunluk"), sf("profil_tipi")
                ))
            imported += 1

        conn_main.commit()
        return jsonify({"durum":"ok","mesaj":f"{imported} grup import edildi"})
    except Exception as e:
        if conn_main:
            try: conn_main.rollback()
            except: pass
        return jsonify({"durum":"hata","mesaj":str(e)}), 500
    finally:
        if conn_main:
            try: conn_main.close()
            except: pass
        if conn_imp:
            try: conn_imp.close()
            except: pass
        if tmp and os.path.exists(tmp):
            try: os.remove(tmp)
            except: pass

# ── TOPLU EXCEL EXPORT ──
@agirlik_bp.route("/api/agirlik/toplu-excel", methods=["POST"])
def agirlik_toplu_excel():
    """
    Secili grup_id'leri Excel'e dokerek sunucu klasorune kaydeder.
    format: "tek" → hepsi tek sheet'te ayrac ile birlesik
            "ayri" → her hesaplama ayri sheet
    """
    conn = None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        data = request.get_json() or {}
        grup_ids = data.get("grup_ids", [])
        format_tipi = (data.get("format") or "tek").lower()  # "tek" veya "ayri"

        if not grup_ids:
            return jsonify({"durum": "hata", "mesaj": "Grup secilmedi"}), 400

        conn = get_agirlik_db()
        kolonlar = _satir_kolonlari(conn)

        # ─── Ortak stiller ───
        header_font  = Font(bold=True, color="FFFFFF", size=10)
        header_fill  = PatternFill("solid", fgColor="012946")
        ayrac_font   = Font(bold=True, color="012946", size=11)
        ayrac_fill   = PatternFill("solid", fgColor="DBEAFE")
        info_font    = Font(bold=True, size=12, color="012946")
        info_fill    = PatternFill("solid", fgColor="F1F5F9")
        toplam_font  = Font(bold=True, size=11)
        toplam_fill  = PatternFill("solid", fgColor="FEF3C7")
        genel_font   = Font(bold=True, size=13, color="FFFFFF")
        genel_fill   = PatternFill("solid", fgColor="012946")
        thin         = Side(border_style="thin", color="CCCCCC")
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)
        align_center = Alignment(horizontal="center", vertical="center")
        align_left   = Alignment(horizontal="left",   vertical="center")
        align_right  = Alignment(horizontal="right",  vertical="center")

        basliklar = ["Lokasyon", "Sayfa", "Sira", "SAP Kodu", "Malzeme",
                     "Kalite", "Standart", "Olcu", "Birim Kg", "Adet", "Toplam Kg"]
        genislikler = [16, 8, 7, 16, 18, 18, 14, 30, 12, 8, 14]

        def tr_format(v):
            try:
                return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except:
                return "0,00"

        def baslik_satiri_yaz(ws, satir):
            """Sutun basliklarini yazar, satiri 1 ileri tasir."""
            for col_idx, b in enumerate(basliklar, start=1):
                c = ws.cell(row=satir, column=col_idx, value=b)
                c.font = header_font
                c.fill = header_fill
                c.alignment = align_center
                c.border = border
            return satir + 1

        def satirlari_yaz(ws, satir, grup_satirlari):
            """Bir grubun satirlarini yazar, son satiri + grup_toplam doner."""
            grup_toplam = 0
            for s in grup_satirlari:
                values = [
                    _safe(s, "lokasyon", kolonlar),
                    _safe(s, "sayfa_no", kolonlar),
                    _safe(s, "sira_no",  kolonlar, ""),
                    _safe(s, "sap_kodu", kolonlar),
                    s["malzeme"],
                    _safe(s, "kalite", kolonlar),
                    _safe(s, "en_std", kolonlar),
                    s["olcu"],
                    s["birim_kg"],
                    s["adet"],
                    s["toplam_kg"],
                ]
                for col_idx, v in enumerate(values, start=1):
                    c = ws.cell(row=satir, column=col_idx, value=v)
                    c.border = border
                    if col_idx in (9, 10, 11):
                        c.alignment = align_right
                        if col_idx in (9, 11):
                            c.number_format = '#,##0.00'
                    else:
                        c.alignment = align_left
                grup_toplam += s["toplam_kg"] or 0
                satir += 1
            return satir, grup_toplam

        def sutun_genisliklerini_ayarla(ws):
            for i, w in enumerate(genislikler, start=1):
                ws.column_dimensions[chr(64 + i)].width = w

        def temiz_sheet_adi(metin, varsayilan):
            """Excel sheet adı: 31 char limit, özel karakter yok."""
            metin = str(metin or varsayilan).strip()
            yasak = ['\\', '/', '?', '*', '[', ']', ':']
            for y in yasak:
                metin = metin.replace(y, "")
            return (metin[:31] or varsayilan)

        wb = Workbook()
        # Default sheet'i sileceğiz, manuel oluşturacağız
        wb.remove(wb.active)

        # ──────────────────────────────────────
        # FORMAT 1: TEK SHEET — Hepsi alt alta, ayrac satırı ile
        # ──────────────────────────────────────
        if format_tipi == "tek":
            ws = wb.create_sheet("Toplu Hesaplama")
            satir = 1
            genel_toplam = 0

            for gid in grup_ids:
                grup = conn.execute("SELECT * FROM hesaplama_gruplari WHERE id=?", (gid,)).fetchone()
                if not grup:
                    continue

                # Ayrac satiri
                grup_aciklama = grup["aciklama"] or f"Hesaplama #{grup['id']}"
                ayrac_metin = "{} | {} | {} kalem | Toplam: {} kg".format(
                    grup_aciklama, grup["tarih"], grup["kalem_sayisi"], tr_format(grup["toplam_kg"])
                )
                ws.cell(row=satir, column=1, value=ayrac_metin)
                ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=len(basliklar))
                c = ws.cell(row=satir, column=1)
                c.font = ayrac_font; c.fill = ayrac_fill; c.alignment = align_left
                ws.row_dimensions[satir].height = 22
                satir += 1

                # Basliklar
                satir = baslik_satiri_yaz(ws, satir)

                # Satırlar
                grup_satirlari = conn.execute(
                    "SELECT * FROM hesaplama_satirlari WHERE grup_id=? ORDER BY id", (gid,)
                ).fetchall()
                satir, grup_toplam = satirlari_yaz(ws, satir, grup_satirlari)

                # Grup toplam satiri
                ws.cell(row=satir, column=1, value="Grup Toplami:")
                ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=10)
                sol = ws.cell(row=satir, column=1)
                sol.font = toplam_font; sol.fill = toplam_fill
                sol.alignment = align_right; sol.border = border
                tc = ws.cell(row=satir, column=11, value=grup_toplam)
                tc.font = toplam_font; tc.fill = toplam_fill
                tc.alignment = align_right; tc.number_format = '#,##0.00'; tc.border = border

                genel_toplam += grup_toplam
                satir += 2  # bos ayrac

            # GENEL TOPLAM
            ws.cell(row=satir, column=1, value="GENEL TOPLAM")
            ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=10)
            gsol = ws.cell(row=satir, column=1)
            gsol.font = genel_font; gsol.fill = genel_fill; gsol.alignment = align_right
            gc = ws.cell(row=satir, column=11, value=genel_toplam)
            gc.font = genel_font; gc.fill = genel_fill
            gc.alignment = align_right; gc.number_format = '#,##0.00'
            ws.row_dimensions[satir].height = 26

            sutun_genisliklerini_ayarla(ws)

        # ──────────────────────────────────────
        # FORMAT 2: AYRI SHEET'LER — Her hesaplama kendi sayfasında
        # ──────────────────────────────────────
        else:
            # Önce özet sayfası
            ozet = wb.create_sheet("Ozet")
            ozet.cell(row=1, column=1, value="HESAPLAMA OZET")
            ozet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            c = ozet.cell(row=1, column=1)
            c.font = genel_font; c.fill = genel_fill; c.alignment = align_center
            ozet.row_dimensions[1].height = 26

            # Özet başlıkları
            ozet_basliklar = ["Sayfa Adi", "Aciklama", "Tarih", "Toplam Kg"]
            for col_idx, b in enumerate(ozet_basliklar, start=1):
                cc = ozet.cell(row=2, column=col_idx, value=b)
                cc.font = header_font; cc.fill = header_fill
                cc.alignment = align_center; cc.border = border

            ozet_satir = 3
            ozet_genel_toplam = 0
            kullanilan_sheet_adlari = set()

            for gid in grup_ids:
                grup = conn.execute("SELECT * FROM hesaplama_gruplari WHERE id=?", (gid,)).fetchone()
                if not grup:
                    continue

                grup_aciklama = grup["aciklama"] or f"Hesaplama #{grup['id']}"

                # Benzersiz sheet adi (çakışma olursa _2, _3 ekle)
                sheet_adi_base = temiz_sheet_adi(grup_aciklama, f"Hesaplama_{grup['id']}")
                sheet_adi = sheet_adi_base
                sayac = 2
                while sheet_adi in kullanilan_sheet_adlari:
                    suffix = f"_{sayac}"
                    sheet_adi = (sheet_adi_base[:31-len(suffix)]) + suffix
                    sayac += 1
                kullanilan_sheet_adlari.add(sheet_adi)

                ws = wb.create_sheet(sheet_adi)

                # En üstte hesaplama bilgisi
                ws.cell(row=1, column=1, value=grup_aciklama)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(basliklar))
                ic = ws.cell(row=1, column=1)
                ic.font = info_font; ic.fill = info_fill; ic.alignment = align_center
                ws.row_dimensions[1].height = 24

                # Tarih + kalem sayısı
                ws.cell(row=2, column=1, value=f"Tarih: {grup['tarih']}  |  {grup['kalem_sayisi']} kalem")
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(basliklar))
                ic2 = ws.cell(row=2, column=1)
                ic2.alignment = align_center
                ic2.font = Font(italic=True, color="64748B", size=10)

                # Basliklar (satır 4)
                satir = baslik_satiri_yaz(ws, 4)

                # Satırlar
                grup_satirlari = conn.execute(
                    "SELECT * FROM hesaplama_satirlari WHERE grup_id=? ORDER BY id", (gid,)
                ).fetchall()
                satir, grup_toplam = satirlari_yaz(ws, satir, grup_satirlari)

                # Toplam satiri
                ws.cell(row=satir, column=1, value="TOPLAM:")
                ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=10)
                sol = ws.cell(row=satir, column=1)
                sol.font = genel_font; sol.fill = genel_fill
                sol.alignment = align_right
                tc = ws.cell(row=satir, column=11, value=grup_toplam)
                tc.font = genel_font; tc.fill = genel_fill
                tc.alignment = align_right; tc.number_format = '#,##0.00'
                ws.row_dimensions[satir].height = 24

                sutun_genisliklerini_ayarla(ws)

                # Özet sayfasına ekle
                ozet.cell(row=ozet_satir, column=1, value=sheet_adi).border = border
                ozet.cell(row=ozet_satir, column=2, value=grup_aciklama).border = border
                ozet.cell(row=ozet_satir, column=3, value=grup["tarih"]).border = border
                tc_ozet = ozet.cell(row=ozet_satir, column=4, value=grup_toplam)
                tc_ozet.border = border
                tc_ozet.alignment = align_right
                tc_ozet.number_format = '#,##0.00'

                ozet_genel_toplam += grup_toplam
                ozet_satir += 1

            # Özet genel toplam
            ozet.cell(row=ozet_satir, column=1, value="GENEL TOPLAM")
            ozet.merge_cells(start_row=ozet_satir, start_column=1, end_row=ozet_satir, end_column=3)
            gs = ozet.cell(row=ozet_satir, column=1)
            gs.font = genel_font; gs.fill = genel_fill; gs.alignment = align_right
            gt = ozet.cell(row=ozet_satir, column=4, value=ozet_genel_toplam)
            gt.font = genel_font; gt.fill = genel_fill
            gt.alignment = align_right; gt.number_format = '#,##0.00'
            ozet.row_dimensions[ozet_satir].height = 26

            # Ozet sütun genişlikleri
            ozet.column_dimensions["A"].width = 28
            ozet.column_dimensions["B"].width = 40
            ozet.column_dimensions["C"].width = 20
            ozet.column_dimensions["D"].width = 16

            # Özet sayfasını en başa al
            wb.move_sheet(ozet, offset=-len(wb.sheetnames) + 1)

        # ─── KAYDET ───
        klasor = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports", "sayim")
        os.makedirs(klasor, exist_ok=True)
        suffix = "ayri_sayfa" if format_tipi == "ayri" else "birlesik"
        dosya_adi = "toplu_hesaplama_{}_{}.xlsx".format(
            datetime.now().strftime("%Y%m%d_%H%M%S"), suffix
        )
        dosya_yolu = os.path.join(klasor, dosya_adi)
        wb.save(dosya_yolu)

        return jsonify({
            "durum": "ok",
            "mesaj": f"{len(grup_ids)} hesaplama kaydedildi",
            "dosya_yolu": dosya_yolu,
            "dosya_adi": dosya_adi,
            "format": format_tipi
        })

    except Exception as e:
        return jsonify({"durum": "hata", "mesaj": str(e)}), 500
    finally:
        if conn:
            try: conn.close()
            except: pass